# -*- coding: utf-8 -*-
"""
装潢 Excel 原生透视表自动化工具
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
工作流程：
  1. 读取 工厂清单 7.XLSX（Plant→区域+分公司 映射）
  2. openpyxl 增强 EXPORT1 2 1.xlsx：在"订单净值"右侧插入"区域"和"分公司"列
  3. 按 工厂↔Plant 匹配填值，保留原 Sheet 全部格式
  4. pandas 筛选 物料∈{1000027307, 1000027308}
  5. COM 生成原生透视表 → Sheet "装潢透视表"（位于 Sheet1 右侧）

格式保留：
  → Sheet1 原有列宽、行高、填充色、字体、边框、合并单元格等均完整保留
  → 仅新增两列（区域、分公司，默认格式）和一个透视表 Sheet

生成效果：
  → 带有折叠/展开层级按钮的原生 Excel 数据透视表
  → 行层级：区域 → 分公司 → 供应商名称 → 采购凭证
  → 值：订单净值（求和）
  → 位置：紧邻 "装潢" Sheet 右侧
"""

import os
import sys
import time
import traceback
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
import ttkbootstrap as ttk
from download_period import monthly_download_dir


# ═══════════════════ 配置区 ═══════════════════

# 脚本所在目录（数据文件也放这里，替换同名文件即可更新）
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR = os.path.join(SCRIPT_DIR, "assets")
TKE_LOGO_PATH = os.path.join(ASSETS_DIR, "tke-move-beyond.png")

# 主数据文件
EXCEL_FILE = os.path.join(SCRIPT_DIR, "采购记录0701.xlsx")
DOWNLOADS_ROOT_DIR = os.path.join(SCRIPT_DIR, "downloads")
DOWNLOAD_DIR = monthly_download_dir(DOWNLOADS_ROOT_DIR)

# 工厂清单（Plant → 区域 + 分公司 映射表，仅装潢品类需要）
FACTORY_FILE = os.path.join(SCRIPT_DIR, "工厂清单 7.XLSX")

# 通用 Sheet 名称
SOURCE_SHEET = "Sheet1"                  # 数据源（主数据所在 Sheet）
WORKBENCH_THEME = "flatly"

# 透视表结构（所有品类通用）
ROW_FIELDS = ["区域", "分公司", "供应商名称", "采购凭证"]
VALUE_FIELD = "订单净值"
VALUE_NAME = "求和项:订单净值"

# 校验所需列
CHECK_COLS = ["工厂", "物料", "供应商名称", "采购凭证", "订单净值"]

# 透视表右侧扩展列（预留，后续手动填入，所有品类通用）
EXTRA_COLUMNS = [
    "E2E项目名",
    "WBS",
    "E2E订单数量",
    "项目总金额(审批)",
    "订单差异(单独计算)",
    "支持文件名(FLD 审批)",
    "Price",
    "PlanCost",
    "总saving",
    "订单是否下完=price-订单净值",
]

# ═══════════════════ 品类配置 ═══════════════════

CATEGORIES = {
    "装潢": {
        "label": "装潢",
        "filter_materials": [1000027307, 1000027308],
        "data_sheet": "装潢数据",
        "target_sheet": "装潢透视表",
        "pivot_table_name": "装潢透视表",
        "content_filter": "装潢",                    # PM Tracking Content 列筛选关键词
        "insert_cols_after_order": [],               # 区域/分公司已由独立按钮处理
        "insert_col_at_end": None,                   # 在 Sheet1 末尾插入的列（None=不插入）
    },
    "外包板": {
        "label": "外包板",
        "filter_materials": [1000027309],
        "data_sheet": "外包板数据",
        "target_sheet": "外包板透视表",
        "pivot_table_name": "外包板透视表",
        "content_filter": "外包板",
        "insert_cols_after_order": [],
        "insert_col_at_end": None,
    },
    "线槽": {
        "label": "线槽",
        "filter_materials": [1000027323],
        "data_sheet": "井道线槽数据",
        "workflow": "data_price",
        "required_source_columns": ["物料", "订单净值", "短文本", "净价", "采购订单数量"],
        "data_extra_columns": ["老价格", "新价格", "Saving"],
        "price_list_sheet": "井道线槽25.10降价",
        "price_source_columns": {
            "sap_description": 5,  # E=SAP Discription-Ner
            "old_price": 8,        # H=老价格
            "new_price": 10,       # J=新价格
        },
        "insert_cols_after_order": [],
        "insert_col_at_end": None,
    },
    "无线五方通话": {
        "label": "无线五方通话",
        "filter_materials": [8001366263, 8001366265, 8001366266, 8001366267],
        "data_sheet": "五方通话数据",
        "workflow": "data_price",
        "required_source_columns": ["物料", "订单净值", "短文本", "净价", "采购订单数量"],
        "data_extra_columns": ["老价格", "新价格", "Saving"],
        "price_list_sheet": "五方2026.05降价",
        "price_source_columns": {
            "sap_description": 5,  # E=SAP Discription-Ner
            "old_price": 8,        # H=老价格
            "new_price": 10,       # J=新价格
        },
        "insert_cols_after_order": [],
        "insert_col_at_end": None,
    },
    "井道照明": {
        "label": "井道照明",
        "filter_materials": [1000027319],
        "data_sheet": "井道照明数据",
        "target_sheet": "井道照明Saving",
        "pivot_table_name": "井道照明Saving",
        "workflow": "lighting_saving",
        "required_source_columns": ["物料", "订单净值", "短文本"],
        "pivot_row_fields": ["描述"],
        "pivot_value_field": "PO数量",
        "pivot_value_name": "求和项:PO数量",
        "insert_cols_after_order": [],
        "insert_col_at_end": None,
    },
    "空调": {
        "label": "空调",
        "filter_materials": [1000027316],
        "data_sheet": "空调数据",
        "target_sheet": "空调透视表",
        "pivot_table_name": "空调透视表",
        "content_filter": "AC空调",
        "content_filter_values": ("AC", "AC空调"),
        "content_filter_display": "AC / AC空调",
        "content_filter_exact_on_multiple_matches": True,
        "insert_cols_after_order": ["老价格", "新价格", "Saving"],
        "insert_col_at_end": "CHECK",
        "price_list_sheet": "AC 25.10降价",
    },
    "IC卡": {
        "label": "IC卡",
        "filter_materials": [1000027312],
        "data_sheet": "IC卡数据",
        "target_sheet": "IC卡透视表",
        "pivot_table_name": "IC卡透视表",
        "content_filter": "IC卡",
        "content_filter_values": ("IC卡",),
        "content_filter_display": "IC卡",
        "content_filter_exact_on_multiple_matches": True,
        "insert_cols_after_order": ["老价格", "新价格", "Saving"],
        "insert_col_at_end": "CHECK",
        "price_list_sheet": "IC卡25.10降价",
    },
    "监控": {
        "label": "监控",
        "filter_materials": [1000027313],
        "data_sheet": "监控数据",
        "target_sheet": "监控透视表",
        "pivot_table_name": "监控透视表",
        "content_filter": "监控",
        "content_filter_values": ("监控",),
        "content_filter_display": "监控",
        "content_filter_exact_on_multiple_matches": True,
        "insert_cols_after_order": ["老价格", "新价格", "Saving"],
        "insert_col_at_end": "CHECK",
        "price_list_sheet": "Monitor 25.10降价",
    },
    "LCD": {
        "label": "LCD",
        "filter_materials": [1000027318],
        "data_sheet": "LCD数据",
        "target_sheet": "LCD透视表",
        "pivot_table_name": "LCD透视表",
        "content_filter": "LCD",
        "content_filter_values": ("LCD",),
        "content_filter_display": "LCD",
        "content_filter_exact_on_multiple_matches": True,
        "insert_cols_after_order": ["老价格", "新价格", "Saving"],
        "insert_col_at_end": "CHECK",
        "price_list_sheet": "LCD 2026.1月降价",
    },
}

# 向后兼容：保留全局变量指向装潢默认配置（旧代码无参调用仍然有效）
ACTIVE_CATEGORY = "装潢"

def _cfg(category=None):
    """获取品类配置，默认使用 ACTIVE_CATEGORY"""
    return CATEGORIES[category or ACTIVE_CATEGORY]

# 向后兼容别名（旧代码引用这些全局变量时自动指向装潢配置）
DATA_SHEET = CATEGORIES["装潢"]["data_sheet"]
TARGET_SHEET = CATEGORIES["装潢"]["target_sheet"]
FILTER_MATERIALS = CATEGORIES["装潢"]["filter_materials"]

# Excel COM 常量
XlDatabase = 1
XlRowField = 1
XlSum = -4157
XlPivotTableVersion15 = 6


# ═══════════════════ 工具函数 ═══════════════════

def _col_letter(idx):
    """1→A, 26→Z, 27→AA, ..."""
    s = ""
    while idx > 0:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


def _com_start():
    """启动 Excel COM 实例"""
    # 清除 pywin32 COM 缓存，避免 CLSIDToClassMap 错误
    import shutil
    for tmp_base in (
        os.environ.get("TEMP", ""),
        os.path.join(os.environ.get("USERPROFILE", ""), "AppData", "Local", "Temp"),
    ):
        gen_py = os.path.join(tmp_base, "gen_py")
        if os.path.isdir(gen_py):
            try:
                shutil.rmtree(gen_py, ignore_errors=True)
            except Exception:
                pass

    import pythoncom
    pythoncom.CoInitialize()
    from win32com.client import Dispatch
    excel = Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    return excel


def _com_stop(excel, wb=None):
    """安全关闭 COM 资源"""
    if wb is not None:
        try:
            wb.Close(SaveChanges=False)
        except Exception:
            pass
    if excel is not None:
        try:
            excel.ScreenUpdating = True
            excel.DisplayAlerts = True
            excel.Quit()
        except Exception:
            pass
    try:
        import pythoncom
        pythoncom.CoUninitialize()
    except Exception:
        pass


def _com_try_open(excel, path):
    """尝试用 COM 打开文件，失败返回 None"""
    try:
        return excel.Workbooks.Open(path)
    except Exception:
        return None


def _com_create_pivot(
    wb_com, total_cols, total_rows, source_name, target_sheet, pivot_table_name,
    row_fields=None, value_field=None, value_name=None,
):
    """在 COM 工作簿中创建/更新透视表"""
    row_fields = row_fields or ROW_FIELDS
    value_field = value_field or VALUE_FIELD
    value_name = value_name or VALUE_NAME
    # 删除旧透视表 Sheet
    for s in list(wb_com.Sheets):
        if s.Name == target_sheet:
            s.Delete()
            break

    # 定位数据源区域
    src_ws = wb_com.Sheets(source_name)
    data_range = src_ws.Range(f"A1:{_col_letter(total_cols)}{total_rows + 1}")

    # 在 Sheet1 右侧新建透视表 Sheet
    sheet1 = wb_com.Sheets(SOURCE_SHEET)
    new_ws = wb_com.Sheets.Add(After=sheet1)
    new_ws.Name = target_sheet

    # 创建数据透视表缓存
    pivot_cache = wb_com.PivotCaches().Create(
        SourceType=XlDatabase, SourceData=data_range, Version=XlPivotTableVersion15
    )

    # 在工作表上创建透视表
    pivot_table = pivot_cache.CreatePivotTable(
        TableDestination=new_ws.Cells(1, 1), TableName=pivot_table_name
    )

    # 添加行字段（层级顺序）
    for i, field_name in enumerate(row_fields, start=1):
        pf = pivot_table.PivotFields(field_name)
        pf.Orientation = XlRowField
        pf.Position = i

    # 添加值字段（求和）
    pivot_table.AddDataField(
        pivot_table.PivotFields(value_field), value_name, XlSum
    )
    return new_ws, pivot_table


def _lighting_pivot_count_expression(description):
    """返回从井道照明透视表读取单个描述汇总数量的 Excel 表达式。"""
    escaped_description = description.replace('"', '""')
    return (
        'IFERROR(GETPIVOTDATA("求和项:PO数量",$A$1,"描述",'
        f'"{escaped_description}"),0)'
    )


def _apply_lighting_saving_layout(worksheet):
    """在井道照明 Saving Sheet 右侧写入固定业务规则和可刷新的计算公式。"""
    rules = (
        (
            "有机房 7米方案", 680.043, 565.562, 114.481,
            "SAP PO 数量 (=Shaft Lighting 044 + 045)",
            f"={_lighting_pivot_count_expression('Shaft Lighting 044')}+"
            f"{_lighting_pivot_count_expression('Shaft Lighting 045')}",
        ),
        (
            "有机房 3+2 方案", 529.018, 469.891, 59.127,
            "SAP PO 数量 (=Shaft Lighting 061 - 070)",
            f"={_lighting_pivot_count_expression('Shaft Lighting 061')}-"
            f"{_lighting_pivot_count_expression('Shaft Lighting 070')}",
        ),
        (
            "有机房 3+2 方案", 529.018, 511.481, 17.537,
            "SAP PO 数量 (=Shaft Lighting 053- (044+045) - (061-070))",
            f"={_lighting_pivot_count_expression('Shaft Lighting 053')}-N6-N7",
        ),
        (
            "无机房 7米方案", 358.7745, 274.023, 84.7515,
            "SAP PO 数量 (=Shaft Lighting 069)",
            f"={_lighting_pivot_count_expression('Shaft Lighting 069')}",
        ),
        (
            "无机房 3+2 方案", 325.238, 242.831, 82.407,
            "SAP PO 数量 (=Shaft Lighting 070)",
            f"={_lighting_pivot_count_expression('Shaft Lighting 070')}",
        ),
        (
            "无机房 3+2 方案", 325.238, 320.241, 4.997,
            "SAP PO 数量 (=Shaft Lighting071-069-070)",
            f"={_lighting_pivot_count_expression('Shaft Lighting 071')}-N9-N10",
        ),
    )
    headers = (
        "2024 年 11 月 1 号版本价格", "2025 年 10 月 30 号版本价格",
        "Saving (单台)", "台量统计方法", "实际梯台数", "Saving 总额",
    )

    worksheet.Range("I5:J5").Merge()
    worksheet.Range("I5").Value = headers[0]
    for column, header in zip(("K", "L", "M", "N", "O"), headers[1:]):
        worksheet.Range(f"{column}5").Value = header

    for row_number, rule in enumerate(rules, start=6):
        scheme, old_price, new_price, unit_saving, count_method, count_formula = rule
        worksheet.Range(f"I{row_number}").Value = scheme
        worksheet.Range(f"J{row_number}").Value = old_price
        worksheet.Range(f"K{row_number}").Value = new_price
        worksheet.Range(f"L{row_number}").Value = unit_saving
        worksheet.Range(f"M{row_number}").Value = count_method
        worksheet.Range(f"N{row_number}").Formula = count_formula
        worksheet.Range(f"O{row_number}").Formula = f"=N{row_number}*L{row_number}"

    worksheet.Range("O12").Formula = "=SUM(O6:O11)"
    worksheet.Range("J6:L11").NumberFormat = "0.000"
    worksheet.Range("N6:N11").NumberFormat = "0.000"
    worksheet.Range("O6:O12").NumberFormat = "0.000"


# ═══════════════════ 数据增强 ═══════════════════

def _build_plant_mapping(_log):
    """
    读取 工厂清单 7.XLSX，建立 Plant → (区域, 分公司) 映射字典。
    按列位置读取：A=Plant, J=区域(英文), K=分公司(拼音)
    （由独立的「匹配区域/分公司」按钮调用，品类无关）
    """
    import pandas as pd
    _log("正在读取工厂清单...")
    df_factory = pd.read_excel(FACTORY_FILE, sheet_name=0, header=None)

    # ── 诊断：打印前 3 行原始数据 ──
    debug_lines = ["[诊断] 工厂清单前 3 行 A/J/K 列原始值:"]
    for i in range(min(3, len(df_factory))):
        a_raw = repr(df_factory.iloc[i, 0])
        j_raw = repr(df_factory.iloc[i, 9])
        k_raw = repr(df_factory.iloc[i, 10])
        debug_lines.append(f"  行{i+1}: A={a_raw}  J={j_raw}  K={k_raw}")

    plant_map = {}
    skipped = 0
    for _, row in df_factory.iterrows():
        plant = str(row[0]).strip().upper().replace('\u3000', '').replace('\xa0', '')   # A 列: Plant
        region = str(row[9]).strip()  # J 列: 区域 (South/East 等)
        branch = str(row[10]).strip() # K 列: 分公司 (Zhongshan 拼音)
        if plant and plant != "NAN" and plant != "PLANT":
            plant_map[plant] = (region if region != "nan" else "", branch if branch != "nan" else "")
        else:
            skipped += 1

    # ── 诊断：打印前 5 个映射 key ──
    sample_keys = list(plant_map.keys())[:5]
    debug_lines.append(f"[诊断] 跳过行数: {skipped}, 有效映射条目: {len(plant_map)}")
    debug_lines.append(f"[诊断] 映射样例 (前5个): {[repr(k) for k in sample_keys]}")

    _log(f"工厂映射条目: {len(plant_map)}")
    return plant_map, "\n".join(debug_lines)


# ═══════════════════ 独立工厂映射 ═══════════════════

def apply_factory_mapping(status_callback=None):
    """
    独立功能：读取工厂清单，在 Sheet1 插入/更新「区域」「分公司」列。
    品类无关，跑一次即可。正式员工无论先做装潢还是空调，都可先点此按钮。
    """
    import openpyxl

    _log = lambda msg: status_callback and status_callback(msg)

    if not os.path.exists(EXCEL_FILE):
        return False, f"找不到主数据文件：\n{EXCEL_FILE}"
    if not os.path.exists(FACTORY_FILE):
        return False, f"找不到工厂清单文件：\n{FACTORY_FILE}"

    # ── 1. 读工厂清单 ──
    try:
        plant_map, factory_debug = _build_plant_mapping(_log)
    except Exception as e:
        return False, f"读取工厂清单失败：\n{e}"

    # ── 2. 打开 Sheet1 ──
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SOURCE_SHEET]

    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None:
            headers[str(v).strip()] = c

    if "工厂" not in headers:
        wb.close()
        return False, "Sheet1 缺少必需列：工厂"
    if "订单净值" not in headers:
        wb.close()
        return False, "Sheet1 缺少必需列：订单净值"

    factory_col = headers["工厂"]
    order_col = headers["订单净值"]

    # ── 3. 插入/定位 区域、分公司 列 ──
    has_region = "区域" in headers
    has_branch = "分公司" in headers

    if not has_region and not has_branch:
        ws.insert_cols(order_col + 1, 2)
        region_col = order_col + 1
        branch_col = order_col + 2
        ws.cell(1, region_col, "区域")
        ws.cell(1, branch_col, "分公司")
        _log("已在「订单净值」右侧插入「区域」和「分公司」列")
    elif has_region:
        region_col = headers["区域"]
        branch_col = headers.get("分公司", region_col + 1)
        _log("区域列已存在，将更新值")
    else:
        branch_col = headers["分公司"]
        ws.insert_cols(branch_col, 1)
        region_col = branch_col
        branch_col = branch_col + 1
        ws.cell(1, region_col, "区域")
        _log("补插「区域」列")

    # ── 4. 填值 ──
    matched, unmatched = 0, 0
    unmatched_samples = []
    for r in range(2, ws.max_row + 1):
        fv = str(ws.cell(r, factory_col).value or "").strip().upper().replace('\u3000', '').replace('\xa0', '')
        if fv in plant_map:
            reg, brh = plant_map[fv]
            ws.cell(r, region_col, reg)
            ws.cell(r, branch_col, brh)
            matched += 1
        else:
            unmatched += 1
            if len(unmatched_samples) < 5:
                unmatched_samples.append(repr(fv))

    _log(f"区域/分公司匹配: 成功 {matched}, 未匹配 {unmatched}")

    wb.save(EXCEL_FILE)
    wb.close()

    # ── 摘要 ──
    summary = (
        f"区域/分公司匹配完成！\n\n"
        f"匹配成功：{matched} 行\n"
        f"未匹配：{unmatched} 行\n"
        f"工厂映射条目总数：{len(plant_map)}\n\n"
        f"{factory_debug}"
    )
    if unmatched_samples:
        summary += f"\n\n未匹配样例：{unmatched_samples[:5]}"

    return True, summary


# ═══════════════════ 数据增强 ═══════════════════

def _enhance_and_filter(cfg, _log):
    """
    品类驱动的 Sheet1 增强 + 筛选流程：
      1. 打开 Sheet1，定位「订单净值」列
      2. 按品类配置在订单净值右侧插入列（cfg["insert_cols_after_order"]）
      3. 在 Sheet1 末尾插入列（cfg["insert_col_at_end"]）
      4. pandas 筛选物料 → 写入品类数据 Sheet
    返回: (total_cols, total_rows) 筛选后数据的行列数（供 COM 透视表使用）
    """
    import openpyxl, pandas as pd

    # ── 1. 打开原文件，找列位置 ──
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SOURCE_SHEET]

    # 扫描表头，定位各列
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None:
            headers[str(v).strip()] = c

    if "订单净值" not in headers:
        wb.close()
        raise RuntimeError("Sheet1 缺少必需列：订单净值")
    order_col = headers["订单净值"]

    debug_lines = []
    label = cfg["label"]

    # ── 2. 品类差异化：在订单净值右侧插入列 ──
    insert_cols = cfg["insert_cols_after_order"]
    if insert_cols:
        # 检查这些列是否已存在（可能上次运行已插入）
        all_exist = all(c in headers for c in insert_cols)
        if not all_exist:
            # 在订单净值右边一次性插入所有列
            ws.insert_cols(order_col + 1, len(insert_cols))
            for i, col_name in enumerate(insert_cols):
                ws.cell(1, order_col + 1 + i, col_name)
            _log(f"[{label}] 已在「订单净值」右侧插入列：{', '.join(insert_cols)}")
            # 刷新 headers（插入列后列号已变化）
            headers = {}
            for c in range(1, ws.max_column + 1):
                v = ws.cell(1, c).value
                if v is not None:
                    headers[str(v).strip()] = c
            order_col = headers["订单净值"]
        else:
            _log(f"[{label}] 列 {', '.join(insert_cols)} 已存在，跳过插入")

    # ── 3. 在 Sheet1 末尾插入列 ──
    end_col_name = cfg["insert_col_at_end"]
    if end_col_name:
        # 刷新 headers（前面的插入可能改变了列数）
        headers = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            if v is not None:
                headers[str(v).strip()] = c
        if end_col_name not in headers:
            last_col = ws.max_column
            ws.cell(1, last_col + 1, end_col_name)
            _log(f"[{label}] 已在 Sheet1 末尾插入「{end_col_name}」列")
        else:
            _log(f"[{label}] 列「{end_col_name}」已存在，跳过插入")

    debug_str = "\n".join(debug_lines)

    wb.save(EXCEL_FILE)
    wb.close()

    # ── 4. pandas 读增强后数据 → 筛选 → 写入品类数据 Sheet ──
    filter_materials = cfg["filter_materials"]
    data_sheet = cfg["data_sheet"]
    _log(f"[{label}] 正在筛选物料 {filter_materials}...")
    df_all = pd.read_excel(EXCEL_FILE, sheet_name=SOURCE_SHEET)
    # 确保物料列为数值类型用于比较
    df_all["物料"] = pd.to_numeric(df_all["物料"], errors="coerce")
    df_filtered = df_all[df_all["物料"].isin(filter_materials)].copy()
    df_filtered = df_filtered.reset_index(drop=True)

    if len(df_filtered) == 0:
        raise RuntimeError(f"[{label}] 筛选后无数据！物料 {filter_materials} 在数据中不存在。")
    _log(f"[{label}] 筛选后数据行数: {len(df_filtered)}")

    # 井道照明只在独立数据 Sheet 中增加描述和计数辅助列，不能污染 Sheet1。
    if cfg.get("workflow") == "lighting_saving":
        if "短文本" not in df_filtered.columns:
            raise RuntimeError(f"[{label}] 筛选数据缺少必需列：短文本")
        source_text = df_filtered["短文本"].where(
            df_filtered["短文本"].notna(), ""
        ).astype(str)
        df_filtered["描述"] = source_text.str.slice(2).str.strip()
        df_filtered["PO数量"] = 1

    # 轻量价格类只需要在品类数据 Sheet 中追加计算列，不能污染 Sheet1，
    # 也不会为后续透视表或其他品类预先插入无关列。
    for column_name in cfg.get("data_extra_columns", []):
        if column_name not in df_filtered.columns:
            df_filtered[column_name] = None

    # 写入品类数据 Sheet
    wb2 = openpyxl.load_workbook(EXCEL_FILE)
    if data_sheet in wb2.sheetnames:
        del wb2[data_sheet]
    ds = wb2.create_sheet(title=data_sheet)
    # 写表头
    for ci, col_name in enumerate(df_filtered.columns, 1):
        ds.cell(1, ci, col_name)
    # 写数据
    for ri, row in df_filtered.iterrows():
        for ci, val in enumerate(row, 1):
            ds.cell(ri + 2, ci, val)
    wb2.save(EXCEL_FILE)
    wb2.close()

    return len(df_filtered.columns), len(df_filtered), debug_str


# ═══════════════════ 扩展列 ═══════════════════

def add_extra_columns(category=None):
    """
    在透视表 Sheet 中，透视表右侧追加 EXTRA_COLUMNS 表头。
    重复调用会自动更新列名（旧列名自动替换为新列名，已有数据不丢失）。
    """
    import openpyxl

    cfg = _cfg(category)
    target_sheet = cfg["target_sheet"]
    label = cfg["label"]

    if not os.path.exists(EXCEL_FILE):
        return False, f"找不到文件：\n{EXCEL_FILE}"

    wb = openpyxl.load_workbook(EXCEL_FILE)
    if target_sheet not in wb.sheetnames:
        wb.close()
        return False, f"Sheet \"{target_sheet}\" 不存在，请先生成透视表。"

    ws = wb[target_sheet]

    # 找到第 1 行最右侧已用列
    max_col = ws.max_column
    last_used = 0
    for c in range(max_col, 0, -1):
        if ws.cell(1, c).value is not None:
            last_used = c
            break

    if last_used == 0:
        wb.close()
        return False, f"Sheet \"{target_sheet}\" 第 1 行为空，无法定位。"

    # 定位扩展列起始列：优先找已有「E2E项目名」，否则从 last_used + 1 开始
    start_col = None
    for c in range(1, max_col + 1):
        if str(ws.cell(1, c).value or "").strip() == EXTRA_COLUMNS[0]:
            start_col = c
            break
    if start_col is None:
        start_col = last_used + 1
    elif str(ws.cell(1, start_col + 1).value or "").strip() != "WBS":
        # 旧版已有 9 列扩展列。WBS 插在第二列时必须整体右移旧列，
        # 否则会把数量、金额等已回填数据误配到新的表头。
        ws.insert_cols(start_col + 1, 1)

    # 写入/更新表头（覆盖旧列名，保留已有数据）
    updated = 0
    for i, col_name in enumerate(EXTRA_COLUMNS):
        cell = ws.cell(1, start_col + i)
        existing = str(cell.value or "").strip()
        if existing == col_name:
            continue
        cell.value = col_name
        updated += 1

    wb.save(EXCEL_FILE)
    wb.close()

    col_range = f"{_col_letter(start_col)}~{_col_letter(start_col + len(EXTRA_COLUMNS) - 1)}"
    action = f"更新了 {updated} 个列名" if updated > 0 else "列名已是最新，无需更新"
    return True, f"已在「{target_sheet}」透视表右侧配置 {len(EXTRA_COLUMNS)} 个扩展列表头。\n列范围：{col_range}\n{action}\n列名：{', '.join(EXTRA_COLUMNS)}"


# ═══════════════════ NI PM Saving Tracking 匹配 ═══════════════════

TRACKING_FILE = os.path.join(SCRIPT_DIR, "NI PM Saving Tracking.xlsx")
TRACKING_SHEET = "Base Data"
TRACKING_PROJECT_COL = "Project Name"
SUPPORT_FILE_NAME_COL = "支持文件名(FLD 审批)"
APPROVAL_AMOUNT_COL = "项目总金额(审批)"
PRICE_APPROVAL_TOLERANCE = 0.01
FUZZY_MATCH_THRESHOLD = 0.6  # n-gram 相似度阈值，精确子串匹配失败时启用


def _ngram_similarity(s1, s2, n=2):
    """计算两个字符串的 n-gram 相似度（0.0~1.0），用于模糊匹配兜底。"""
    s1 = s1.lower()
    s2 = s2.lower()
    if len(s1) < n or len(s2) < n:
        return 0.0
    ng1 = {s1[i:i+n] for i in range(len(s1) - n + 1)}
    ng2 = {s2[i:i+n] for i in range(len(s2) - n + 1)}
    if not ng1 or not ng2:
        return 0.0
    intersection = ng1 & ng2
    return len(intersection) / min(len(ng1), len(ng2))


def _as_number(value):
    """将金额单元格转换为数值；空白或非数值返回 None，供人工复核提示使用。"""
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def match_pm_tracking_data(category=None):
    """
    读取透视表中的 E2E项目名，去 NI PM Saving Tracking.xlsx
    的 Base Data Sheet 的 Project Name 列中做模糊匹配（包含即匹配）。

    匹配成功：从 Tracking 提取 Price/PlanCost，读取透视表订单净值，
    计算总saving（= 订单净值 × PlanCost ÷ Price）和 订单是否下完（= Price - 订单净值），
    写入透视表对应列。Price 为 0 或空时填入 "N/A"。
    未匹配：跳过不写。

    返回: (success: bool, message: str)
    """
    import openpyxl

    cfg = _cfg(category)
    target_sheet = cfg["target_sheet"]
    content_filter = cfg["content_filter"]
    content_filter_values = tuple(cfg.get("content_filter_values", (content_filter,)))
    label = cfg["label"]
    requires_fld_support_file = cfg.get("requires_fld_support_file", False)
    content_filter_exact_on_multiple_matches = cfg.get(
        "content_filter_exact_on_multiple_matches", False
    )

    if not os.path.exists(EXCEL_FILE):
        return False, f"找不到文件：\n{EXCEL_FILE}"
    if not os.path.exists(TRACKING_FILE):
        return False, f"找不到 NI PM Saving Tracking 文件：\n{TRACKING_FILE}"

    # ── 1. 打开透视表，扫描第1行定位各列 ──
    wb = openpyxl.load_workbook(EXCEL_FILE)
    if target_sheet not in wb.sheetnames:
        wb.close()
        return False, f"Sheet「{target_sheet}」不存在，请先生成透视表。"

    ws = wb[target_sheet]

    e2e_col = None
    order_value_col = None   # 订单净值（列B "求和项:订单净值"）
    po_col = None             # 采购凭证
    price_col = None
    plancost_col = None
    total_saving_col = None
    saving_diff_col = None
    support_file_col = None
    approval_amount_col = None

    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(1, col).value or "").strip()
        if val == "E2E项目名":
            e2e_col = col
        elif val == "求和项:订单净值":
            order_value_col = col
        elif val == "采购凭证":
            po_col = col
        elif val == "Price":
            price_col = col
        elif val == "PlanCost":
            plancost_col = col
        elif val == "总saving":
            total_saving_col = col
        elif val == "订单是否下完=price-订单净值":
            saving_diff_col = col
        elif val == SUPPORT_FILE_NAME_COL:
            support_file_col = col
        elif val == APPROVAL_AMOUNT_COL:
            approval_amount_col = col

    if e2e_col is None:
        wb.close()
        return False, f"「{target_sheet}」中未找到「E2E项目名」列，请先点击「添加扩展列」按钮。"
    if order_value_col is None:
        wb.close()
        return False, f"「{target_sheet}」中未找到「求和项:订单净值」列，透视表可能未正确生成。"
    if price_col is None or plancost_col is None or total_saving_col is None or saving_diff_col is None:
        wb.close()
        return False, (
            f"「{target_sheet}」中缺少扩展列（Price/PlanCost/总saving/订单是否下完=price-订单净值），"
            f"请先点击「添加扩展列」按钮。"
        )
    if approval_amount_col is None:
        wb.close()
        return False, (
            f"「{target_sheet}」中未找到「{APPROVAL_AMOUNT_COL}」列，"
            "请先点击「添加扩展列」按钮。"
        )
    if requires_fld_support_file and support_file_col is None:
        wb.close()
        return False, (
            f"「{target_sheet}」中未找到「{SUPPORT_FILE_NAME_COL}」列，"
            "请先完成附件回填。"
        )

    # ── 2. 收集待匹配的 E2E项目名和采购凭证（跳过空值）──
    project_names = []  # (row_number, project_name, po_number)
    for row in range(2, ws.max_row + 1):
        val = str(ws.cell(row, e2e_col).value or "").strip()
        if not val:
            continue
        if requires_fld_support_file:
            support_file_name = str(ws.cell(row, support_file_col).value or "").strip()
            if not support_file_name:
                continue
        po_val = str(ws.cell(row, po_col).value or "").strip() if po_col else ""
        project_names.append((row, val, po_val))

    if not project_names:
        wb.close()
        return False, "透视表中未找到任何 E2E项目名，请先运行「打开网站查询」。"

    # ── 3. 打开 NI PM Saving Tracking（只读）──
    wb_track = openpyxl.load_workbook(TRACKING_FILE, read_only=True, data_only=True)
    if TRACKING_SHEET not in wb_track.sheetnames:
        wb_track.close()
        wb.close()
        return False, (
            f"NI PM Saving Tracking 中不存在 Sheet「{TRACKING_SHEET}」。\n"
            f"当前 Sheet 列表：{wb_track.sheetnames}"
        )

    ws_track = wb_track[TRACKING_SHEET]

    # 定位 Tracking 中的 Project Name / Price / PlanCost 列（表头在第2行）
    HEADER_ROW = 2
    project_col = None
    track_price_col = None
    track_plancost_col = None
    track_content_col = None

    for col in range(1, ws_track.max_column + 1):
        val = str(ws_track.cell(HEADER_ROW, col).value or "").strip()
        if val == TRACKING_PROJECT_COL:
            project_col = col
        elif val == "Price":
            track_price_col = col
        elif val == "PlanCost":
            track_plancost_col = col
        elif val == "Content":
            track_content_col = col

    # 诊断：找不到列时打印表头所有非空列
    def _tracking_headers():
        h = []
        for c in range(1, ws_track.max_column + 1):
            v = str(ws_track.cell(HEADER_ROW, c).value or "")
            if v:
                h.append(f"第{c}列={v}")
        return f"共{ws_track.max_column}列，非空表头：" + (', '.join(h) if h else '(空)')

    if project_col is None:
        headers_str = _tracking_headers()
        wb_track.close()
        wb.close()
        return False, (
            f"NI PM Saving Tracking 的「{TRACKING_SHEET}」Sheet 中未找到列「{TRACKING_PROJECT_COL}」。\n"
            f"{headers_str}"
        )

    missing_track = []
    if track_price_col is None:
        missing_track.append("Price")
    if track_plancost_col is None:
        missing_track.append("PlanCost")
    if track_content_col is None:
        missing_track.append("Content")
    if missing_track:
        headers_str = _tracking_headers()
        wb_track.close()
        wb.close()
        return False, (
            f"NI PM Saving Tracking 中缺少列：{', '.join(missing_track)}。\n"
            f"{headers_str}"
        )

    # 收集 Tracking 中所有行数据（iter_rows 批量读取，比逐格快100倍+）
    tracking_rows = []
    max_col = max(project_col, track_price_col, track_plancost_col,
                  track_content_col or 0)
    idx_pn = project_col - 1
    idx_price = track_price_col - 1
    idx_pc = track_plancost_col - 1
    idx_content = (track_content_col - 1) if track_content_col else None
    for row_vals in ws_track.iter_rows(
        min_row=HEADER_ROW + 1,
        min_col=1, max_col=max_col,
        values_only=True,
    ):
        pn_val = str(row_vals[idx_pn] or "").strip()
        if pn_val:
            content_val = str(row_vals[idx_content] or "").strip() if idx_content is not None else ""
            tracking_rows.append((
                pn_val.lower(),  # 预计算小写，加速后续匹配
                row_vals[idx_price],
                row_vals[idx_pc],
                content_val,
            ))

    wb_track.close()

    # Content 是品类边界，不是同名项目出现多行时才使用的附加条件。
    # 必须先限定到当前自动化的品类，再做项目名精确/模糊匹配；否则长项目名
    # 可能因包含了一个装潢短名称而把装潢 Price 错写入空调透视表。
    def _matches_current_category(t_content):
        if content_filter_exact_on_multiple_matches:
            return t_content in content_filter_values
        return t_content.startswith(content_filter)

    category_tracking_rows = [
        track_row for track_row in tracking_rows
        if _matches_current_category(track_row[3])
    ]

    # ── 4. 逐行匹配 + 写入 ──
    matched = 0
    written = 0
    unmatched = 0
    unmatched_samples = []
    na_count = 0
    fuzzy_match_count = 0
    content_multi_count = 0
    content_multi_list = []
    other_multi_count = 0
    other_multi_list = []
    amount_mismatch_list = []
    amount_uncomparable_list = []

    for row_num, pn, po_num in project_names:
        pn_lower = pn.lower()

        # 第一轮：精确子串匹配
        all_matches = []
        for t_pn, t_price, t_plancost, t_content in category_tracking_rows:
            if pn_lower in t_pn:  # t_pn 已在收集时预计算为小写
                all_matches.append((t_pn, t_price, t_plancost, t_content))

        # 兜底：精确匹配无结果时，用 n-gram 模糊匹配
        is_fuzzy = False
        if not all_matches:
            best_score = 0.0
            best_match = None
            for t_pn, t_price, t_plancost, t_content in category_tracking_rows:
                score = _ngram_similarity(pn_lower, t_pn)
                if score > best_score:
                    best_score = score
                    best_match = (t_pn, t_price, t_plancost, t_content)
            if best_score >= FUZZY_MATCH_THRESHOLD and best_match:
                all_matches = [best_match]
                is_fuzzy = True

        found_track = None
        if len(all_matches) == 1:
            found_track = all_matches[0]
        elif len(all_matches) > 1:
            # 候选已通过当前品类的 Content 过滤；仍有多行时属于业务歧义，
            # 不猜测、不写入，交由人工确认。
            content_multi_count += 1
            content_multi_list.append((po_num, pn))
            continue

        if found_track is None:
            unmatched += 1
            unmatched_samples.append((po_num, pn))
            continue

        matched += 1
        if is_fuzzy:
            fuzzy_match_count += 1
        _, t_price, t_plancost, _ = found_track

        # 读取订单净值
        order_value = ws.cell(row_num, order_value_col).value

        # 转换数值
        try:
            price_num = float(t_price) if t_price is not None else 0.0
        except (ValueError, TypeError):
            price_num = 0.0
        try:
            plancost_num = float(t_plancost) if t_plancost is not None else 0.0
        except (ValueError, TypeError):
            plancost_num = 0.0
        try:
            order_num = float(order_value) if order_value is not None else 0.0
        except (ValueError, TypeError):
            order_num = 0.0

        # 写入 Price 和 PlanCost（保持原始值类型）
        ws.cell(row_num, price_col, t_price)
        ws.cell(row_num, plancost_col, t_plancost)

        # 计算总saving 和 订单是否下完
        if price_num == 0:
            ws.cell(row_num, total_saving_col, "N/A")
            ws.cell(row_num, saving_diff_col, "N/A")
            na_count += 1
        else:
            total_saving = order_num * plancost_num / price_num
            saving_diff = price_num - order_num
            ws.cell(row_num, total_saving_col, round(total_saving, 2))
            ws.cell(row_num, saving_diff_col, round(saving_diff, 2))
            written += 1

        # PM 匹配完成后，以透视表最终写入的 Price 对比项目总金额（审批）。
        # 不一致或任一金额为空/非数值都仅记录；全部项目处理完后统一提示人工。
        written_price = ws.cell(row_num, price_col).value
        approval_amount = ws.cell(row_num, approval_amount_col).value
        price_for_check = _as_number(written_price)
        approval_for_check = _as_number(approval_amount)
        if price_for_check is None or approval_for_check is None:
            amount_uncomparable_list.append((po_num, pn, written_price, approval_amount))
        else:
            raw_amount_difference = price_for_check - approval_for_check
            if abs(raw_amount_difference) > PRICE_APPROVAL_TOLERANCE:
                amount_mismatch_list.append(
                    (
                        po_num,
                        pn,
                        written_price,
                        approval_amount,
                        round(raw_amount_difference, 2),
                    )
                )

    # ── 5. 保存 ──
    wb.save(EXCEL_FILE)
    wb.close()

    # ── 6. 返回摘要 ──
    sep = "─" * 42
    summary_lines = [
        f"{sep}",
        f"  PM Tracking 匹配报告",
        f"{sep}",
        f"",
        f"  待匹配项目总数：{len(project_names)}",
        f"",
    ]

    # ✅ 成功
    skipped_total = content_multi_count + other_multi_count
    summary_lines.append(f"  ✅ 匹配成功：{matched} 个（Price/PlanCost/总saving/订单是否下完 已写入）")
    if fuzzy_match_count > 0:
        summary_lines.append(f"     其中模糊匹配（n-gram 兜底）：{fuzzy_match_count} 个")
    if na_count > 0:
        summary_lines.append(f"     Price 为 0/空（填入 N/A）：{na_count} 个")

    if amount_mismatch_list:
        summary_lines.append("")
        summary_lines.append(
            f"  ⚠ Price 与{APPROVAL_AMOUNT_COL}不一致（需人工介入）："
            f"{len(amount_mismatch_list)} 个（允许误差 ±{PRICE_APPROVAL_TOLERANCE:.2f}）"
        )
        for po, pn, price, approval, difference in amount_mismatch_list:
            summary_lines.append(
                f"     · PO {po} — {pn[:40]} | Price={price} | 审批={approval} | 差额={difference}"
            )

    if amount_uncomparable_list:
        summary_lines.append("")
        summary_lines.append(
            f"  ⚠ Price 与{APPROVAL_AMOUNT_COL}无法比较（需人工确认）："
            f"{len(amount_uncomparable_list)} 个"
        )
        for po, pn, price, approval in amount_uncomparable_list:
            summary_lines.append(
                f"     · PO {po} — {pn[:40]} | Price={price} | 审批={approval}"
            )

    # ⚠ 跳过
    if skipped_total > 0:
        summary_lines.append(f"")
        summary_lines.append(f"  ⚠ 跳过（需人工确认）：{skipped_total} 个")
    if content_multi_count > 0:
        for po, pn in content_multi_list:
            summary_lines.append(f"     · PO {po} — {pn[:40]}")
            summary_lines.append(
                f"       原因：Content 列「{' / '.join(content_filter_values)}」匹配到多行"
            )
    if other_multi_count > 0:
        for po, pn in other_multi_list:
            summary_lines.append(f"     · PO {po} — {pn[:40]}")
            summary_lines.append(f"       原因：匹配到多行且均非{label}")

    # ❌ 未匹配
    if unmatched > 0:
        summary_lines.append(f"")
        summary_lines.append(f"  ❌ 未匹配（共 {unmatched} 个）：")
        for po, pn in unmatched_samples:
            summary_lines.append(f"     · PO {po}  —  {pn[:50]}")

    # 底部文件信息
    summary_lines.extend([
        f"",
        f"{sep}",
        f"  Tracking：{os.path.basename(TRACKING_FILE)}",
        f"  Sheet：{TRACKING_SHEET}  |  匹配列：{TRACKING_PROJECT_COL}  |  总行数：{len(tracking_rows)}",
        f"{sep}",
    ])

    summary_text = "\n".join(summary_lines)

    # ── 写入 txt 文件（方便复制分享）──
    report_path = os.path.join(SCRIPT_DIR, "PM匹配报告.txt")
    try:
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(summary_text)
    except Exception:
        pass

    return True, summary_text


# ═══════════════════ 主流程 ═══════════════════

PRICE_INSTALLATION_TEXT = "安装调试费（含辅料）"


def _fill_data_price_category(category, price_file, status_callback=None):
    """填充只含数据 Sheet 的轻量品类：价格匹配 + Saving，不依赖透视表。"""
    import math
    import openpyxl

    cfg = CATEGORIES.get(category)
    if not cfg or cfg.get("workflow") != "data_price":
        return False, f"品类「{category}」未配置轻量价格流程。"

    label = cfg["label"]
    data_sheet_name = cfg["data_sheet"]
    price_list_sheet = cfg["price_list_sheet"]
    source_columns = cfg["price_source_columns"]

    def _status(message):
        if status_callback:
            status_callback(message)

    def _normalise(value):
        return "" if value is None else str(value).strip()

    def _number(value):
        if value is None or (isinstance(value, str) and not value.strip()):
            return None
        try:
            number = float(str(value).replace(",", "").strip())
        except (TypeError, ValueError):
            return None
        return number if math.isfinite(number) else None

    if not os.path.exists(EXCEL_FILE):
        return False, f"找不到采购记录文件：\n{EXCEL_FILE}"
    if not price_file or not os.path.isfile(price_file):
        return False, f"找不到{label}价格表：\n{price_file or '未选择文件'}"

    _status(f"正在读取{label}价格表并按短文本建立匹配索引...")
    try:
        price_book = openpyxl.load_workbook(
            price_file, read_only=True, data_only=True
        )
    except Exception as exc:
        return False, f"无法打开{label}价格表：\n{exc}"

    if price_list_sheet not in price_book.sheetnames:
        available = "、".join(price_book.sheetnames)
        price_book.close()
        return False, (
            f"价格表中未找到 Sheet「{price_list_sheet}」。\n"
            f"当前可用 Sheet：{available}"
        )

    price_by_short_text = {}
    max_source_column = max(source_columns.values())
    try:
        price_sheet = price_book[price_list_sheet]
        for row in price_sheet.iter_rows(
            min_row=2, max_col=max_source_column, values_only=True
        ):
            short_text = _normalise(row[source_columns["sap_description"] - 1])
            if not short_text:
                continue
            old_price = row[source_columns["old_price"] - 1]
            new_price = row[source_columns["new_price"] - 1]
            if old_price is None or new_price is None:
                continue
            price_by_short_text.setdefault(short_text, set()).add(
                (old_price, new_price)
            )
    finally:
        price_book.close()

    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
    except Exception as exc:
        return False, f"无法打开采购记录文件：\n{exc}"

    if data_sheet_name not in workbook.sheetnames:
        workbook.close()
        return False, f"采购记录中未找到 Sheet「{data_sheet_name}」，请先生成线槽数据。"

    worksheet = workbook[data_sheet_name]
    headers = {
        _normalise(worksheet.cell(1, column).value): column
        for column in range(1, worksheet.max_column + 1)
        if _normalise(worksheet.cell(1, column).value)
    }
    required_headers = (
        "短文本", "净价", "采购订单数量", "老价格", "新价格", "Saving",
    )
    missing_headers = [header for header in required_headers if header not in headers]
    if missing_headers:
        workbook.close()
        return False, f"「{data_sheet_name}」缺少列：{', '.join(missing_headers)}。"

    short_text_col = headers["短文本"]
    net_price_col = headers["净价"]
    order_quantity_col = headers["采购订单数量"]
    old_price_col = headers["老价格"]
    new_price_col = headers["新价格"]
    saving_col = headers["Saving"]

    matched_rows = 0
    saving_rows = 0
    saving_skipped = 0
    unmatched_texts = set()
    ambiguous_texts = set()
    _status(f"正在匹配{label}老价格、新价格并计算 Saving...")
    for row_number in range(2, worksheet.max_row + 1):
        short_text = _normalise(worksheet.cell(row_number, short_text_col).value)
        if not short_text:
            continue
        candidates = price_by_short_text.get(short_text, set())
        if len(candidates) == 0:
            unmatched_texts.add(short_text)
            continue
        if len(candidates) > 1:
            ambiguous_texts.add(short_text)
            continue

        old_price, new_price = next(iter(candidates))
        worksheet.cell(row_number, old_price_col, old_price)
        worksheet.cell(row_number, new_price_col, new_price)
        matched_rows += 1

        old_price_number = _number(old_price)
        net_price_number = _number(worksheet.cell(row_number, net_price_col).value)
        order_quantity_number = _number(
            worksheet.cell(row_number, order_quantity_col).value
        )
        if (
            old_price_number is None
            or net_price_number is None
            or order_quantity_number is None
        ):
            saving_skipped += 1
            continue

        saving = round(
            (old_price_number - net_price_number) * order_quantity_number, 2
        )
        worksheet.cell(row_number, saving_col, 0 if saving == 0 else saving)
        saving_rows += 1

    try:
        workbook.save(EXCEL_FILE)
    except Exception as exc:
        workbook.close()
        return False, f"无法保存采购记录文件：\n{exc}"
    workbook.close()

    summary = [
        f"{label}老/新价格与 Saving 填充完成",
        f"短文本 → SAP Discription-Ner 匹配：{matched_rows} 行",
        f"Saving 已计算：{saving_rows} 行（老价格 - 净价）× 采购订单数量",
        f"价格表：{os.path.basename(price_file)} / {price_list_sheet}（H=老价格，J=新价格）",
    ]
    if unmatched_texts:
        summary.append("未匹配并保持为空：" + "、".join(sorted(unmatched_texts)))
    if ambiguous_texts:
        summary.append("价格存在歧义，未写入：" + "、".join(sorted(ambiguous_texts)))
    if saving_skipped:
        summary.append(f"Saving 保持为空：{saving_skipped} 行缺少有效老价格、净价或采购订单数量")
    return True, "\n".join(summary)


def fill_category_old_new_prices(category, price_file, status_callback=None):
    """填充指定品类价格，并用有 FLD 的 PO 汇总值覆盖其全部明细行。

    常规项目仅接受“短文本 = SAP Discription-Ner”的精确匹配；这是为了
    避免名称相近时把价格写到错误物料。仅“安装调试费（含辅料）”可回退到
    Discription，因为该名称在实际台账中未规范写入 SAP 字段；回退结果的
    全部老/新价格必须一致，否则保留为空供人工确认。

    价格表是无 FLD 明细的基础价格。对于已在空调透视表完成 PM Tracking
    的 FLD PO，整组明细统一改为：老价格 = 总saving，新价格 = 订单净值，
    Saving = 老价格 - 新价格。无 FLD 的明细则按（老价格 - 净价）× 采购
    订单数量计算 Saving。最后逐行计算 CHECK = 新价格 - 净价；差额为零时
    明确写入数值 0，而不是保留空白。计算所需值缺失时不猜测，保留原值。
    重复执行时，已同时具有老价格和新价格的无 FLD 行视为已完成，四个结果
    列均保持不变；有 FLD 的 PO 仍按最新透视表数据刷新。
    """
    import math
    import openpyxl

    cfg = CATEGORIES.get(category)
    if not cfg or not cfg.get("price_list_sheet"):
        return False, f"品类「{category}」未配置老/新价格填充。"
    if cfg.get("workflow") == "data_price":
        return _fill_data_price_category(category, price_file, status_callback)
    label = cfg["label"]
    price_list_sheet = cfg["price_list_sheet"]

    def _status(message):
        if status_callback:
            status_callback(message)

    def _normalise(value):
        return str(value or "").strip()

    def _is_blank(value):
        return value is None or str(value).strip() == ""

    def _normalise_po(value):
        """消除 Excel 将整数 PO 读成浮点数时产生的 .0。"""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return _normalise(value)

    def _consistent_price_pair(candidates):
        """同一备用描述允许多行，但它们的老/新价格必须完全一致。"""
        pairs = {(candidate[0], candidate[1]) for candidate in candidates}
        if len(pairs) != 1:
            return None
        old_price, new_price = next(iter(pairs))
        if old_price is None or new_price is None:
            return None
        return old_price, new_price

    def _merged_anchor(worksheet_to_read, row_number, column_number):
        """返回单元格所在合并区域的左上角，普通单元格返回自身。"""
        for merged_range in worksheet_to_read.merged_cells.ranges:
            if (
                merged_range.min_row <= row_number <= merged_range.max_row
                and merged_range.min_col <= column_number <= merged_range.max_col
            ):
                return merged_range.min_row, merged_range.min_col
        return row_number, column_number

    def _read_merged_value(worksheet_to_read, row_number, column_number):
        anchor_row, anchor_col = _merged_anchor(
            worksheet_to_read, row_number, column_number
        )
        return worksheet_to_read.cell(anchor_row, anchor_col).value

    def _write_merged_value(worksheet_to_write, row_number, column_number, value):
        """尊重用户已有的合并格式；合并区域只写入其可编辑的左上角。"""
        anchor_row, anchor_col = _merged_anchor(
            worksheet_to_write, row_number, column_number
        )
        worksheet_to_write.cell(anchor_row, anchor_col, value)

    def _contiguous_row_groups(row_numbers):
        """将同一 PO 的行拆成连续区间，绝不跨越其他 PO 合并单元格。"""
        groups = []
        for row_number in sorted(set(row_numbers)):
            if not groups or row_number != groups[-1][-1] + 1:
                groups.append([row_number])
            else:
                groups[-1].append(row_number)
        return groups

    def _merge_fld_display_group(worksheet_to_write, po_number, row_numbers, columns):
        """将一个连续 FLD PO 区间的汇总列合并为可点击的一组展示。"""
        if len(row_numbers) < 2:
            return 0

        from openpyxl.styles import Alignment, PatternFill

        first_row = row_numbers[0]
        last_row = row_numbers[-1]
        fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        alignment = Alignment(horizontal="center", vertical="center")
        ranges_to_unmerge = []

        for column_number in columns:
            overlapping_ranges = [
                merged_range
                for merged_range in worksheet_to_write.merged_cells.ranges
                if (
                    merged_range.min_col == merged_range.max_col == column_number
                    and merged_range.min_row <= last_row
                    and merged_range.max_row >= first_row
                )
            ]
            for merged_range in overlapping_ranges:
                # 只撤销同一 PO 已有的汇总列合并，避免误伤用户手工制作的其他区域。
                merged_pos = {
                    _normalise_po(
                        worksheet_to_write.cell(row, po_col).value
                    )
                    for row in range(merged_range.min_row, merged_range.max_row + 1)
                }
                if merged_pos == {po_number}:
                    ranges_to_unmerge.append(merged_range)
                else:
                    return 0

        # 三列均确认安全后才解除旧合并，避免其中一列冲突时留下半完成状态。
        for merged_range in ranges_to_unmerge:
            worksheet_to_write.unmerge_cells(str(merged_range))

        for column_number in columns:
            for row_number in row_numbers:
                cell = worksheet_to_write.cell(row_number, column_number)
                cell.fill = fill
                cell.alignment = alignment
            worksheet_to_write.merge_cells(
                start_row=first_row,
                start_column=column_number,
                end_row=last_row,
                end_column=column_number,
            )
        return 1

    def _number(value):
        try:
            number = float(value)
        except (TypeError, ValueError):
            return None
        return number if math.isfinite(number) else None

    if not os.path.exists(EXCEL_FILE):
        return False, f"找不到采购记录文件：\n{EXCEL_FILE}"
    if not price_file or not os.path.isfile(price_file):
        return False, f"找不到{label}价格表：\n{price_file or '未选择文件'}"

    _status(f"正在读取{label}价格表并建立精确匹配索引...")
    try:
        price_book = openpyxl.load_workbook(price_file, read_only=True, data_only=True)
    except Exception as exc:
        return False, f"无法打开{label}价格表：\n{exc}"

    if price_list_sheet not in price_book.sheetnames:
        available = "、".join(price_book.sheetnames)
        price_book.close()
        return False, (
            f"价格表中未找到 Sheet「{price_list_sheet}」。\n"
            f"当前可用 Sheet：{available}"
        )

    price_by_sap_description = {}
    price_by_description = {}
    try:
        price_sheet = price_book[price_list_sheet]
        for row in price_sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 10:
                continue
            pair = (row[7], row[9])  # H=老价格，J=新价格
            sap_description = _normalise(row[4])  # E=SAP Discription-Ner
            description = _normalise(row[5])      # F=Discription
            if sap_description:
                price_by_sap_description.setdefault(sap_description, []).append(pair)
            if description:
                price_by_description.setdefault(description, []).append(pair)
    finally:
        price_book.close()

    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
    except Exception as exc:
        return False, f"无法打开采购记录文件：\n{exc}"

    air_cfg = cfg
    air_data_sheet = air_cfg["data_sheet"]
    air_pivot_sheet = air_cfg["target_sheet"]
    if air_data_sheet not in workbook.sheetnames:
        workbook.close()
        return False, f"采购记录中未找到 Sheet「{air_data_sheet}」。"
    if air_pivot_sheet not in workbook.sheetnames:
        workbook.close()
        return False, (
            f"采购记录中未找到 Sheet「{air_pivot_sheet}」。\n"
            "请先生成透视表、完成附件回填和 PM Tracking 匹配后，再执行第⑥步。"
        )

    worksheet = workbook[air_data_sheet]
    headers = {
        _normalise(worksheet.cell(1, column).value): column
        for column in range(1, worksheet.max_column + 1)
        if _normalise(worksheet.cell(1, column).value)
    }
    required_headers = (
        "采购凭证", "短文本", "老价格", "新价格", "Saving", "CHECK", "净价", "采购订单数量",
    )
    missing_headers = [header for header in required_headers if header not in headers]
    if missing_headers:
        workbook.close()
        return False, f"「{air_data_sheet}」缺少列：{', '.join(missing_headers)}。"

    po_col = headers["采购凭证"]
    short_text_col = headers["短文本"]
    old_price_col = headers["老价格"]
    new_price_col = headers["新价格"]
    saving_col = headers["Saving"]
    check_col = headers["CHECK"]
    net_price_col = headers["净价"]
    purchase_order_quantity_col = headers["采购订单数量"]

    pivot_worksheet = workbook[air_pivot_sheet]
    pivot_headers = {
        _normalise(pivot_worksheet.cell(1, column).value): column
        for column in range(1, pivot_worksheet.max_column + 1)
        if _normalise(pivot_worksheet.cell(1, column).value)
    }
    pivot_order_value_col = pivot_headers.get("求和项:订单净值")
    pivot_total_saving_col = pivot_headers.get("总saving")
    pivot_po_columns = [
        column for column in (
            pivot_headers.get("采购凭证"),
            pivot_headers.get("行标签"),
        ) if column is not None
    ]
    if pivot_order_value_col is None or pivot_total_saving_col is None:
        workbook.close()
        return False, (
            f"「{air_pivot_sheet}」缺少「求和项:订单净值」或「总saving」列。\n"
            "请先完成附件回填和 PM Tracking 匹配后，再执行第⑥步。"
        )
    if not pivot_po_columns:
        workbook.close()
        return False, (
            f"「{air_pivot_sheet}」未找到「采购凭证」或「行标签」列，"
            "无法定位 FLD PO。"
        )

    po_to_rows = {}
    for row_number in range(2, worksheet.max_row + 1):
        po_number = _normalise_po(worksheet.cell(row_number, po_col).value)
        if po_number:
            po_to_rows.setdefault(po_number, []).append(row_number)

    # 第五列允许记录任意支持文件名，不能再把“非空”当成 FLD 标志。
    # 复用网页下载模块的文件名规则，同时扫描两个历史分类目录中的实际文件。
    try:
        import web_query
        classify_downloaded_po = web_query.classify_downloaded_po_attachments
    except (ImportError, AttributeError) as exc:
        workbook.close()
        return False, f"无法加载附件分类逻辑：\n{exc}"

    fld_pos_seen = set()
    no_fld_pos_seen = set()
    download_missing_pos = set()
    download_scan_error_pos = set()
    download_scan_errors = {}
    for po_number in po_to_rows:
        classification = classify_downloaded_po(
            po_number,
            category_label=label,
            download_dir=DOWNLOAD_DIR,
        )
        status = classification.get("status")
        if status == "fld":
            fld_pos_seen.add(po_number)
        elif status == "non_fld":
            no_fld_pos_seen.add(po_number)
        elif status == "missing":
            download_missing_pos.add(po_number)
        else:
            download_scan_error_pos.add(po_number)
            download_scan_errors[po_number] = classification.get("errors") or []

    unclassified_download_pos = download_missing_pos | download_scan_error_pos

    # 原生透视表有两种布局：表格形式用“采购凭证”列，紧凑形式将层级值放在“行标签”列。
    # 只接受能在空调数据中找到的 PO，避免把区域、分公司、供应商名称误识别为采购凭证。
    fld_pivot_values = {}
    ambiguous_fld_pos = set()
    for row_number in range(2, pivot_worksheet.max_row + 1):
        po_number = ""
        for candidate_col in pivot_po_columns:
            candidate = _normalise_po(
                _read_merged_value(pivot_worksheet, row_number, candidate_col)
            )
            if candidate in po_to_rows:
                po_number = candidate
                break
        if not po_number or po_number not in fld_pos_seen:
            continue

        old_value = _read_merged_value(pivot_worksheet, row_number, pivot_total_saving_col)
        new_value = _read_merged_value(pivot_worksheet, row_number, pivot_order_value_col)
        old_number = _number(old_value)
        new_number = _number(new_value)
        if old_number is None or new_number is None:
            continue

        existing = fld_pivot_values.get(po_number)
        if existing and (existing[2], existing[3]) != (old_number, new_number):
            ambiguous_fld_pos.add(po_number)
            continue
        fld_pivot_values[po_number] = (old_value, new_value, old_number, new_number)

    for po_number in ambiguous_fld_pos:
        fld_pivot_values.pop(po_number, None)
    invalid_fld_pos = fld_pos_seen - set(fld_pivot_values) - ambiguous_fld_pos

    # 保护已经完成过按钮⑥的无 FLD 行。首次执行时老/新价格为空，仍会正常
    # 匹配、计算；以后两列均已有值时，不再覆盖老价格、新价格、Saving、CHECK。
    # 如果该 PO 后续被识别为有 FLD，则不进入保护集，仍按最新透视表刷新。
    protected_non_fld_rows = set()
    for row_number in range(2, worksheet.max_row + 1):
        po_number = _normalise_po(worksheet.cell(row_number, po_col).value)
        if po_number not in no_fld_pos_seen:
            continue
        old_price = _read_merged_value(worksheet, row_number, old_price_col)
        new_price = _read_merged_value(worksheet, row_number, new_price_col)
        if not _is_blank(old_price) and not _is_blank(new_price):
            protected_non_fld_rows.add(row_number)

    primary_matches = 0
    installation_fallback_matches = 0
    unmatched_texts = set()
    ambiguous_texts = set()

    _status(f"正在填充{label}基础价格，随后计算 FLD 覆盖与 CHECK...")
    for row_number in range(2, worksheet.max_row + 1):
        po_number = _normalise_po(worksheet.cell(row_number, po_col).value)
        if not po_number or po_number in unclassified_download_pos:
            continue
        if row_number in protected_non_fld_rows:
            continue
        short_text = _normalise(worksheet.cell(row_number, short_text_col).value)
        if not short_text:
            continue

        primary_candidates = price_by_sap_description.get(short_text, [])
        price_pair = _consistent_price_pair(primary_candidates)
        if price_pair is not None:
            primary_matches += 1
        elif primary_candidates:
            ambiguous_texts.add(short_text)
            continue
        elif short_text == PRICE_INSTALLATION_TEXT:
            # 这是唯一经业务确认的非标准名称回退，不能扩展为模糊匹配。
            fallback_candidates = price_by_description.get(short_text, [])
            price_pair = _consistent_price_pair(fallback_candidates)
            if price_pair is not None:
                installation_fallback_matches += 1
            elif fallback_candidates:
                ambiguous_texts.add(short_text)
                continue
            else:
                unmatched_texts.add(short_text)
                continue
        else:
            unmatched_texts.add(short_text)
            continue

        _write_merged_value(worksheet, row_number, old_price_col, price_pair[0])
        _write_merged_value(worksheet, row_number, new_price_col, price_pair[1])

    _status(f"正在用{label}透视表中的 FLD PO 汇总值覆盖全部对应明细...")
    fld_override_rows = 0
    for po_number, (old_value, new_value, old_number, new_number) in fld_pivot_values.items():
        saving_value = round(old_number - new_number, 2)
        for row_number in po_to_rows[po_number]:
            _write_merged_value(worksheet, row_number, old_price_col, old_value)
            _write_merged_value(worksheet, row_number, new_price_col, new_value)
            _write_merged_value(worksheet, row_number, saving_col, saving_value)
            fld_override_rows += 1

    _status("正在计算无 FLD 明细的 Saving（老价格 - 净价）× 采购订单数量...")
    non_fld_saving_written = 0
    non_fld_saving_skipped = 0
    for row_number in range(2, worksheet.max_row + 1):
        po_number = _normalise_po(worksheet.cell(row_number, po_col).value)
        # 只要 PO 有 FLD 附件，就必须保留透视表的项目汇总 Saving，即使该汇总
        # 当前缺值或有歧义，也不能退回为普通物料的逐行计算。
        if po_number not in no_fld_pos_seen:
            continue
        if row_number in protected_non_fld_rows:
            continue

        old_price = _number(_read_merged_value(worksheet, row_number, old_price_col))
        net_price = _number(_read_merged_value(worksheet, row_number, net_price_col))
        order_quantity = _number(
            _read_merged_value(worksheet, row_number, purchase_order_quantity_col)
        )
        if old_price is None or net_price is None or order_quantity is None:
            non_fld_saving_skipped += 1
            continue

        saving_value = round((old_price - net_price) * order_quantity, 2)
        if saving_value == 0:
            saving_value = 0
        _write_merged_value(worksheet, row_number, saving_col, saving_value)
        non_fld_saving_written += 1

    _status("正在计算 CHECK（新价格 - 净价）...")
    check_written = 0
    check_skipped = 0
    for row_number in range(2, worksheet.max_row + 1):
        po_number = _normalise_po(worksheet.cell(row_number, po_col).value)
        if not po_number or po_number in unclassified_download_pos:
            continue
        if row_number in protected_non_fld_rows:
            continue
        new_price = _number(_read_merged_value(worksheet, row_number, new_price_col))
        net_price = _number(_read_merged_value(worksheet, row_number, net_price_col))
        if new_price is None or net_price is None:
            check_skipped += 1
            continue

        check_value = round(new_price - net_price, 2)
        # Excel 显示 -0.0 容易造成误判；业务要求差额为零时必须写入整数 0。
        if check_value == 0:
            check_value = 0
        _write_merged_value(worksheet, row_number, check_col, check_value)
        check_written += 1

    _status("正在将连续的 FLD PO 汇总列合并展示...")
    fld_merged_groups = 0
    for po_number in fld_pivot_values:
        for row_group in _contiguous_row_groups(po_to_rows[po_number]):
            fld_merged_groups += _merge_fld_display_group(
                worksheet,
                po_number,
                row_group,
                (old_price_col, new_price_col, saving_col),
            )

    try:
        workbook.save(EXCEL_FILE)
    except Exception as exc:
        workbook.close()
        return False, f"无法保存采购记录文件：\n{exc}"
    workbook.close()

    summary = [
        f"{label}老/新价格填充完成",
        (
            f"实际附件分类：FLD {len(fld_pos_seen)} 个 PO / "
            f"无 FLD {len(no_fld_pos_seen)} 个 PO / "
            f"未判定 {len(unclassified_download_pos)} 个 PO"
        ),
        f"主匹配（短文本 → SAP Discription-Ner）：{primary_matches} 行",
        f"备用匹配（安装调试费（含辅料）→ Discription）：{installation_fallback_matches} 行",
        f"FLD PO 覆盖：{len(fld_pivot_values)} 个 PO / {fld_override_rows} 条明细",
        "FLD 覆盖规则：老价格 = 总saving；新价格 = 订单净值；Saving = 两者差额",
        f"FLD 分组合并：{fld_merged_groups} 组（老价格、新价格、Saving；淡黄色）",
        f"无 FLD 已有结果保护：{len(protected_non_fld_rows)} 条（重复执行不覆盖四列）",
        f"无 FLD Saving：已计算 {non_fld_saving_written} 条（老价格 - 净价）× 采购订单数量",
        f"CHECK：已计算 {check_written} 条（新价格 - 净价；差额为零写 0）",
    ]
    if unmatched_texts:
        summary.append("未匹配并保持为空：" + "、".join(sorted(unmatched_texts)))
    if ambiguous_texts:
        summary.append("价格存在歧义，未写入：" + "、".join(sorted(ambiguous_texts)))
    if invalid_fld_pos:
        summary.append("FLD PO 缺少有效总saving/订单净值，保留价格表结果：" + "、".join(sorted(invalid_fld_pos)))
    if ambiguous_fld_pos:
        summary.append("FLD PO 汇总值不一致，保留价格表结果：" + "、".join(sorted(ambiguous_fld_pos)))
    if download_missing_pos:
        summary.append(
            "未找到任何下载目录，四列未修改："
            + "、".join(sorted(download_missing_pos))
        )
    if download_scan_error_pos:
        summary.append(
            "附件目录扫描失败，四列未修改："
            + "、".join(sorted(download_scan_error_pos))
        )
        for po_number in sorted(download_scan_error_pos):
            errors = download_scan_errors.get(po_number) or []
            if errors:
                summary.append(
                    f"  · {po_number}：{'; '.join(str(error) for error in errors[:2])}"
                )
    if non_fld_saving_skipped:
        summary.append(
            f"无 FLD Saving 保持原值：{non_fld_saving_skipped} 条缺少有效老价格、净价或采购订单数量"
        )
    if check_skipped:
        summary.append(f"CHECK 保持原值：{check_skipped} 条缺少有效新价格或净价")
    summary.append(f"价格表：{os.path.basename(price_file)} / {price_list_sheet}")
    return True, "\n".join(summary)


def fill_air_conditioning_old_new_prices(price_file, status_callback=None):
    """兼容既有空调调用入口。"""
    return fill_category_old_new_prices("空调", price_file, status_callback)


def generate_pivot_table(category=None, status_callback=None):
    """
    品类驱动的完整工作流：
      1. 按品类配置：插入品类专属列（差异化预处理）
      2. pandas 筛选物料 → 写入品类数据 Sheet
      3. COM 生成原生透视表 → 品类透视表 Sheet（Sheet1 右侧）
    """
    _log = lambda msg: status_callback and status_callback(msg)
    cfg = _cfg(category)
    label = cfg["label"]
    filter_materials = cfg["filter_materials"]
    data_sheet = cfg["data_sheet"]
    is_data_price_workflow = cfg.get("workflow") == "data_price"
    is_lighting_saving_workflow = cfg.get("workflow") == "lighting_saving"
    target_sheet = cfg.get("target_sheet")
    pivot_table_name = cfg.get("pivot_table_name")
    required_source_columns = cfg.get("required_source_columns", CHECK_COLS)
    pivot_row_fields = cfg.get("pivot_row_fields", ROW_FIELDS)
    pivot_value_field = cfg.get("pivot_value_field", VALUE_FIELD)
    pivot_value_name = cfg.get("pivot_value_name", VALUE_NAME)

    # ── 1. 检查文件 ──
    if not os.path.exists(EXCEL_FILE):
        return False, f"找不到主数据文件：\n{EXCEL_FILE}"

    # ── 2. 校验 Sheet1 列名 ──
    _log(f"[{label}] 正在校验列名...")
    try:
        import pandas as pd
        df_head = pd.read_excel(EXCEL_FILE, sheet_name=SOURCE_SHEET, nrows=0)
        missing = [c for c in required_source_columns if c not in df_head.columns]
        if missing:
            return False, (
                f"\"{SOURCE_SHEET}\" 中缺少以下列：\n"
                f"{', '.join(missing)}\n\n"
                f"当前列名：{list(df_head.columns)}"
            )
    except ValueError:
        return False, f"工作簿中不存在名为 \"{SOURCE_SHEET}\" 的 Sheet。"
    except Exception as e:
        return False, f"读取 Excel 失败：\n{e}"

    # ── 3. 增强 Sheet1 + 筛选 → 写入品类数据 Sheet ──
    try:
        total_cols, total_rows, match_debug = _enhance_and_filter(cfg, _log)
    except Exception as e:
        return False, f"数据增强/筛选失败：\n{e}"

    # 线槽这一类没有透视表：筛选结果本身就是交付数据，后续只做价格匹配。
    # 这里提前返回，确保不会启动 Excel COM，也不会意外创建空的透视表 Sheet。
    if is_data_price_workflow:
        return True, (
            f"[{label}] 数据生成成功！\n\n"
            f"筛选物料：{filter_materials}\n"
            f"筛选后数据行数：{total_rows}\n"
            f"目标 Sheet：\"{data_sheet}\"\n"
            f"已新增列：老价格 / 新价格 / Saving\n\n"
            f"本品类不生成透视表；请继续执行②，匹配价格并计算 Saving。\n\n"
            f"════════════ 诊断信息 ════════════\n{match_debug}"
        )

    # ── 4. COM 生成原生透视表 ──
    _log(f"[{label}] 正在调用 Excel 引擎生成原生透视表...")
    excel = None
    wb = None
    try:
        excel = _com_start()
        _log(f"[{label}] COM 打开增强后的文件...")
        wb = excel.Workbooks.Open(EXCEL_FILE)
        if wb is None:
            return False, "COM 无法打开文件，请关闭其他正在使用该文件的程序后重试。"

        created_sheet, _ = _com_create_pivot(
            wb, total_cols, total_rows,
            source_name=data_sheet,
            target_sheet=target_sheet,
            pivot_table_name=pivot_table_name,
            row_fields=pivot_row_fields,
            value_field=pivot_value_field,
            value_name=pivot_value_name,
        )
        if is_lighting_saving_workflow:
            _apply_lighting_saving_layout(created_sheet)

        _log("正在保存...")
        wb.Save()
        wb.Close(SaveChanges=True)
        wb = None
        time.sleep(0.3)
    except Exception as e:
        return False, f"COM 透视表生成失败：\n{e}"
    finally:
        _com_stop(excel, wb)

    summary = (
        f"[{label}] 原生透视表生成成功！\n\n"
        f"筛选物料：{filter_materials}\n"
        f"筛选后数据行数：{total_rows}\n"
        f"行层级：{' → '.join(pivot_row_fields)}\n"
        f"值字段：{pivot_value_name}（求和）\n"
        f"目标 Sheet：\"{target_sheet}\"（紧邻 Sheet1 右侧）\n\n"
        f"Sheet1 原有格式完整保留，仅新增了品类相关列。\n"
    )
    if is_lighting_saving_workflow:
        summary += "右侧已写入井道照明固定规则、实际梯台数与 Saving 总额公式。\n"
    summary += f"\n════════════ 诊断信息 ════════════\n{match_debug}"
    return True, summary


# ═══════════════════ GUI 界面 ═══════════════════

WINDOW_DEFAULT_WIDTH = 760
WINDOW_DEFAULT_HEIGHT = 780
WINDOW_SAFE_MARGIN = 32


def fit_window_geometry(screen_width, screen_height,
                        desired_width=WINDOW_DEFAULT_WIDTH,
                        desired_height=WINDOW_DEFAULT_HEIGHT,
                        margin=WINDOW_SAFE_MARGIN):
    """返回始终落在可视屏幕内的初始窗口尺寸与坐标。

    Windows 高 DPI 缩放时，Tk 取得的逻辑屏幕高度可能小于默认窗口高度。
    先收缩窗口、再留出四周安全边距，避免标题栏和系统关闭按钮被放到屏幕外。
    """
    screen_width = max(1, int(screen_width))
    screen_height = max(1, int(screen_height))
    margin = max(0, int(margin))
    max_width = max(1, screen_width - margin * 2)
    max_height = max(1, screen_height - margin * 2)
    width = min(max(1, int(desired_width)), max_width)
    height = min(max(1, int(desired_height)), max_height)
    x = max(0, (screen_width - width) // 2)
    y = max(0, (screen_height - height) // 2)
    return width, height, x, y


class PivotTableApp:
    """Tkinter 主窗口 — 支持装潢、外包板、空调、IC卡、监控、LCD多品类切换。"""

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("采购自动化工作台")
        # 显式保留系统标题栏，用户可通过它自由拖动窗口和关闭程序。
        self.root.overrideredirect(False)
        self.root.resizable(True, True)
        self.root.configure(bg="#eef3f8")

        self.active_category = "装潢"  # 默认品类
        self._web_stop_event = None

        # 高 DPI/低分辨率屏幕下，默认 760×780 可能高于逻辑屏幕；
        # 必须先收缩到可视区域内，否则标题栏会被挤到屏幕上方。
        self.root.update_idletasks()
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        width, height, x, y = fit_window_geometry(sw, sh)
        self.root.minsize(min(640, width), min(520, height))
        self.root.geometry(f"{width}x{height}+{x}+{y}")
        self.root.protocol("WM_DELETE_WINDOW", self._on_close_request)

        self._build_ui()

    def _on_close_request(self):
        """提供界面内关闭入口，并在任务运行时提醒用户。"""
        task_buttons = (
            getattr(self, "btn_factory", None),
            getattr(self, "btn", None),
            getattr(self, "btn_extra", None),
            getattr(self, "btn_web", None),
            getattr(self, "btn_retry_missing", None),
            getattr(self, "btn_backfill", None),
            getattr(self, "btn_tracking", None),
            getattr(self, "btn_air_price", None),
        )
        is_busy = any(
            button is not None and str(button.cget("state")) == str(tk.DISABLED)
            for button in task_buttons
        )
        if is_busy and not messagebox.askyesno(
            "关闭工作台",
            "当前任务仍在运行。关闭工作台会退出本程序；已打开的 Edge 浏览器可能仍会保留。是否继续？",
            parent=self.root,
        ):
            return
        self.root.destroy()

    def _is_content_widget(self, widget):
        """判断事件来源是否位于可滚动的工作台正文中。"""
        while widget is not None:
            if widget is self.content_canvas:
                return True
            widget = getattr(widget, "master", None)
        return False

    def _on_content_mousewheel(self, event):
        """Windows 鼠标滚轮：仅在正文区域滚动。"""
        if not self._is_content_widget(event.widget):
            return
        delta = getattr(event, "delta", 0)
        if not delta:
            return
        units = max(1, abs(int(delta)) // 120)
        self.content_canvas.yview_scroll(-units if delta > 0 else units, "units")

    # ── 品类切换 ──
    def _switch_category(self, category):
        self.active_category = category
        self._update_category_buttons()
        self._update_button_labels()
        self._update_hint_labels()
        self._apply_category_workflow_ui()
        self._update_workflow_guidance()
        workflow = CATEGORIES[category].get("workflow")
        steps = "①" if workflow == "lighting_saving" else "① → ②" if workflow == "data_price" else "① → ② → ③ → ④ → ⑤"
        self._update_status(
            f"已切换到「{CATEGORIES[category]['label']}」模式 — 按 {steps} 依次完成",
            "#3b82f6"
        )

    def _update_category_buttons(self):
        for cat, btn in self.cat_buttons.items():
            btn.config(
                bootstyle="primary" if cat == self.active_category else "secondary-outline"
            )

    def _update_button_labels(self):
        cfg = CATEGORIES[self.active_category]
        label = cfg["label"]
        if cfg.get("workflow") == "lighting_saving":
            self.btn.config(text=f"① 生成{label}数据与Saving")
            return
        if cfg.get("workflow") == "data_price":
            self.btn.config(text=f"① 生成{label}数据")
            self.btn_air_price.config(text=f"② 填充{label}价格并计算 Saving")
            return
        self.btn.config(text=f"① 生成{label}透视表")
        self.btn_extra.config(text=f"② 添加扩展列")
        self.btn_web.config(text="③ 打开网站查询")
        self.btn_retry_missing.config(text="↻ 补查缺失 PO")
        self.btn_backfill.config(text="④ 回填已下载文件")
        self.btn_tracking.config(text="⑤ 匹配 PM Tracking")
        if hasattr(self, "btn_air_price") and CATEGORIES[self.active_category].get("price_list_sheet"):
            self.btn_air_price.config(text=f"⑥ 填充{label}老/新价格")

    def _update_hint_labels(self):
        cfg = CATEGORIES[self.active_category]
        label = cfg["label"]
        materials = cfg["filter_materials"]
        if cfg.get("workflow") == "lighting_saving":
            self.hint1.config(
                text=f"筛选物料 {materials} → 新增描述 / PO数量 → 生成「{cfg['target_sheet']}」原生透视表与 Saving 计算"
            )
            return
        if cfg.get("workflow") == "data_price":
            self.hint1.config(
                text=f"筛选物料 {materials} → 生成「{cfg['data_sheet']}」；本品类不创建透视表"
            )
            self.hint_air_price.config(
                text="短文本精确匹配价格表 E 列 SAP Discription-Ner；H 列取老价格、J 列取新价格；Saving =（老价格 - 净价）× 采购订单数量"
            )
            return
        self.hint1.config(text=f"筛选物料 {materials} → COM 引擎生成原生透视表")
        self.hint2.config(text=f"在{label}透视表右侧追加 E2E项目名 / WBS / Price / PlanCost 等 10 列空白表头")
        self.hint3.config(
            text="③ 首次查询下载全部 PO；若日志提示缺失，点击右侧「补查缺失 PO」仅重查缺失项（也可扫描旧运行结果）。"
        )
        self.hint_backfill.config(text=f"读取「{label}」已下载附件 → 解析审批价格 / 计算差异 → 写回{label}透视表")
        content_display = cfg.get("content_filter_display", cfg["content_filter"])
        self.hint4.config(text=f"E2E项目名模糊匹配 → Content 筛选「{content_display}」→ 计算总 saving / 订单是否下完")

    def _update_workflow_guidance(self):
        """仅更新当前品类标识，不改变任何任务逻辑。"""
        label = CATEGORIES[self.active_category]["label"]
        self.lbl_workflow_category.config(text=f"当前工作流：{label}")

    def _apply_category_workflow_ui(self):
        """按品类工作流显示对应按钮，线槽不暴露无关的网页/透视表操作。"""
        cfg = CATEGORIES[self.active_category]
        is_data_price = cfg.get("workflow") == "data_price"
        is_lighting_saving = cfg.get("workflow") == "lighting_saving"

        # 每次先收起再按固定顺序重新布局，避免来回切换后控件顺序错乱。
        for widget in (
            self.btn_factory, self.hint_factory, self.c1_sep_factory,
            self.btn, self.hint1, self.c1_sep_main,
            self.btn_extra, self.hint2, self.c1_bottom_spacer,
            self.web_action_row, self.hint3, self.c2_sep_web,
            self.btn_backfill, self.hint_backfill, self.c2_sep_backfill,
            self.btn_tracking, self.hint4, self.c2_bottom_spacer,
            self.air_price_frame,
        ):
            widget.pack_forget()

        if is_lighting_saving:
            self.workflow_outer.pack_forget()
            self.c2_outer.pack_forget()
            self.c1_title.config(text=f"01  生成{cfg['label']}数据与Saving")
            self.btn.pack(fill=tk.X, padx=20, pady=(4, 0))
            self.hint1.pack(anchor=tk.W, padx=22, pady=(2, 0))
            self.c1_bottom_spacer.pack()
            return

        if not self.c2_outer.winfo_manager():
            self.c2_outer.pack(
                fill=tk.X, padx=24, pady=(12, 0), before=self.log_frame,
            )

        if is_data_price:
            self.workflow_outer.pack_forget()
            self.c1_title.config(text=f"01  生成{cfg['label']}数据")
            self.btn.pack(fill=tk.X, padx=20, pady=(4, 0))
            self.hint1.pack(anchor=tk.W, padx=22, pady=(2, 0))
            self.c1_bottom_spacer.pack()

            self.c2_title.config(text="02  价格匹配与 Saving")
            self.air_price_frame.pack(fill=tk.X)
            self.c2_bottom_spacer.pack()
            return

        if not self.workflow_outer.winfo_manager():
            self.workflow_outer.pack(
                fill=tk.X, padx=24, pady=(18, 0), before=self.c1_outer,
            )
        self.c1_title.config(text="01  数据准备")
        self.btn_factory.pack(fill=tk.X, padx=20, pady=(4, 0))
        self.hint_factory.pack(anchor=tk.W, padx=22, pady=(2, 0))
        self.c1_sep_factory.pack(fill=tk.X, padx=20, pady=8)
        self.btn.pack(fill=tk.X, padx=20, pady=(4, 0))
        self.hint1.pack(anchor=tk.W, padx=22, pady=(2, 0))
        self.c1_sep_main.pack(fill=tk.X, padx=20, pady=8)
        self.btn_extra.pack(fill=tk.X, padx=20, pady=(4, 0))
        self.hint2.pack(anchor=tk.W, padx=22, pady=(2, 0))
        self.c1_bottom_spacer.pack()

        self.c2_title.config(text="02  查询下载与回填核对")
        self.web_action_row.pack(fill=tk.X, padx=20, pady=(4, 0))
        self.hint3.pack(anchor=tk.W, padx=22, pady=(2, 0))
        self.c2_sep_web.pack(fill=tk.X, padx=20, pady=8)
        self.btn_backfill.pack(fill=tk.X, padx=20, pady=(4, 0))
        self.hint_backfill.pack(anchor=tk.W, padx=22, pady=(2, 0))
        self.c2_sep_backfill.pack(fill=tk.X, padx=20, pady=8)
        self.btn_tracking.pack(fill=tk.X, padx=20, pady=(4, 0))
        self.hint4.pack(anchor=tk.W, padx=22, pady=(2, 0))
        self.c2_bottom_spacer.pack()
        if cfg.get("price_list_sheet"):
            self.air_price_frame.pack(fill=tk.X)

    def _build_ui(self):
        WIN_BG = "#eef3f8"
        CARD_BG = "#ffffff"
        HINT_FG = "#66788a"
        TITLE_FG = "#172b4d"
        CARD_BORDER = "#d9e2ec"

        # ── 顶部标题栏 ──
        category_buttons_per_row = 4
        category_rows = (len(CATEGORIES) + category_buttons_per_row - 1) // category_buttons_per_row
        header_height = 180 + max(0, category_rows - 2) * 34
        header = tk.Frame(self.root, bg="#102a43", height=header_height)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        # 标题与品类切换占上行，窗口控制固定在下行。这样窗口变窄时不会让
        # “停止查询”被标题、品类切换按钮挤出可视区域。
        header_main = tk.Frame(header, bg="#102a43")
        header_main.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        header_actions = tk.Frame(header, bg="#0b2239", height=38)
        header_actions.pack(side=tk.BOTTOM, fill=tk.X)
        header_actions.pack_propagate(False)

        # 品类较多时使用两行均分布局；窄窗口也不会把最后一个入口或品牌标志
        # 挤出可视范围。每个按钮仍是独立可点击的工作台切换入口。
        cat_frame = tk.Frame(header_main, bg="#102a43")
        cat_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=16, pady=(0, 6))
        self.cat_buttons = {}
        categories_order = list(CATEGORIES.keys())
        for column in range(category_buttons_per_row):
            cat_frame.grid_columnconfigure(column, weight=1)
        for i, cat_key in enumerate(categories_order):
            cat_cfg = CATEGORIES[cat_key]
            is_active = (cat_key == self.active_category)
            btn = ttk.Button(
                cat_frame, text=f"  {cat_cfg['label']}  ",
                bootstyle="primary" if is_active else "secondary-outline",
                padding=(8, 5),
                command=lambda c=cat_key: self._switch_category(c),
            )
            btn.grid(
                row=i // category_buttons_per_row,
                column=i % category_buttons_per_row,
                sticky=tk.EW,
                padx=2,
                pady=2,
            )
            self.cat_buttons[cat_key] = btn

        brand = tk.Frame(header_main, bg="#102a43")
        brand.pack(side=tk.TOP, anchor=tk.W, padx=(26, 0), pady=(8, 0))
        self.tke_header_logo = _load_tke_logo(
            self.root, max_width=172, max_height=48,
        )
        self.lbl_tke_logo = tk.Label(
            brand, image=self.tke_header_logo, bg="#102a43",
        )
        self.lbl_tke_logo.pack(side=tk.LEFT)

        # 窗口控制始终独占一行，确保缩小时三个关键按钮都不会被裁掉。
        window_controls = tk.Frame(header_actions, bg="#0b2239")
        window_controls.pack(side=tk.RIGHT, padx=(0, 20), pady=4)
        self.btn_close = ttk.Button(
            window_controls, text="关闭工作台",
            bootstyle="danger", padding=(12, 6),
            command=self._on_close_request,
        )
        self.btn_close.pack(side=tk.LEFT)
        self.btn_stop_web = ttk.Button(
            window_controls, text="⏹ 停止查询",
            bootstyle="danger-outline", padding=(12, 6),
            command=self._on_stop_web_click, state=tk.DISABLED,
        )
        self.btn_stop_web.pack(side=tk.LEFT, padx=(6, 0))

        # 内容区可滚动：小屏或高 DPI 缩放时不会再为了容纳所有卡片把窗口
        # 顶出屏幕，窗口依然可以任意拖动和缩放。
        content_shell = tk.Frame(self.root, bg=WIN_BG)
        content_shell.pack(fill=tk.BOTH, expand=True)
        self.content_canvas = tk.Canvas(
            content_shell, bg=WIN_BG, highlightthickness=0, borderwidth=0,
        )
        content_scrollbar = tk.Scrollbar(
            content_shell, orient=tk.VERTICAL, command=self.content_canvas.yview,
        )
        self.content_canvas.configure(yscrollcommand=content_scrollbar.set)
        content_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.content_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        content = tk.Frame(self.content_canvas, bg=WIN_BG)
        self._content_window = self.content_canvas.create_window(
            (0, 0), window=content, anchor=tk.NW,
        )

        def _refresh_scroll_region(_event=None):
            self.content_canvas.configure(scrollregion=self.content_canvas.bbox("all"))

        def _fit_content_width(event):
            self.content_canvas.itemconfigure(self._content_window, width=event.width)
            self.content_canvas.after_idle(_refresh_scroll_region)

        content.bind("<Configure>", _refresh_scroll_region)
        self.content_canvas.bind("<Configure>", _fit_content_width)
        # bind_all 让鼠标位于卡片、文字或按钮上方时同样可以滚动；处理函数会
        # 限定事件来源，因此不影响页头、弹窗等其他区域。
        self.root.bind_all("<MouseWheel>", self._on_content_mousewheel, add="+")

        # ── 卡片样式工厂 ──
        def _card(parent, title):
            """创建一张卡片：白色背景 + 细边框 + 标题"""
            outer = tk.Frame(parent, bg=CARD_BORDER, bd=0)
            inner = tk.Frame(outer, bg=CARD_BG, bd=0)
            inner.pack(padx=1, pady=1, fill=tk.BOTH, expand=True)
            title_row = tk.Frame(inner, bg=CARD_BG)
            title_row.pack(fill=tk.X, padx=20, pady=(13, 4))
            title_label = tk.Label(
                title_row, text=title,
                font=("Microsoft YaHei", 11, "bold"),
                fg=TITLE_FG, bg=CARD_BG,
            )
            title_label.pack(side=tk.LEFT)
            inner._card_title_label = title_label
            return outer, inner

        def _btn(parent, text, bootstyle, command):
            """创建统一风格的按钮"""
            btn = ttk.Button(
                parent, text=text,
                bootstyle=bootstyle, padding=(12, 9),
                command=command,
            )
            btn.pack(fill=tk.X, padx=20, pady=(4, 0))
            return btn

        def _hint(parent, text):
            """按钮下方的说明文字"""
            lbl = tk.Label(
                parent, text=text,
                font=("Microsoft YaHei", 8), fg=HINT_FG, bg=CARD_BG,
                justify=tk.LEFT, wraplength=680,
            )
            lbl.pack(anchor=tk.W, padx=22, pady=(2, 0))
            return lbl

        def _sep(parent):
            """细分隔线"""
            separator = ttk.Separator(parent, orient=tk.HORIZONTAL)
            separator.pack(fill=tk.X, padx=20, pady=8)
            return separator

        # ══════════ 工作流概览：仅帮助用户理解顺序，不参与任何业务判断 ══════════
        cfg_default = CATEGORIES[self.active_category]
        workflow_outer, workflow = _card(content, "推荐流程")
        self.workflow_outer = workflow_outer
        workflow_outer.pack(fill=tk.X, padx=24, pady=(18, 0))

        workflow_header = tk.Frame(workflow, bg=CARD_BG)
        workflow_header.pack(fill=tk.X, padx=20, pady=(2, 6))
        tk.Label(
            workflow_header, text="按顺序完成，减少遗漏与重复查询",
            font=("Microsoft YaHei", 9), fg=HINT_FG, bg=CARD_BG,
        ).pack(side=tk.LEFT)
        self.lbl_workflow_category = tk.Label(
            workflow_header, text=f"当前工作流：{cfg_default['label']}",
            font=("Microsoft YaHei", 9, "bold"), fg="#0d6efd", bg=CARD_BG,
        )
        self.lbl_workflow_category.pack(side=tk.RIGHT)

        workflow_steps = tk.Frame(workflow, bg=CARD_BG)
        workflow_steps.pack(fill=tk.X, padx=20, pady=(0, 7))
        step_specs = (
            ("01", "数据准备", "生成透视表 · 添加扩展列", "#0d6efd"),
            ("02", "查询下载", "网站查询 · 补查缺失 PO", "#0dcaf0"),
            ("03", "回填核对", "回填附件 · PM Tracking", "#198754"),
        )
        for index, (number, title, detail, color) in enumerate(step_specs):
            step = tk.Frame(workflow_steps, bg="#f8fafc", highlightbackground="#d9e2ec", highlightthickness=1)
            step.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0 if index == 0 else 4, 0))
            tk.Label(
                step, text=number, font=("Microsoft YaHei", 9, "bold"),
                fg="#ffffff", bg=color, width=4, pady=5,
            ).pack(side=tk.LEFT, padx=(7, 7), pady=7)
            step_text = tk.Frame(step, bg="#f8fafc")
            step_text.pack(side=tk.LEFT, fill=tk.X, expand=True, pady=6)
            tk.Label(
                step_text, text=title, font=("Microsoft YaHei", 9, "bold"),
                fg=TITLE_FG, bg="#f8fafc", anchor=tk.W,
            ).pack(fill=tk.X)
            tk.Label(
                step_text, text=detail, font=("Microsoft YaHei", 8),
                fg=HINT_FG, bg="#f8fafc", anchor=tk.W,
            ).pack(fill=tk.X)

        self.lbl_workflow_path = tk.Label(
            workflow,
            text="01 数据准备  →  02 查询下载  →  03 回填核对",
            font=("Microsoft YaHei", 8, "bold"), fg="#52677d", bg=CARD_BG,
            anchor=tk.W,
        )
        self.lbl_workflow_path.pack(fill=tk.X, padx=22)
        tk.Frame(workflow, bg=CARD_BG, height=8).pack()

        # ══════════ 卡片 1：Excel 数据处理 ══════════
        c1_outer, c1 = _card(content, "01  数据准备")
        self.c1_outer = c1_outer
        self.c1_title = c1._card_title_label
        c1_outer.pack(fill=tk.X, padx=24, pady=(12, 0))

        # ── 独立按钮：匹配区域/分公司（品类无关）──
        self.btn_factory = _btn(
            c1, "🔧 匹配区域/分公司", "warning", self._on_factory_click,
        )
        self.hint_factory = _hint(c1, "读取工厂清单 → Sheet1 插入区域/分公司列 → 按 Plant 填值（跑一次即可）")

        self.c1_sep_factory = _sep(c1)

        self.btn = _btn(
            c1, f"① 生成{cfg_default['label']}透视表", "success", self._on_click,
        )
        self.hint1 = _hint(c1, f"筛选物料 {cfg_default['filter_materials']} → COM 引擎生成原生透视表")

        self.c1_sep_main = _sep(c1)

        self.btn_extra = _btn(
            c1, "② 添加扩展列", "primary", self._on_extra_click,
        )
        self.hint2 = _hint(c1, f"在{cfg_default['label']}透视表右侧追加 E2E项目名 / WBS / Price / PlanCost 等 10 列空白表头")
        self.c1_bottom_spacer = tk.Frame(c1, bg=CARD_BG, height=10)
        self.c1_bottom_spacer.pack()

        # ══════════ 卡片 2：网站查询与匹配 ══════════
        c2_outer, c2 = _card(content, "02  查询下载与回填核对")
        self.c2_outer = c2_outer
        self.c2_title = c2._card_title_label
        c2_outer.pack(fill=tk.X, padx=24, pady=(12, 0))

        # ③ 与“补查缺失 PO”并列：两者属于同一网页下载步骤，避免用户误以为
        # 补查是另一个独立流程。
        web_action_row = tk.Frame(c2, bg=CARD_BG)
        web_action_row.pack(fill=tk.X, padx=20, pady=(4, 0))
        self.web_action_row = web_action_row
        self.btn_web = ttk.Button(
            web_action_row, text="③ 打开网站查询",
            bootstyle="info", padding=(12, 9),
            command=self._on_web_click,
        )
        self.btn_web.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))
        self.btn_retry_missing = ttk.Button(
            web_action_row, text="↻ 补查缺失 PO",
            bootstyle="secondary-outline", padding=(12, 9),
            command=self._on_retry_missing_click,
        )
        self.btn_retry_missing.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(4, 0))
        self.hint3 = _hint(
            c2,
            "③ 首次查询下载全部 PO；若日志提示缺失，点击右侧「补查缺失 PO」仅重查缺失项（也可扫描旧运行结果）。",
        )

        self.c2_sep_web = _sep(c2)

        self.btn_backfill = _btn(
            c2, "④ 回填已下载文件", "success", self._on_backfill_click,
        )
        self.hint_backfill = _hint(c2, "读取已下载附件 → 解析审批价格 / 计算差异 → 写回当前品类透视表")

        self.c2_sep_backfill = _sep(c2)

        self.btn_tracking = _btn(
            c2, "⑤ 匹配 PM Tracking", "info", self._on_tracking_click,
        )
        default_content_display = cfg_default.get("content_filter_display", cfg_default["content_filter"])
        self.hint4 = _hint(c2, f"E2E项目名模糊匹配 → Content 筛选「{default_content_display}」→ 计算 总saving / 订单是否下完")
        self.c2_bottom_spacer = tk.Frame(c2, bg=CARD_BG, height=10)
        self.c2_bottom_spacer.pack()

        # ── 日志输出区 ──
        self.air_price_frame = tk.Frame(c2, bg=CARD_BG)
        self.btn_air_price = _btn(
            self.air_price_frame, "⑥ 填充老/新价格", "primary", self._on_air_price_click,
        )
        self.hint_air_price = _hint(
            self.air_price_frame,
            "FLD PO 用透视表覆盖；无 FLD Saving =（老价格 - 净价）× 采购订单数量；CHECK = 新价格 - 净价",
        )

        log_frame = tk.Frame(content, bg="#102a43")
        self.log_frame = log_frame
        log_frame.pack(fill=tk.BOTH, expand=True, padx=24, pady=(12, 0))
        log_label = tk.Label(
            log_frame, text="运行日志", font=("Microsoft YaHei", 9, "bold"),
            fg="#b8d4eb", bg="#102a43", anchor=tk.W,
        )
        log_label.pack(fill=tk.X, padx=10, pady=(7, 2))
        self.log_text = tk.Text(
            log_frame, bg="#0b1f33", fg="#d8e7f3", font=("Consolas", 9),
            wrap=tk.WORD, relief=tk.FLAT, borderwidth=0,
            insertbackground="#d8e7f3", state=tk.DISABLED, padx=10, pady=8,
        )
        log_scroll = tk.Scrollbar(log_frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scroll.set)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 5))

        # 初次启动同样走统一布局，确保装潢默认视图和后续品类切换一致。
        self._apply_category_workflow_ui()

        # ── 底部状态栏 ──
        # 不固定高度：高 DPI 或中文字体行高变化时，状态文字仍须完整可见。
        status_bar = tk.Frame(self.root, bg="#102a43")
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        self.lbl_status = tk.Label(
            status_bar, text="就绪 — 请按 ① → ② → ③ → ④ → ⑤ 顺序操作",
            font=("Microsoft YaHei", 8), fg="#b8d4eb", bg="#102a43",
        )
        self.lbl_status.pack(side=tk.LEFT, padx=26, pady=6)
        self.lbl_workflow_state = tk.Label(
            status_bar, text="● 就绪 — 请从 ① 生成透视表开始",
            font=("Microsoft YaHei", 8, "bold"), fg="#27ae60", bg="#102a43",
        )
        self.lbl_workflow_state.pack(side=tk.RIGHT, padx=26, pady=6)

    def _on_factory_click(self):
        """点击匹配区域/分公司（品类无关）"""
        self.btn_factory.config(state=tk.DISABLED, text="正在匹配...")
        self._update_status("正在匹配区域/分公司...", "#f59e0b")

        def _run():
            try:
                success, msg = apply_factory_mapping(
                    status_callback=lambda m: self.root.after(0, self._update_status, m, "#f59e0b")
                )
                self.root.after(0, lambda: self._factory_done(success, msg))
            except Exception as e:
                err_msg = f"错误：\n{e}\n\n{traceback.format_exc()}"
                self.root.after(0, lambda: self._factory_done(False, err_msg))

        self.root.after(100, _run)

    def _factory_done(self, success, msg):
        self.btn_factory.config(state=tk.NORMAL, text="🔧 匹配区域/分公司")
        if success:
            self._update_status("区域/分公司匹配完成！", "#27ae60")
            self._show_copyable_dialog("✅ 工厂映射", msg)
        else:
            self._update_status("匹配失败，见弹窗", "#e74c3c")
            self._show_copyable_dialog("❌ 匹配失败", msg, is_error=True)

    def _on_click(self):
        """点击生成透视表"""
        category = self.active_category
        label = CATEGORIES[category]["label"]
        self.btn.config(state=tk.DISABLED, text="处理中，请稍候...")
        self._update_status(f"[{label}] 正在处理...", "#e67e22")

        def _run():
            try:
                success, msg = generate_pivot_table(
                    category=category,
                    status_callback=lambda m: self.root.after(0, self._update_status, m, "#e67e22")
                )
                self.root.after(0, lambda: self._done(success, msg))
            except Exception as e:
                err_msg = f"未预料的错误：\n{e}\n\n{traceback.format_exc()}"
                self.root.after(0, lambda: self._done(False, err_msg))

        self.root.after(100, _run)

    def _update_status(self, msg, color):
        self.lbl_status.config(text=msg, fg=color)
        self.lbl_workflow_state.config(text=f"● {msg}", fg=color)
        # 同时写入日志区
        ts = time.strftime("%H:%M:%S")
        self._log(f"{ts}  {msg}")

    def _log(self, msg):
        """写入日志区（线程安全）"""
        try:
            self.log_text.configure(state=tk.NORMAL)
            self.log_text.insert(tk.END, msg + "\n")
            self.log_text.see(tk.END)
            self.log_text.configure(state=tk.DISABLED)
        except Exception:
            pass

    def _show_copyable_dialog(self, title, msg, is_error=False):
        """显示可选中复制的完成提示，不锁定主工作台。"""
        dlg = tk.Toplevel(self.root)
        dlg.title(title)
        dialog_width, dialog_height = 620, 440
        self.root.update_idletasks()
        screen_x = self.root.winfo_vrootx()
        screen_y = self.root.winfo_vrooty()
        screen_right = screen_x + self.root.winfo_vrootwidth()
        screen_bottom = screen_y + self.root.winfo_vrootheight()
        x = self.root.winfo_rootx() + (self.root.winfo_width() - dialog_width) // 2
        y = self.root.winfo_rooty() + (self.root.winfo_height() - dialog_height) // 2
        x = max(screen_x, min(x, screen_right - dialog_width))
        y = max(screen_y, min(y, screen_bottom - dialog_height))
        dlg.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        dlg.minsize(460, 300)
        dlg.resizable(True, True)
        dlg.configure(bg="#1a2332")
        dlg.transient(self.root)

        # 完成提示必须始终是非模态的：即便窗口映射或焦点异常，主工作台
        # 仍可点击、移动和关闭，绝不能再被一个不可见的提示窗锁死。
        def _close_dialog():
            if dlg.winfo_exists():
                dlg.destroy()

        dlg.protocol("WM_DELETE_WINDOW", _close_dialog)
        dlg.bind("<Escape>", lambda _event: _close_dialog())
        # 标题
        tk.Label(
            dlg, text=title, font=("Microsoft YaHei", 12, "bold"),
            fg="#e74c3c" if is_error else "#27ae60", bg="#1a2332",
        ).pack(pady=(12, 0))
        # 可复制的文本框
        text_frame = tk.Frame(dlg, bg="#0a1018")
        text_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=8)
        text_widget = tk.Text(
            text_frame, bg="#0a1018", fg="#b0c0d0",
            font=("Consolas", 10), wrap=tk.WORD,
            relief=tk.FLAT, borderwidth=0, padx=10, pady=10,
        )
        text_widget.insert("1.0", msg)
        text_widget.configure(state=tk.DISABLED)
        scrollbar = tk.Scrollbar(text_frame, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        text_widget.pack(fill=tk.BOTH, expand=True)
        # 提示
        tk.Label(
            dlg, text="💡 可按 Ctrl+A 全选, Ctrl+C 复制内容",
            font=("Microsoft YaHei", 8), fg="#7b8ca0", bg="#1a2332",
        ).pack(pady=(0, 6))
        # 关闭按钮
        ttk.Button(
            dlg, text="关闭", bootstyle="primary", padding=(30, 6),
            command=_close_dialog,
        ).pack(pady=(0, 14))

        # 创建后立即提高层级并请求焦点；不使用永久置顶，避免妨碍用户继续
        # 操作 Excel 或浏览器。
        dlg.lift()
        dlg.focus_force()
        return dlg

    def _on_extra_click(self):
        """点击添加扩展列"""
        category = self.active_category
        label = CATEGORIES[category]["label"]
        self.btn_extra.config(state=tk.DISABLED, text="正在添加...")
        self._update_status(f"[{label}] 正在添加扩展列...", "#3498db")

        def _run():
            try:
                success, msg = add_extra_columns(category=category)
                self.root.after(0, lambda: self._extra_done(success, msg))
            except Exception as e:
                err_msg = f"未预料的错误：\n{e}\n\n{traceback.format_exc()}"
                self.root.after(0, lambda: self._extra_done(False, err_msg))

        self.root.after(100, _run)

    def _extra_done(self, success, msg):
        self.btn_extra.config(state=tk.NORMAL, text="② 添加扩展列")
        if success:
            self._update_status("扩展列添加成功！", "#27ae60")
            self._show_copyable_dialog("✅ 扩展列", msg)
        else:
            self._update_status("添加失败，见弹窗", "#e74c3c")
            self._show_copyable_dialog("❌ 操作失败", msg, is_error=True)

    def _on_stop_web_click(self):
        """请求在当前 PO 完成后的安全检查点停止网页查询。"""
        if self._web_stop_event is None or self._web_stop_event.is_set():
            return
        self._web_stop_event.set()
        self.btn_stop_web.config(
            state=tk.DISABLED, text="⏹ 正在停止...", bootstyle="danger-outline"
        )
        self._update_status("已请求停止：当前 PO 完成后将保存已下载文件并停止后续查询。", "#ef4444")

    def _on_web_click(self):
        """点击查询并下载附件按钮（不回填 Excel）。"""
        category = self.active_category
        label = CATEGORIES[category]["label"]
        stop_event = threading.Event()
        self._web_stop_event = stop_event
        self.btn_web.config(state=tk.DISABLED, text="正在查询并下载...")
        self.btn_retry_missing.config(state=tk.DISABLED, text="↻ 补查缺失 PO")
        self.btn_stop_web.config(
            state=tk.NORMAL, text="⏹ 停止查询", bootstyle="danger"
        )
        self._update_status(f"[{label}] 正在启动 Edge 查询附件...", "#6c63ff")

        def _run():
            try:
                import web_query
                target_sheet = CATEGORIES[category]["target_sheet"]
                category_label = CATEGORIES[category]["label"]
                success, msg = web_query.query_and_download_attachments(
                    target_sheet=target_sheet,
                    category_label=category_label,
                    status_callback=lambda m: self.root.after(0, self._update_status, m, "#6c63ff"),
                    stop_requested=stop_event,
                )
                self.root.after(0, lambda: self._web_done(success, msg, stop_event))
            except ImportError:
                self.root.after(0, lambda: self._web_done(
                    False, "缺少 playwright 模块，请先运行：\npip install playwright\nplaywright install msedge", stop_event
                ))
            except Exception as e:
                err_msg = f"未预料的错误：\n{e}\n\n{traceback.format_exc()}"
                self.root.after(0, lambda: self._web_done(False, err_msg, stop_event))

        threading.Thread(target=_run, daemon=True).start()

    def _web_done(self, success, msg, stop_event=None):
        if self._web_stop_event is stop_event:
            self._web_stop_event = None
        self.btn_web.config(state=tk.NORMAL, text="③ 打开网站查询")
        self.btn_retry_missing.config(state=tk.NORMAL, text="↻ 补查缺失 PO")
        self.btn_stop_web.config(
            state=tk.DISABLED, text="⏹ 停止查询", bootstyle="danger-outline"
        )
        stopped = bool(stop_event and stop_event.is_set()) or msg.startswith("⏹")
        if stopped:
            self._update_status("查询已停止；已完成的附件可直接回填。", "#f59e0b")
            self._show_copyable_dialog("⏹ 查询已停止", msg)
        elif success:
            self._update_status("查询与附件下载完成，可进行回填。", "#27ae60")
            self._show_copyable_dialog("✅ 查询与下载完成", msg)
        else:
            self._update_status("查询与下载失败，见弹窗", "#e74c3c")
            self._show_copyable_dialog("❌ 查询与下载失败", msg, is_error=True)

    def _on_retry_missing_click(self):
        """扫描当前品类已有目录，只重新查询确实缺失的 PO。"""
        category = self.active_category
        label = CATEGORIES[category]["label"]
        stop_event = threading.Event()
        self._web_stop_event = stop_event
        self.btn_web.config(state=tk.DISABLED, text="③ 打开网站查询")
        self.btn_retry_missing.config(state=tk.DISABLED, text="正在核对并补查...")
        self.btn_stop_web.config(
            state=tk.NORMAL, text="⏹ 停止查询", bootstyle="danger"
        )
        self._update_status(f"[{label}] 正在核对本地 PO 目录并补查缺失项...", "#475569")

        def _run():
            try:
                import web_query
                success, msg = web_query.retry_missing_pos(
                    target_sheet=CATEGORIES[category]["target_sheet"],
                    category_label=CATEGORIES[category]["label"],
                    status_callback=lambda m: self.root.after(0, self._update_status, m, "#475569"),
                    stop_requested=stop_event,
                )
                self.root.after(0, lambda: self._retry_missing_done(success, msg, stop_event))
            except ImportError:
                self.root.after(0, lambda: self._retry_missing_done(
                    False, "缺少 playwright 模块，请先运行：\npip install playwright\nplaywright install msedge", stop_event
                ))
            except Exception as e:
                err_msg = f"未预料的错误：\n{e}\n\n{traceback.format_exc()}"
                self.root.after(0, lambda: self._retry_missing_done(False, err_msg, stop_event))

        threading.Thread(target=_run, daemon=True).start()

    def _retry_missing_done(self, success, msg, stop_event=None):
        """恢复网页下载操作后的控件状态；和完整查询共用停止开关。"""
        if self._web_stop_event is stop_event:
            self._web_stop_event = None
        self.btn_web.config(state=tk.NORMAL, text="③ 打开网站查询")
        self.btn_retry_missing.config(state=tk.NORMAL, text="↻ 补查缺失 PO")
        self.btn_stop_web.config(
            state=tk.DISABLED, text="⏹ 停止查询", bootstyle="danger-outline"
        )
        stopped = bool(stop_event and stop_event.is_set()) or msg.startswith("⏹")
        if stopped:
            self._update_status("补查已停止；已完成的附件会保留。", "#f59e0b")
            self._show_copyable_dialog("⏹ 补查已停止", msg)
        elif success:
            self._update_status("缺失 PO 补查完成，可执行回填。", "#27ae60")
            self._show_copyable_dialog("✅ 缺失 PO 补查完成", msg)
        else:
            self._update_status("缺失 PO 补查未完成，见弹窗。", "#e74c3c")
            self._show_copyable_dialog("❌ 缺失 PO 补查未完成", msg, is_error=True)

    def _on_backfill_click(self):
        """点击回填按钮：只处理当前品类已下载的待回填任务。"""
        category = self.active_category
        label = CATEGORIES[category]["label"]
        self.btn_backfill.config(state=tk.DISABLED, text="正在解析并回填...")
        self._update_status(f"[{label}] 正在解析附件并回填 Excel...", "#0f9d8a")

        def _run():
            try:
                import web_query
                success, msg = web_query.backfill_downloaded_results(
                    target_sheet=CATEGORIES[category]["target_sheet"],
                    category_label=CATEGORIES[category]["label"],
                    status_callback=lambda m: self.root.after(0, self._update_status, m, "#0f9d8a")
                )
                self.root.after(0, lambda: self._backfill_done(success, msg))
            except Exception as e:
                err_msg = f"未预料的错误：\n{e}\n\n{traceback.format_exc()}"
                self.root.after(0, lambda: self._backfill_done(False, err_msg))

        threading.Thread(target=_run, daemon=True).start()

    def _backfill_done(self, success, msg):
        self.btn_backfill.config(state=tk.NORMAL, text="④ 回填已下载文件")
        if success:
            self._update_status("Excel 回填完成！", "#27ae60")
            self._show_copyable_dialog("✅ 回填完成", msg)
        else:
            self._update_status("回填未完成，任务已保留", "#e74c3c")
            self._show_copyable_dialog("❌ 回填未完成", msg, is_error=True)

    def _on_tracking_click(self):
        """点击 PM Tracking 匹配按钮"""
        category = self.active_category
        label = CATEGORIES[category]["label"]
        self.btn_tracking.config(state=tk.DISABLED, text="正在匹配...")
        self._update_status(f"[{label}] 正在匹配 NI PM Saving Tracking...", "#16a085")

        def _run():
            try:
                success, msg = match_pm_tracking_data(category=category)
                self.root.after(0, lambda: self._tracking_done(success, msg))
            except Exception as e:
                err_msg = f"错误：\n{e}\n\n{traceback.format_exc()}"
                self.root.after(0, lambda: self._tracking_done(False, err_msg))

        threading.Thread(target=_run, daemon=True).start()

    def _tracking_done(self, success, msg):
        self.btn_tracking.config(state=tk.NORMAL, text="⑤ 匹配 PM Tracking")
        if success:
            self._update_status("PM Tracking 匹配完成！", "#27ae60")
            self._show_copyable_dialog("✅ PM匹配报告", msg)
        else:
            self._update_status("匹配失败，见弹窗", "#e74c3c")
            self._show_copyable_dialog("❌ 匹配失败", msg, is_error=True)

    def _on_air_price_click(self):
        """选择价格表后，执行当前品类对应的价格填充流程。"""
        category = self.active_category
        cfg = CATEGORIES[category]
        if not cfg.get("price_list_sheet"):
            return
        label = cfg["label"]

        price_file = filedialog.askopenfilename(
            parent=self.root,
            title=f"选择{label} FY25/26 价格表",
            initialdir=SCRIPT_DIR,
            filetypes=[("Excel 文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")],
        )
        if not price_file:
            self._update_status(f"未选择价格表，未执行{label}老/新价格填充。", "#f59e0b")
            return

        is_data_price = cfg.get("workflow") == "data_price"
        busy_text = "正在匹配价格并计算 Saving..." if is_data_price else "正在填充价格 / Saving / CHECK..."
        status_text = (
            f"[{label}] 正在按短文本匹配老/新价格，并计算 Saving..."
            if is_data_price
            else f"[{label}] 正在填充基础价格、覆盖 FLD PO，并计算 CHECK..."
        )
        self.btn_air_price.config(state=tk.DISABLED, text=busy_text)
        self._update_status(status_text, "#7c3aed")

        def _run():
            try:
                success, msg = fill_category_old_new_prices(
                    category,
                    price_file,
                    status_callback=lambda m: self.root.after(
                        0, self._update_status, m, "#7c3aed"
                    ),
                )
                self.root.after(0, lambda: self._air_price_done(category, success, msg))
            except Exception as exc:
                err_msg = f"错误：\n{exc}\n\n{traceback.format_exc()}"
                self.root.after(0, lambda: self._air_price_done(category, False, err_msg))

        threading.Thread(target=_run, daemon=True).start()

    def _air_price_done(self, category, success, msg):
        cfg = CATEGORIES[category]
        label = cfg["label"]
        active_cfg = CATEGORIES[self.active_category]
        active_label = active_cfg["label"]
        button_text = (
            f"② 填充{active_label}价格并计算 Saving"
            if active_cfg.get("workflow") == "data_price"
            else f"⑥ 填充{active_label}老/新价格"
        )
        self.btn_air_price.config(state=tk.NORMAL, text=button_text)
        if success:
            complete_text = (
                f"{label}老/新价格与 Saving 填充完成！"
                if cfg.get("workflow") == "data_price"
                else f"{label}基础价格、FLD PO Saving 与 CHECK 填充完成！"
            )
            self._update_status(complete_text, "#27ae60")
            self._show_copyable_dialog(f"✅ {label}价格填充报告", msg)
        else:
            self._update_status(f"{label}价格填充失败，见弹窗", "#e74c3c")
            self._show_copyable_dialog(f"❌ {label}价格填充失败", msg, is_error=True)

    def _done(self, success, msg):
        cfg = CATEGORIES[self.active_category]
        label = cfg["label"]
        if cfg.get("workflow") == "lighting_saving":
            button_text = f"① 生成{label}数据与Saving"
        elif cfg.get("workflow") == "data_price":
            button_text = f"① 生成{label}数据"
        else:
            button_text = f"① 生成{label}透视表"
        self.btn.config(state=tk.NORMAL, text=button_text)
        if success:
            self._update_status(f"[{label}] 生成成功！", "#27ae60")
            self._show_copyable_dialog("✅ 操作成功", msg)
        else:
            self._update_status(f"[{label}] 生成失败，见弹窗", "#e74c3c")
            self._show_copyable_dialog("❌ 操作失败", msg, is_error=True)


# ═══════════════════ 入口 ═══════════════════

def _load_tke_logo(master, max_width=None, max_height=None):
    """加载项目内置 TKE 标志；缺失资源时保持工作台可启动。"""
    try:
        image = tk.PhotoImage(master=master, file=TKE_LOGO_PATH)
    except (OSError, tk.TclError):
        return None

    if not max_width and not max_height:
        return image
    width_scale = (image.width() + max_width - 1) // max_width if max_width else 1
    height_scale = (image.height() + max_height - 1) // max_height if max_height else 1
    scale = max(1, width_scale, height_scale)
    return image.subsample(scale, scale)


def create_workbench_root():
    """创建带统一 ttkbootstrap 主题的工作台窗口。"""
    root = ttk.Window(themename=WORKBENCH_THEME)
    root._tke_window_icon = _load_tke_logo(root)
    if root._tke_window_icon is not None:
        root.iconphoto(True, root._tke_window_icon)
    return root


def main():
    root = create_workbench_root()
    app = PivotTableApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
