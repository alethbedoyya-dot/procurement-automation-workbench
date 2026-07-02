# -*- coding: utf-8 -*-
"""
TKE VIEW 网站自动化查询模块
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
功能：
  1. 从「装潢透视表」Sheet 提取全部采购凭证号
  2. Playwright 启动 Edge → 登录 TKE VIEW（Microsoft SSO，会话持久化）
  3. 逐个 PO：跳转搜索页 → 填PO号 → 查找 → 点击请求ID → 详情页
  4. 详情页提取：项目名称 / 梯台明细合计数量 / 下载支持文件
  5. PDF 解析「总计（不含税）」→ 计算差异 → 写回 Excel

会话管理：
  → 首次运行需手动完成 Microsoft 账户登录
  → 登录后自动保存 cookies + localStorage 到 auth.json
  → 后续运行加载 auth.json，跳过登录
"""

import os
import re
import json
import sys
import time
import subprocess
import traceback

# ═══════════════════ 配置 ═══════════════════

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(SCRIPT_DIR, "EXPORT1 2 1.xlsx")
TARGET_SHEET = "装潢透视表"
HOME_URL = "https://view.tkelevator.com.cn/"
SEARCH_URL = "https://view.tkelevator.com.cn/vivid/niops/purchasing/materials/search"
AUTH_FILE = os.path.join(SCRIPT_DIR, "auth.json")
RUN_LOG_FILE = os.path.join(SCRIPT_DIR, "automation_run_log.txt")
DOWNLOAD_DIR = os.path.join(SCRIPT_DIR, "downloads")
PAGE_SCAN_FILE = os.path.join(SCRIPT_DIR, "page_scan.txt")
PDF_PARSE_TIMEOUT = 45


# ═══════════════════ PO 号提取 ═══════════════════

def extract_all_pos_from_pivot(excel_path=None, sheet_name=None):
    """
    从透视表 Sheet 中提取所有有效的采购凭证号（去重）。
    采购凭证为 7~12 位纯数字。
    """
    import openpyxl

    path = excel_path or EXCEL_FILE
    name = sheet_name or TARGET_SHEET

    if not os.path.exists(path):
        return []

    wb = openpyxl.load_workbook(path, data_only=True)
    if name not in wb.sheetnames:
        wb.close()
        return []

    ws = wb[name]
    seen = set()
    pos = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            val = str(cell.value or "").strip()
            if val and re.match(r'^\d{7,12}$', val) and val not in seen:
                seen.add(val)
                pos.append(val)

    wb.close()
    return pos


# ═══════════════════ 辅助函数 ═══════════════════

def _col_letter(idx):
    """1→A, 26→Z, 27→AA, ..."""
    s = ""
    while idx > 0:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


def _is_on_login_page(page):
    """判断当前页面是否为登录页面。"""
    url = page.url.lower()
    return (
        "login.php" in url
        or "login.microsoftonline" in url
        or "microsoft" in url
    )


# ═══════════════════ 元素定位 ═══════════════════

def _locate_po_input(page, log):
    """多策略定位「PO号」输入框，返回第一个可见 input 或 None。"""
    selectors = [
        "//label[contains(text(),'PO号')]/following-sibling::*/descendant::input[1]",
        "//label[contains(text(),'PO号')]/following::input[1]",
        "//td[contains(text(),'PO')]/following-sibling::td//input[1]",
        "//th[contains(text(),'PO')]/following-sibling::td//input[1]",
        "//span[contains(text(),'PO号')]/ancestor::td/following-sibling::td//input[1]",
        "input[placeholder*='PO']",
        "input[placeholder*='po']",
        "input[name*='po']",
        "input[name*='PO']",
        "input[name*='purchase']",
        "input[id*='po']",
        "input[id*='PO']",
        "//label[contains(text(),'PO')]/following::input[1]",
        "//*[contains(text(),'PO号')]/following::input[1]",
    ]
    for sel in selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=800):
                log(f"  ✓ PO输入框定位成功（{sel[:50]}...）")
                return el
        except Exception:
            continue
    return None


def _locate_search_button(page, log):
    """多策略定位「查找」按钮。找不到时诊断页面所有可见按钮。"""
    selectors = [
        "button:has-text('查找')",
        "//button[contains(text(),'查找')]",
        "//button[normalize-space(text())='查找']",
        "//a[contains(text(),'查找')]",
        "//input[@type='submit'][@value='查找']",
        "//input[@type='button'][@value='查找']",
        "//*[@role='button'][contains(text(),'查找')]",
        "button:has-text('查')",
        "//button[contains(@class,'search')]",
        "//button[contains(@class,'btn')][contains(text(),'查')]",
        "//*[contains(@class,'el-button')][contains(text(),'查找')]",
    ]
    for sel in selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=800):
                log(f"  ✓ 「查找」按钮定位成功（{sel[:50]}）")
                return el
        except Exception:
            continue

    # ── 诊断：列出所有可见按钮 ──
    try:
        all_btns = page.locator("button:visible").all()
        log(f"[诊断] 共 {len(all_btns)} 个可见按钮：")
        for i, btn in enumerate(all_btns[:15]):
            try:
                txt = (btn.inner_text() or "").strip()[:40]
                cls = (btn.get_attribute("class") or "")[:40]
                log(f"  [{i}] text='{txt}' | class='{cls}'")
            except Exception:
                pass
        input_btns = page.locator("input[type='submit']:visible, input[type='button']:visible").all()
        if input_btns:
            log(f"[诊断] 共 {len(input_btns)} 个可见 input 按钮：")
            for i, ib in enumerate(input_btns[:10]):
                try:
                    v = (ib.get_attribute("value") or "")[:40]
                    log(f"  [{i}] value='{v}'")
                except Exception:
                    pass
    except Exception:
        pass

    return None


def _wait_for_search_results(page, log, max_retries=2):
    """等待搜索结果出现，支持重试。"""
    wait_selectors = [
        "//th[contains(text(),'请求ID')]",
        "//a[contains(text(),'LM')]",
        "table:visible",
    ]
    for attempt in range(max_retries):
        if attempt > 0:
            log(f"  搜索结果未出现，第 {attempt + 1} 次重试...")
            try:
                search_btn = _locate_search_button(page, log)
                if search_btn:
                    search_btn.click()
            except Exception:
                pass
            time.sleep(3)
        time.sleep(2)
        for sel in wait_selectors:
            try:
                page.wait_for_selector(sel, state="visible", timeout=8000)
                log("  ✓ 检测到搜索结果")
                return True
            except Exception:
                continue
        time.sleep(3)
    return False


# ═══════════════════ 请求ID点击 & 导航 ═══════════════════

def _click_request_id(page, context, log):
    """
    在搜索结果中找到请求ID并点击，等待新标签页打开。
    返回 (请求ID文本, 详情页page对象 或 None)。
    """
    request_el = None
    selectors = [
        "//a[contains(text(),'LM')]",
        "//td//a",
        "//a[contains(@href,'request')]",
        "//td[contains(text(),'LM')]",
        "//span[contains(text(),'LM')]",
        "//*[contains(text(),'LM80')]",
    ]
    for sel in selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=1000):
                text = el.inner_text().strip()
                if len(text) >= 6:
                    request_el = el
                    log(f"  找到请求ID: {text} (selector: {sel[:40]})")
                    break
        except Exception:
            continue

    # ── 诊断：列出搜索页所有链接 ──
    if request_el is None:
        try:
            all_links = page.locator("a:visible").all()
            log(f"[诊断] 搜索结果页共 {len(all_links)} 个可见链接：")
            for i, link in enumerate(all_links[:15]):
                try:
                    txt = (link.inner_text() or "").strip()[:50]
                    href = (link.get_attribute("href") or "")[:60]
                    log(f"  [{i}] text='{txt}' | href='{href}'")
                except Exception:
                    pass
            tables = page.locator("table:visible").all()
            log(f"[诊断] 共 {len(tables)} 个可见表格")
        except Exception:
            pass

    if request_el is None:
        log("  ⚠ 未找到请求ID链接")
        return ("", None)

    rid_text = request_el.inner_text().strip()
    current_pages = list(context.pages)
    current_url = page.url

    # 多策略点击
    click_methods = [
        ("标准点击", lambda: request_el.click()),
        ("JS点击", lambda: request_el.evaluate("el => el.click()")),
        ("dispatchEvent", lambda: request_el.dispatch_event("click")),
        ("父元素点击", lambda: request_el.locator("..").click()),
    ]

    for method_name, method_fn in click_methods:
        try:
            log(f"  尝试点击: {method_name}")
            method_fn()
            for _ in range(7):
                time.sleep(1.5)
                # 检测新标签页
                new_pages = [p for p in context.pages if p not in current_pages]
                if new_pages:
                    np = new_pages[0]
                    np.bring_to_front()
                    try:
                        np.wait_for_load_state("domcontentloaded", timeout=15000)
                    except Exception:
                        pass
                    log(f"  ✓ {method_name} 成功（新标签页）")
                    return (rid_text, np)
                # 检测 URL 变化
                if page.url != current_url and "search" not in page.url.lower():
                    try:
                        page.wait_for_load_state("domcontentloaded", timeout=15000)
                    except Exception:
                        pass
                    log(f"  ✓ {method_name} 成功（URL变化）")
                    return (rid_text, page)
                # 检测详情页特征文本
                for marker in ["VIEW编号", "项目经理", "梯台明细", "物料供应商"]:
                    try:
                        if page.locator(f"//*[contains(text(),'{marker}')]").is_visible(timeout=500):
                            log(f"  ✓ {method_name} 成功（检测到「{marker}」）")
                            return (rid_text, page)
                    except Exception:
                        continue
            log(f"  {method_name} 未检测到导航，尝试下一个...")
        except Exception as e:
            log(f"  ✗ {method_name} 异常: {e}")

    log("  ⚠ 所有点击方法均未触发导航")
    return (rid_text, None)


# ═══════════════════ 详情页数据提取 ═══════════════════

def _extract_project_name(page, log):
    """从详情页文本中提取「项目名称」字段值。"""
    try:
        body_text = page.inner_text("body")
    except Exception:
        return None
    if not body_text:
        return None

    lines = [l.strip() for l in body_text.split("\n") if l.strip()]

    # 搜索 "项目名称" 标签，取紧随其后的中文值
    for i, line in enumerate(lines):
        if "项目名称" not in line:
            continue
        # 尝试在当前行取「项目名称」之后的文本
        m = re.search(r'项目名称\s*[：:]*\s*([\u4e00-\u9fff][\u4e00-\u9fff\w（）()\-·\s]+)', line)
        if m:
            val = m.group(1).strip()
            if len(val) > 2:
                log(f"  项目名称（同行）: {val[:60]}")
                return val
        # 尝试下一行
        if i + 1 < len(lines):
            next_line = lines[i + 1].strip()
            if next_line and not next_line.startswith("项目") and any('\u4e00' <= c <= '\u9fff' for c in next_line):
                log(f"  项目名称（下行）: {next_line[:60]}")
                return next_line

    # 兜底：正则搜索
    m = re.search(r'项目名称\s*[：:]*\s*([\u4e00-\u9fff][\u4e00-\u9fff\w（）()\-·]+)', body_text)
    if m:
        val = m.group(1).strip()
        if len(val) > 2:
            log(f"  项目名称（正则）: {val[:60]}")
            return val

    log("  ⚠ 未提取到项目名称")
    return None


def _extract_order_qty(page, log):
    """
    从详情页「梯台明细」表中提取合计行的数量值。
    方法：找表格→定位「数量」列→定位「合计」行→取交叉单元格。
    """
    try:
        # 滚动到梯台明细区域
        try:
            target = page.locator("//*[contains(text(),'梯台明细')]").first
            target.scroll_into_view_if_needed()
            time.sleep(0.5)
        except Exception:
            pass

        for table in page.locator("table").all():
            try:
                if not table.is_visible():
                    continue
                tbl_text = table.inner_text()
                if "数量" not in tbl_text or "合计" not in tbl_text:
                    continue

                # 找「数量」列索引
                qty_col = None
                ths = table.locator("th").all()
                for i, th in enumerate(ths):
                    if (th.inner_text() or "").strip() == "数量":
                        qty_col = i
                        break
                if qty_col is None:
                    first_row = table.locator("tr").first
                    tds = first_row.locator("td").all()
                    for i, td in enumerate(tds):
                        if (td.inner_text() or "").strip() == "数量":
                            qty_col = i
                            break
                if qty_col is None:
                    continue

                # 找「合计」行 → 取数量列单元格
                for row in table.locator("tr").all():
                    row_text = row.inner_text() or ""
                    if "合计" not in row_text:
                        continue
                    cells = row.locator("td").all()
                    if len(cells) > qty_col:
                        val = cells[qty_col].inner_text().strip()
                        if val:
                            log(f"  梯台明细合计数量: {val}")
                            return val
            except Exception:
                continue
    except Exception:
        pass

    log("  ⚠ 未提取到梯台明细合计数量")
    return None


def _download_support_files(page, log, subfolder=""):
    """
    滚动到「审批记录」区域，下载「支持文件」列中的全部文件。
    subfolder: 存放在 downloads/子文件夹/ 下（如项目名称）。
    返回已下载的文件名列表。
    """
    target_dir = os.path.join(DOWNLOAD_DIR, subfolder) if subfolder else DOWNLOAD_DIR
    os.makedirs(target_dir, exist_ok=True)

    # 滚动到审批记录区域
    try:
        target = page.locator("//*[contains(text(),'审批记录')]").first
        target.scroll_into_view_if_needed()
        time.sleep(0.5)
    except Exception:
        pass

    downloaded = []
    file_links = []

    # 策略：找含「支持文件」表头的表格
    try:
        table = page.locator("table:has(th:has-text('支持文件'))").first
        if table.is_visible(timeout=2000):
            rows = table.locator("tr").all()
            for row in rows[1:]:
                try:
                    links_in_row = row.locator("td").last.locator("a").all()
                    file_links.extend(links_in_row)
                except Exception:
                    pass
    except Exception:
        pass

    if not file_links:
        try:
            links = page.locator(
                "//th[contains(text(),'支持文件')]/ancestor::table//tr/td[last()]//a"
            ).all()
            file_links.extend(links)
        except Exception:
            pass

    if not file_links:
        try:
            links = page.locator(
                "a[href$='.pdf'], a[href$='.msg'], a[href$='.doc'], "
                "a[href$='.docx'], a[href$='.xls'], a[href$='.xlsx'], a[href$='.zip']"
            ).all()
            file_links.extend(links)
        except Exception:
            pass

    for link in file_links:
        try:
            if not link.is_visible():
                continue
            file_name = link.inner_text().strip()
            if not file_name:
                continue

            log(f"  正在下载: {file_name}")
            with page.expect_download(timeout=30000) as download_info:
                link.click()
            download = download_info.value

            safe_name = file_name.replace("/", "_").replace("\\", "_").replace(":", "_")
            # 如果没有后缀，用 suggested_filename 的后缀
            if "." not in safe_name.rsplit("/", 1)[-1]:
                suggested = download.suggested_filename or ""
                if "." in suggested:
                    ext = suggested.rsplit(".", 1)[-1]
                else:
                    ext = "pdf"
                safe_name = f"{safe_name}.{ext}"

            save_path = os.path.join(target_dir, safe_name)
            download.save_as(save_path)
            downloaded.append(safe_name)
            log(f"    ✓ 已保存: {subfolder + '/' if subfolder else ''}{safe_name}")
            time.sleep(0.5)
        except Exception as e:
            log(f"    ✗ 下载失败: {str(e)[:100]}")

    if not downloaded:
        log("  未发现可下载的支持文件。")
    return downloaded


# ═══════════════════ PDF 金额提取 ═══════════════════

def _extract_total_from_pdf(pdf_path, log):
    """
    读取 PDF 文件，提取「总计（不含税）」后的金额。
    使用子进程执行 pdfplumber 解析，防止卡死主流程。
    """
    log(f"  [PDF] 解析: {os.path.basename(pdf_path)}")
    worker_code = r'''
import json, sys, pdfplumber

pdf_path = sys.argv[1]
logs = []

def emit(status, total=None, error=""):
    print(json.dumps({
        "status": status, "total": total, "error": error, "logs": logs,
    }, ensure_ascii=False))

try:
    with pdfplumber.open(pdf_path) as pdf:
        for page_no, page in enumerate(pdf.pages, 1):
            logs.append(f"  [PDF] 第 {page_no}/{len(pdf.pages)} 页")
            # 归一化旋转（摆正颠倒/横排页面）
            rot = getattr(page, 'rotation', 0) or 0
            if rot != 0:
                logs.append(f"  [PDF] 检测到旋转 {rot}°，正在摆正...")
                try:
                    page = page.rotate(360 - rot)
                except Exception:
                    pass
            text = page.extract_text()
            if not text:
                continue
            lines = text.split("\n")
            for i, line in enumerate(lines):
                if "总计（不含税）" not in line and "总计(不含税)" not in line:
                    continue
                logs.append(f"  [PDF] 匹配行: {line[:150]}")
                # 当前行取数字
                for p in line.split():
                    p = p.strip().replace(",", "").replace("\uffe5", "").replace("\xa5", "")
                    try:
                        v = float(p)
                        if v > 100:
                            logs.append(f"  [PDF] 提取总计: {p}")
                            emit("ok", p)
                            sys.exit(0)
                    except ValueError:
                        continue
                # 下一行取数字
                if i + 1 < len(lines):
                    for p in lines[i + 1].split():
                        p = p.strip().replace(",", "").replace("\uffe5", "").replace("\xa5", "")
                        try:
                            v = float(p)
                            if v > 100:
                                logs.append(f"  [PDF] 提取总计(下行): {p}")
                                emit("ok", p)
                                sys.exit(0)
                        except ValueError:
                            continue
    emit("ok")
except Exception as e:
    emit("error", error=str(e))
'''

    try:
        completed = subprocess.run(
            [sys.executable, "-c", worker_code, pdf_path],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
            timeout=PDF_PARSE_TIMEOUT,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
    except subprocess.TimeoutExpired:
        log(f"  [PDF] 解析超时（{PDF_PARSE_TIMEOUT}秒），跳过")
        return None
    except Exception as e:
        log(f"  [PDF] 子进程启动失败: {e}")
        return None

    output = (completed.stdout or "").strip()
    if not output:
        stderr = (completed.stderr or "").strip()
        if stderr:
            log(f"  [PDF] 无结果: {stderr[:200]}")
        return None

    try:
        data = json.loads(output.splitlines()[-1])
    except Exception:
        log(f"  [PDF] 解析JSON失败: {output[:200]}")
        return None

    for line in data.get("logs", []):
        log(line)

    if data.get("status") == "error":
        log(f"  [PDF] 读取失败: {data.get('error')}")
        return None

    return data.get("total")


# ═══════════════════ Excel 读写 ═══════════════════

def _read_order_value_from_pivot(po):
    """从透视表读取指定 PO 的订单净值（列B「求和项:订单净值」）。"""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        if TARGET_SHEET not in wb.sheetnames:
            wb.close()
            return None
        ws = wb[TARGET_SHEET]
        po_str = str(po).strip()
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row, 1).value or "").strip() == po_str:
                val = ws.cell(row, 2).value
                wb.close()
                return str(val) if val is not None else None
        wb.close()
    except Exception:
        pass
    return None


def _write_to_pivot_excel(po, project_name=None, order_qty=None,
                          support_files=None, total_amount=None, order_diff=None):
    """
    将提取数据写回「装潢透视表」Sheet，按采购凭证行对齐。
    """
    import openpyxl

    if not os.path.exists(EXCEL_FILE):
        return f"Excel 文件不存在: {EXCEL_FILE}"

    wb = openpyxl.load_workbook(EXCEL_FILE)
    if TARGET_SHEET not in wb.sheetnames:
        wb.close()
        return f"Sheet「{TARGET_SHEET}」不存在"

    ws = wb[TARGET_SHEET]

    # 1. 扫描第 1 行定位扩展列
    ext_cols = {}
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(1, col).value or "").strip()
        if val == "E2E项目名":
            ext_cols["E2E项目名"] = col
        elif val == "E2E订单数量":
            ext_cols["E2E订单数量"] = col
        elif val == "项目总金额(审批)":
            ext_cols["项目总金额(审批)"] = col
        elif val == "订单差异(单独计算)":
            ext_cols["订单差异(单独计算)"] = col
        elif val == "支持文件名(FLD 审批)":
            ext_cols["支持文件名(FLD 审批)"] = col

    if not ext_cols:
        wb.close()
        return "未找到扩展列，请先点击「添加扩展列」按钮"

    # 2. 扫描列A 找匹配 PO 的行
    po_str = str(po).strip()
    target_rows = []
    for row in range(2, ws.max_row + 1):
        val = str(ws.cell(row, 1).value or "").strip()
        if val == po_str:
            target_rows.append(row)

    if not target_rows:
        wb.close()
        return f"在列A中未找到采购凭证 {po_str}"

    # 3. 写入数据
    filled = []
    for tr in target_rows:
        if project_name and "E2E项目名" in ext_cols:
            ws.cell(tr, ext_cols["E2E项目名"], project_name)
            filled.append("E2E项目名")
        if order_qty and "E2E订单数量" in ext_cols:
            try:
                ws.cell(tr, ext_cols["E2E订单数量"], float(str(order_qty).replace(",", "")))
            except ValueError:
                ws.cell(tr, ext_cols["E2E订单数量"], order_qty)
            filled.append("E2E订单数量")
        if total_amount is not None and "项目总金额(审批)" in ext_cols:
            try:
                ws.cell(tr, ext_cols["项目总金额(审批)"], float(str(total_amount).replace(",", "")))
            except ValueError:
                ws.cell(tr, ext_cols["项目总金额(审批)"], total_amount)
            filled.append("项目总金额(审批)")
        if order_diff is not None and "订单差异(单独计算)" in ext_cols:
            try:
                ws.cell(tr, ext_cols["订单差异(单独计算)"], float(str(order_diff).replace(",", "")))
            except ValueError:
                ws.cell(tr, ext_cols["订单差异(单独计算)"], order_diff)
            filled.append("订单差异(单独计算)")
        if support_files and "支持文件名(FLD 审批)" in ext_cols:
            ws.cell(tr, ext_cols["支持文件名(FLD 审批)"], support_files)
            filled.append("支持文件名(FLD 审批)")

    wb.save(EXCEL_FILE)
    wb.close()
    return f"已写入 {len(target_rows)} 行（PO {po_str}）：{', '.join(filled)}"


# ═══════════════════ 单 PO 处理 ═══════════════════

def _process_one_po(po, context, log):
    """
    处理单个 PO 的完整浏览器流程：
    搜索 → 点击请求ID → 提取项目名称/数量 → 下载附件。
    返回 dict 或 None。
    """
    result = {"po": po}

    try:
        po_page = context.new_page()
        po_page.set_default_timeout(30000)

        # ── 1. 导航到搜索页 ──
        po_page.goto(SEARCH_URL, wait_until="domcontentloaded", timeout=20000)
        time.sleep(1.5)

        # ── 2. 填入 PO 号 ──
        po_input = _locate_po_input(po_page, log)
        if po_input is None:
            log(f"  ⚠ 无法定位 PO 输入框，跳过")
            return None

        po_input.click()
        po_input.fill("")
        time.sleep(0.2)
        po_input.type(po, delay=80)
        time.sleep(0.3)

        # ── 3. 点击查找 ──
        search_btn = _locate_search_button(po_page, log)
        if search_btn is None:
            # 兜底：按 Enter 键触发搜索
            log("  未找到「查找」按钮，尝试按 Enter 键提交...")
            po_input.press("Enter")
        else:
            search_btn.click()
        log("  已触发搜索...")

        if not _wait_for_search_results(po_page, log):
            log(f"  ⚠ 搜索结果未出现，跳过")
            return None

        # ── 4. 点击请求ID → 详情页 ──
        rid_text, detail_page = _click_request_id(po_page, context, log)
        result["request_id"] = rid_text

        if detail_page is None:
            log(f"  ⚠ 未能进入详情页")
            result["project_name"] = None
            result["order_qty"] = None
            result["downloaded"] = []
        else:
            log(f"  详情页已打开，等待加载...")
            time.sleep(2)
            try:
                detail_page.wait_for_load_state("domcontentloaded", timeout=10000)
            except Exception:
                pass

            # 写入 page_scan.txt 供调试
            try:
                detail_text = detail_page.inner_text("body")
                with open(PAGE_SCAN_FILE, "w", encoding="utf-8") as f:
                    f.write(detail_text)
            except Exception:
                pass

            # ── 5. 提取项目名称（先于下载，用作子文件夹名）──
            result["project_name"] = _extract_project_name(detail_page, log)
            project_dir = result["project_name"] or str(po)
            # 清理文件夹非法字符
            project_dir = re.sub(r'[\\/:*?"<>|]', '_', project_dir)
            result["download_subdir"] = project_dir

            # ── 6. 下载支持文件（存入项目子文件夹）──
            result["downloaded"] = _download_support_files(detail_page, log, subfolder=project_dir)
            log(f"  附件下载: {len(result['downloaded'])} 个")

            # ── 7. 提取梯台明细合计数量 ──
            result["order_qty"] = _extract_order_qty(detail_page, log)

            # 关闭详情页
            if detail_page != po_page:
                try:
                    detail_page.close()
                except Exception:
                    pass

        # ★ 保留搜索页给下一个 PO 复用，不关闭页面
        log(f"  ✓ PO {po} 网页阶段完成")
        return result

    except Exception as e:
        log(f"  ✗ PO {po} 异常: {e}")
        return None


def _finalize_po_result(result, log):
    """
    本地处理：PDF解析 + 差异计算 + Excel写回。
    不依赖浏览器，避免 PDF 解析阻塞后续 PO 查询。
    """
    po = result.get("po")

    # ── PDF 解析 → 提取总计（不含税）──
    total_amount = None
    downloaded = result.get("downloaded", [])
    if downloaded:
        log("  解析附件...")
        subdir = result.get("download_subdir", "")
        for fname in downloaded:
            fpath = os.path.join(DOWNLOAD_DIR, subdir, fname) if subdir else os.path.join(DOWNLOAD_DIR, fname)
            if not os.path.exists(fpath):
                log(f"  附件不存在，跳过: {fname}")
                continue
            if fname.lower().endswith(".pdf"):
                total_amount = _extract_total_from_pdf(fpath, log)
                if total_amount:
                    break
        if total_amount:
            log(f"  总计（不含税）: {total_amount}")
        else:
            log("  未从附件中解析到总计金额")
    result["total_amount"] = total_amount

    # ── 读取订单净值 → 计算差异 ──
    order_value = _read_order_value_from_pivot(po)
    result["order_value"] = order_value
    order_diff = None

    if total_amount is not None and order_value is not None:
        try:
            ta = float(str(total_amount).replace(",", ""))
            ov = float(str(order_value).replace(",", ""))
            order_diff = round(ta - ov, 2)
            log(f"  订单差异: {order_diff} (={ta} - {ov})")
        except (ValueError, TypeError):
            log("  ⚠ 差异计算失败")
    result["order_diff"] = order_diff

    # ── 统计非 PDF 附件 ──
    non_pdf = [f for f in downloaded if not f.lower().endswith(".pdf")]
    result["non_pdf_count"] = len(non_pdf)

    # ── 写回 Excel ──
    support_files_str = "; ".join(downloaded) if downloaded else ""
    diff_str = str(order_diff) if order_diff is not None else ""
    total_str = str(total_amount) if total_amount is not None else ""

    excel_msg = _write_to_pivot_excel(
        po,
        project_name=result.get("project_name"),
        order_qty=result.get("order_qty"),
        support_files=support_files_str,
        total_amount=total_str,
        order_diff=diff_str,
    )
    log(f"  {excel_msg}")
    log(f"  ✓ PO {po} 本地处理完成")
    return result


# ═══════════════════ 主入口 ═══════════════════

def open_website_and_search(status_callback=None):
    """
    主流程：
      1. 从透视表提取全部 PO 号
      2. 启动 Edge → 登录 → 循环处理每个 PO（搜索/提取/下载）
      3. 所有 PO 网页阶段完成后 → 本地解析 PDF → 写回 Excel
      4. 汇总结果
    """
    with open(RUN_LOG_FILE, "w", encoding="utf-8") as f:
        f.write(f"自动化运行日志 {time.strftime('%Y-%m-%d %H:%M:%S')}\n")

    def log(msg):
        if status_callback:
            status_callback(msg)
        try:
            with open(RUN_LOG_FILE, "a", encoding="utf-8") as f:
                f.write(f"{time.strftime('%H:%M:%S')} {msg}\n")
        except Exception:
            pass

    # ── 1. 提取全部 PO ──
    log("正在从透视表提取全部采购凭证号...")
    all_pos = extract_all_pos_from_pivot()
    if not all_pos:
        return False, "透视表中未找到任何采购凭证号。\n请确认透视表已生成。"
    total = len(all_pos)
    log(f"共提取到 {total} 个采购凭证号")

    # ── 2. 启动浏览器 ──
    log("正在启动 Edge 浏览器...")
    from playwright.sync_api import sync_playwright
    playwright = sync_playwright().start()

    try:
        browser = playwright.chromium.launch(
            channel="msedge",
            headless=False,
            args=["--start-maximized"],
        )

        context_kwargs = {"no_viewport": True, "accept_downloads": True}
        if os.path.exists(AUTH_FILE):
            log("检测到已保存的登录状态，尝试自动登录...")
            try:
                context = browser.new_context(**{**context_kwargs, "storage_state": AUTH_FILE})
                log("已加载登录状态。")
            except Exception:
                os.remove(AUTH_FILE)
                context = browser.new_context(**context_kwargs)
                log("登录状态文件已过期，需要重新登录。")
        else:
            context = browser.new_context(**context_kwargs)

        page = context.new_page()

        # ── 3. 登录检测 ──
        log("正在打开 TKE VIEW...")
        page.goto(HOME_URL, wait_until="domcontentloaded", timeout=30000)
        time.sleep(2)

        if _is_on_login_page(page):
            log("════════════════════════════════")
            log("⚠ 需要登录 — 请在浏览器中完成 Microsoft 账户验证")
            log("  登录完成后程序将自动继续...")
            log("════════════════════════════════")
            try:
                page.wait_for_url(
                    lambda url: "login.php" not in url.lower()
                    and "login.microsoftonline" not in url.lower(),
                    timeout=300000
                )
                log("✅ 登录成功！正在保存登录状态...")
                context.storage_state(path=AUTH_FILE)
                log("登录状态已保存。")
                time.sleep(2)
            except Exception:
                return False, "登录超时（超过 5 分钟），请重试。"
        else:
            log("✅ 已登录，会话有效。")

        try:
            page.close()
        except Exception:
            pass

        # ── 4. 循环处理每个 PO（测试模式：前5个）──
        test_pos = all_pos[:5]
        total = len(test_pos)
        log(f"[测试模式] 仅处理前 {total} 个 PO")

        success_count = 0
        fail_count = 0
        failed_pos = []
        web_results = []
        crashed = False

        for i, po in enumerate(test_pos):
            log("")
            log(f"━━━ [{i+1}/{total}] PO: {po} ━━━")
            try:
                result = _process_one_po(po, context, log)
                if result is not None:
                    success_count += 1
                    web_results.append(result)
                    log(f"✓ [{i+1}/{total}] PO {po} 网页阶段完成")
                else:
                    fail_count += 1
                    failed_pos.append(po)
                    log(f"⚠ [{i+1}/{total}] PO {po} 未完成")
            except Exception as e:
                log(f"  ✗ 浏览器异常: {e}")
                fail_count += 1
                failed_pos.append(po)
                error_str = str(e).lower()
                if any(kw in error_str for kw in ("epipe", "closed", "target closed", "connection")):
                    log("⚠ 浏览器连接断开，停止后续处理。")
                    crashed = True
                    break

            # PO 之间短暂休息
            time.sleep(0.5)

        # 保留浏览器供查看
        if not crashed:
            log("网页阶段完成，浏览器窗口保留供查看。")

        # ── 5. 本地解析 PDF + 写回 Excel ──
        finalize_success = 0
        finalize_fail = 0
        total_non_pdf = 0
        if web_results:
            log("")
            log(f"开始本地解析并写回 Excel（共 {len(web_results)} 个 PO）...")
            for j, result in enumerate(web_results, 1):
                po = result.get("po")
                log("")
                log(f"━━━ [本地 {j}/{len(web_results)}] PO: {po} ━━━")
                try:
                    _finalize_po_result(result, log)
                    finalize_success += 1
                    total_non_pdf += result.get("non_pdf_count", 0)
                except Exception as e:
                    finalize_fail += 1
                    log(f"  ✗ PO {po} 本地处理异常: {e}")

        # ── 6. 汇总 ──
        log("")
        if crashed:
            log(f"⚠ 浏览器异常中断！网页阶段完成: {success_count} / 失败: {fail_count}")
        else:
            log(f"✅ 全部完成！网页阶段: 成功 {success_count} / 失败 {fail_count} / 总数 {total}")
        log(f"本地解析写回: 成功 {finalize_success} / 失败 {finalize_fail}")

        auth_info = "登录状态已保存，下次自动登录。" if os.path.exists(AUTH_FILE) else ""
        summary = (
            f"✅ 全部查询完成！\n\n"
            f"网页查询成功: {success_count} / 失败: {fail_count} / 总数: {total}\n"
            f"本地解析写回成功: {finalize_success} / 失败: {finalize_fail}"
        )
        if total_non_pdf > 0:
            summary += f"\n\n⚠ 有 {total_non_pdf} 个附件不是 PDF 格式，金额未提取（后续版本将支持更多格式）。"
        if failed_pos:
            summary += f"\n\n失败 PO ({len(failed_pos)} 个):\n"
            for fp in failed_pos[:20]:
                summary += f"  - {fp}\n"
            if len(failed_pos) > 20:
                summary += f"  ... 等共 {len(failed_pos)} 个\n"
        summary += f"\n\n{auth_info}"
        return True, summary

    except Exception as e:
        return False, f"网页操作异常：\n{e}\n\n{traceback.format_exc()}"
