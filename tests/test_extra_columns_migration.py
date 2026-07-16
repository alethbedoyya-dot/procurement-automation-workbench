import runpy
import sys
import tempfile
import types
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl


PROJECT_DIR = Path(__file__).resolve().parents[1]
GUI_SCRIPT = next(
    path
    for path in PROJECT_DIR.glob("*.py")
    if path.name != "web_query.py" and path.stat().st_size > 50_000
)


class ExtraColumnsMigrationTests(unittest.TestCase):
    def test_adding_wbs_to_a_legacy_sheet_preserves_existing_extension_data(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            excel_path = Path(temp_dir) / "采购记录.xlsx"
            old_columns = [
                "E2E项目名",
                "E2E订单数量",
                "项目总金额(审批)",
                "订单差异(单独计算)",
                "支持文件名(FLD 审批)",
                "Price",
                "PlanCost",
                "总saving",
                "订单是否下完=price-订单净值",
            ]
            legacy_values = ["项目 A", 12, 1000, 25, "FLD审批.msg", 900, 800, 100, 0]
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "装潢透视表"
            worksheet.append(["采购凭证", *old_columns])
            worksheet.append(["4000000001", *legacy_values])
            workbook.save(excel_path)
            workbook.close()

            ttk_stub = types.ModuleType("ttkbootstrap")
            with patch.dict(sys.modules, {"ttkbootstrap": ttk_stub}):
                module = runpy.run_path(str(GUI_SCRIPT))
            function_globals = module["add_extra_columns"].__globals__

            with patch.dict(function_globals, {"EXCEL_FILE": str(excel_path)}):
                success, _ = module["add_extra_columns"]("装潢")

            self.assertTrue(success)
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            worksheet = workbook["装潢透视表"]
            self.assertEqual(
                [worksheet.cell(1, column).value for column in range(2, 12)],
                ["E2E项目名", "WBS", *old_columns[1:]],
            )
            self.assertEqual(worksheet["B2"].value, legacy_values[0])
            self.assertIsNone(worksheet["C2"].value)
            self.assertEqual(
                [worksheet.cell(2, column).value for column in range(4, 12)],
                legacy_values[1:],
            )
            workbook.close()


if __name__ == "__main__":
    unittest.main()
