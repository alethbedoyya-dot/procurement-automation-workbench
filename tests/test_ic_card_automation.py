import gc
import runpy
import sys
import tempfile
import types
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl


PROJECT_DIR = Path(__file__).resolve().parents[1]
GUI_SCRIPT = next(
    path
    for path in PROJECT_DIR.glob("*.py")
    if path.name != "web_query.py" and path.stat().st_size > 50_000
)


class ICCardAutomationTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        ttk_stub = types.ModuleType("ttkbootstrap")
        with patch.dict(sys.modules, {"ttkbootstrap": ttk_stub}):
            self.module = runpy.run_path(str(GUI_SCRIPT))
        self.module_globals = self.module["fill_air_conditioning_old_new_prices"].__globals__
        self.excel_path = Path(self.temp_dir.name) / "采购记录.xlsx"
        self.price_path = Path(self.temp_dir.name) / "价格表.xlsx"
        self.tracking_path = Path(self.temp_dir.name) / "NI PM Saving Tracking.xlsx"
        self.download_dir = Path(self.temp_dir.name) / "downloads"
        self.module_globals["EXCEL_FILE"] = str(self.excel_path)
        self.module_globals["DOWNLOAD_DIR"] = str(self.download_dir)
        self.module_globals["TRACKING_FILE"] = str(self.tracking_path)

    def tearDown(self):
        gc.collect()
        self.temp_dir.cleanup()

    def test_ic_card_reuses_air_conditioning_configuration_and_price_flow(self):
        cfg = self.module["CATEGORIES"]["IC卡"]
        self.assertEqual(cfg["filter_materials"], [1000027312])
        self.assertEqual(cfg["data_sheet"], "IC卡数据")
        self.assertEqual(cfg["target_sheet"], "IC卡透视表")
        self.assertEqual(cfg["content_filter_values"], ("IC卡",))
        self.assertEqual(cfg["price_list_sheet"], "IC卡25.10降价")

        workbook = openpyxl.Workbook()
        data_sheet = workbook.active
        data_sheet.title = "IC卡数据"
        data_sheet.append([
            "采购凭证", "短文本", "订单净值", "老价格", "新价格", "Saving", "CHECK", "净价", "采购订单数量",
        ])
        data_sheet.append(["PO-IC-001", "IC CARD 001", 100, None, None, None, None, 90, 2])
        pivot_sheet = workbook.create_sheet("IC卡透视表")
        pivot_sheet.append(["采购凭证", "求和项:订单净值", "支持文件名(FLD 审批)", "总saving"])
        workbook.save(self.excel_path)
        workbook.close()

        (self.download_dir / "IC卡" / "无FLD文件" / "PO-IC-001").mkdir(parents=True)

        price_book = openpyxl.Workbook()
        price_sheet = price_book.active
        price_sheet.title = "IC卡25.10降价"
        price_sheet.append([None] * 10)
        price_sheet.append([None, None, None, None, "IC CARD 001", "普通物料", None, 120, None, 110])
        price_book.save(self.price_path)
        price_book.close()

        success, _ = self.module["fill_category_old_new_prices"]("IC卡", str(self.price_path))

        self.assertTrue(success)
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        data_sheet = workbook["IC卡数据"]
        self.assertEqual(
            tuple(data_sheet.cell(2, column).value for column in range(4, 8)),
            (120, 110, 60, 20),
        )
        workbook.close()

    def test_outsource_board_reuses_decoration_five_step_configuration(self):
        cfg = self.module["CATEGORIES"]["外包板"]

        self.assertEqual(cfg["filter_materials"], [1000027309])
        self.assertEqual(cfg["data_sheet"], "外包板数据")
        self.assertEqual(cfg["target_sheet"], "外包板透视表")
        self.assertEqual(cfg["pivot_table_name"], "外包板透视表")
        self.assertEqual(cfg["content_filter"], "外包板")
        self.assertEqual(cfg["insert_cols_after_order"], [])
        self.assertIsNone(cfg["insert_col_at_end"])
        self.assertNotIn("price_list_sheet", cfg)

    def test_outsource_board_pm_tracking_uses_outsource_content(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "外包板透视表"
        worksheet.append([
            "采购凭证", "求和项:订单净值", "E2E项目名", "项目总金额(审批)",
            "支持文件名(FLD 审批)", "Price", "PlanCost", "总saving",
            "订单是否下完=price-订单净值",
        ])
        worksheet.append(["PO-OUT-001", 200, "外包板项目测试", 500, None, None, None, None, None])
        workbook.save(self.excel_path)
        workbook.close()

        tracking_book = openpyxl.Workbook()
        tracking_sheet = tracking_book.active
        tracking_sheet.title = "Base Data"
        tracking_sheet.append(["无关表头"])
        tracking_sheet.append(["Project Name", "Price", "PlanCost", "Content"])
        tracking_sheet.append(["外包板项目测试采购", 500, 100, "外包板"])
        tracking_sheet.append(["外包板项目测试装潢", 900, 200, "装潢"])
        tracking_book.save(self.tracking_path)
        tracking_book.close()

        success, message = self.module["match_pm_tracking_data"](category="外包板")

        self.assertTrue(success, message)
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        worksheet = workbook["外包板透视表"]
        self.assertEqual(
            tuple(worksheet.cell(2, column).value for column in range(6, 10)),
            (500, 100, 40, 300),
        )
        workbook.close()

    def test_line_tray_creates_only_data_sheet_then_fills_prices_and_saving(self):
        workbook = openpyxl.Workbook()
        source_sheet = workbook.active
        source_sheet.title = "Sheet1"
        source_sheet.append([
            "物料", "订单净值", "短文本", "净价", "采购订单数量",
        ])
        source_sheet.append([1000027323, 200, "线槽物料A", 70, 2])
        source_sheet.append([1000027323, 300, "未匹配线槽", 50, 3])
        source_sheet.append([1000027323, 0, "线槽物料零", 0, 4])
        source_sheet.append([1000027316, 400, "其他物料", 10, 1])
        workbook.save(self.excel_path)
        workbook.close()

        success, message = self.module["generate_pivot_table"](category="线槽")

        self.assertTrue(success, message)
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        self.assertIn("井道线槽数据", workbook.sheetnames)
        self.assertNotIn("线槽透视表", workbook.sheetnames)
        data_sheet = workbook["井道线槽数据"]
        self.assertEqual(
            [data_sheet.cell(1, column).value for column in range(1, 9)],
            ["物料", "订单净值", "短文本", "净价", "采购订单数量", "老价格", "新价格", "Saving"],
        )
        self.assertEqual(data_sheet.max_row, 4)
        workbook.close()

        price_book = openpyxl.Workbook()
        price_sheet = price_book.active
        price_sheet.title = "井道线槽25.10降价"
        price_sheet.append([None] * 10)
        price_sheet.append([None, None, None, None, "线槽物料A", None, None, 100, None, 80])
        price_sheet.append([None, None, None, None, "线槽物料零", None, None, 0, None, 0])
        price_book.save(self.price_path)
        price_book.close()

        success, message = self.module["fill_category_old_new_prices"](
            "线槽", str(self.price_path)
        )

        self.assertTrue(success, message)
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        data_sheet = workbook["井道线槽数据"]
        self.assertEqual(
            tuple(data_sheet.cell(2, column).value for column in range(6, 9)),
            (100, 80, 60),
        )
        self.assertEqual(
            tuple(data_sheet.cell(3, column).value for column in range(6, 9)),
            (None, None, None),
        )
        self.assertEqual(
            tuple(data_sheet.cell(4, column).value for column in range(6, 9)),
            (0, 0, 0),
        )
        workbook.close()

    def test_five_party_call_creates_only_data_sheet_then_fills_prices_and_saving(self):
        workbook = openpyxl.Workbook()
        source_sheet = workbook.active
        source_sheet.title = "Sheet1"
        source_sheet.append([
            "物料", "订单净值", "短文本", "净价", "采购订单数量",
        ])
        source_sheet.append([8001366263, 200, "五方物料A", 60, 3])
        source_sheet.append([8001366265, 300, "五方物料B", 80, 2])
        source_sheet.append([8001366266, 400, "未匹配五方", 50, 1])
        source_sheet.append([8001366267, 100, "五方物料零", 0, 2])
        source_sheet.append([1000027323, 300, "其他物料", 10, 1])
        workbook.save(self.excel_path)
        workbook.close()

        success, message = self.module["generate_pivot_table"](
            category="无线五方通话"
        )

        self.assertTrue(success, message)
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        self.assertIn("五方通话数据", workbook.sheetnames)
        self.assertNotIn("无线五方通话透视表", workbook.sheetnames)
        data_sheet = workbook["五方通话数据"]
        self.assertEqual(
            [data_sheet.cell(1, column).value for column in range(1, 9)],
            ["物料", "订单净值", "短文本", "净价", "采购订单数量", "老价格", "新价格", "Saving"],
        )
        self.assertEqual(data_sheet.max_row, 5)
        workbook.close()

        price_book = openpyxl.Workbook()
        price_sheet = price_book.active
        price_sheet.title = "五方2026.05降价"
        price_sheet.append([None] * 10)
        price_sheet.append([None, None, None, None, "五方物料A", None, None, 100, None, 80])
        price_sheet.append([None, None, None, None, "五方物料B", None, None, 120, None, 90])
        price_sheet.append([None, None, None, None, "五方物料零", None, None, 0, None, 0])
        price_book.save(self.price_path)
        price_book.close()

        success, message = self.module["fill_category_old_new_prices"](
            "无线五方通话", str(self.price_path)
        )

        self.assertTrue(success, message)
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        data_sheet = workbook["五方通话数据"]
        self.assertEqual(
            tuple(data_sheet.cell(2, column).value for column in range(6, 9)),
            (100, 80, 120),
        )
        self.assertEqual(
            tuple(data_sheet.cell(3, column).value for column in range(6, 9)),
            (120, 90, 80),
        )
        self.assertEqual(
            tuple(data_sheet.cell(4, column).value for column in range(6, 9)),
            (None, None, None),
        )
        self.assertEqual(
            tuple(data_sheet.cell(5, column).value for column in range(6, 9)),
            (0, 0, 0),
        )
        workbook.close()

    def test_shaft_lighting_prepares_description_and_saving_layout(self):
        workbook = openpyxl.Workbook()
        source_sheet = workbook.active
        source_sheet.title = "Sheet1"
        source_sheet.append(["物料", "订单净值", "短文本"])
        source_sheet.append([1000027319, 100, "GL Shaft Lighting 044"])
        source_sheet.append([1000027319, 200, "ZL Shaft Lighting 045"])
        source_sheet.append([1000027319, 300, "AA Shaft Lighting 071"])
        source_sheet.append([1000027318, 400, "LCD Shaft Lighting 044"])
        workbook.save(self.excel_path)
        workbook.close()

        cfg = self.module["CATEGORIES"]["井道照明"]
        total_cols, total_rows, _ = self.module["_enhance_and_filter"](cfg, lambda _msg: None)

        self.assertEqual((total_cols, total_rows), (5, 3))
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        data_sheet = workbook["井道照明数据"]
        self.assertEqual(
            [data_sheet.cell(1, column).value for column in range(1, 6)],
            ["物料", "订单净值", "短文本", "描述", "PO数量"],
        )
        self.assertEqual(
            tuple(data_sheet.cell(2, column).value for column in range(4, 6)),
            ("Shaft Lighting 044", 1),
        )
        self.assertEqual(
            tuple(data_sheet.cell(3, column).value for column in range(4, 6)),
            ("Shaft Lighting 045", 1),
        )
        workbook.close()

        class FakeRange:
            def __init__(self):
                self.Value = None
                self.Formula = None
                self.NumberFormat = None
                self.merged = False

            def Merge(self):
                self.merged = True

        class FakeWorksheet:
            def __init__(self):
                self.ranges = {}

            def Range(self, address):
                return self.ranges.setdefault(address, FakeRange())

        worksheet = FakeWorksheet()
        self.module["_apply_lighting_saving_layout"](worksheet)

        self.assertTrue(worksheet.ranges["I5:J5"].merged)
        self.assertEqual(worksheet.ranges["M6"].Value, "SAP PO 数量 (=Shaft Lighting 044 + 045)")
        self.assertEqual(
            worksheet.ranges["N6"].Formula,
            '=IFERROR(GETPIVOTDATA("求和项:PO数量",$A$1,"描述","Shaft Lighting 044"),0)+IFERROR(GETPIVOTDATA("求和项:PO数量",$A$1,"描述","Shaft Lighting 045"),0)',
        )
        self.assertEqual(worksheet.ranges["N8"].Formula, '=IFERROR(GETPIVOTDATA("求和项:PO数量",$A$1,"描述","Shaft Lighting 053"),0)-N6-N7')
        self.assertEqual(worksheet.ranges["O6"].Formula, "=N6*L6")
        self.assertEqual(worksheet.ranges["O12"].Formula, "=SUM(O6:O11)")

    def test_shaft_lighting_uses_description_and_po_count_for_native_pivot(self):
        workbook = openpyxl.Workbook()
        source_sheet = workbook.active
        source_sheet.title = "Sheet1"
        source_sheet.append(["物料", "订单净值", "短文本"])
        source_sheet.append([1000027319, 100, "GL Shaft Lighting 044"])
        workbook.save(self.excel_path)
        workbook.close()

        class FakeRange:
            def __init__(self):
                self.Value = None
                self.Formula = None
                self.NumberFormat = None

            def Merge(self):
                pass

        class FakeWorksheet:
            def __init__(self):
                self.ranges = {}

            def Range(self, address):
                return self.ranges.setdefault(address, FakeRange())

        class FakeWorkbook:
            def Save(self):
                pass

            def Close(self, **_kwargs):
                pass

        class FakeExcel:
            def __init__(self):
                self.Workbooks = types.SimpleNamespace(Open=lambda _path: FakeWorkbook())

        captured = {}
        pivot_sheet = FakeWorksheet()

        def fake_create_pivot(*_args, **kwargs):
            captured.update(kwargs)
            return pivot_sheet, object()

        with patch.dict(self.module_globals, {
            "_com_start": FakeExcel,
            "_com_stop": lambda *_args: None,
            "_com_create_pivot": fake_create_pivot,
        }):
            success, message = self.module["generate_pivot_table"](category="井道照明")

        self.assertTrue(success, message)
        self.assertEqual(captured["row_fields"], ["描述"])
        self.assertEqual(captured["value_field"], "PO数量")
        self.assertEqual(captured["value_name"], "求和项:PO数量")
        self.assertEqual(pivot_sheet.ranges["O6"].Formula, "=N6*L6")

    def test_monitor_configuration_is_ready_for_the_shared_six_step_workflow(self):
        cfg = self.module["CATEGORIES"]["监控"]

        self.assertEqual(cfg["filter_materials"], [1000027313])
        self.assertEqual(cfg["data_sheet"], "监控数据")
        self.assertEqual(cfg["target_sheet"], "监控透视表")
        self.assertEqual(cfg["content_filter_values"], ("监控",))
        self.assertTrue(cfg["content_filter_exact_on_multiple_matches"])
        self.assertEqual(cfg["price_list_sheet"], "Monitor 25.10降价")

    def test_lcd_configuration_is_ready_for_the_shared_six_step_workflow(self):
        cfg = self.module["CATEGORIES"]["LCD"]

        self.assertEqual(cfg["filter_materials"], [1000027318])
        self.assertEqual(cfg["data_sheet"], "LCD数据")
        self.assertEqual(cfg["target_sheet"], "LCD透视表")
        self.assertEqual(cfg["content_filter_values"], ("LCD",))
        self.assertTrue(cfg["content_filter_exact_on_multiple_matches"])
        self.assertEqual(cfg["price_list_sheet"], "LCD 2026.1月降价")

    def test_monitor_pm_tracking_uses_exact_monitor_content(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "监控透视表"
        worksheet.append([
            "采购凭证", "求和项:订单净值", "E2E项目名", "项目总金额(审批)",
            "支持文件名(FLD 审批)", "Price", "PlanCost", "总saving",
            "订单是否下完=price-订单净值",
        ])
        worksheet.append(["PO-MON-PM", 200, "监控项目测试", None, None, None, None, None, None])
        workbook.save(self.excel_path)
        workbook.close()

        tracking_book = openpyxl.Workbook()
        tracking_sheet = tracking_book.active
        tracking_sheet.title = "Base Data"
        tracking_sheet.append(["无关表头"])
        tracking_sheet.append(["Project Name", "Price", "PlanCost", "Content"])
        tracking_sheet.append(["监控项目测试采购", 500, 100, "监控"])
        tracking_sheet.append(["监控项目测试历史", 900, 200, "监控-历史"])
        tracking_book.save(self.tracking_path)
        tracking_book.close()

        success, _ = self.module["match_pm_tracking_data"](category="监控")

        self.assertTrue(success)
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        worksheet = workbook["监控透视表"]
        self.assertEqual(
            tuple(worksheet.cell(2, column).value for column in range(6, 10)),
            (500, 100, 40, 300),
        )
        workbook.close()

    def test_lcd_pm_tracking_uses_exact_lcd_content(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "LCD透视表"
        worksheet.append([
            "采购凭证", "求和项:订单净值", "E2E项目名", "项目总金额(审批)",
            "支持文件名(FLD 审批)", "Price", "PlanCost", "总saving",
            "订单是否下完=price-订单净值",
        ])
        worksheet.append(["PO-LCD-PM", 200, "LCD项目测试", None, None, None, None, None, None])
        workbook.save(self.excel_path)
        workbook.close()

        tracking_book = openpyxl.Workbook()
        tracking_sheet = tracking_book.active
        tracking_sheet.title = "Base Data"
        tracking_sheet.append(["无关表头"])
        tracking_sheet.append(["Project Name", "Price", "PlanCost", "Content"])
        tracking_sheet.append(["LCD项目测试采购", 500, 100, "LCD"])
        tracking_sheet.append(["LCD项目测试历史", 900, 200, "LCD-历史"])
        tracking_book.save(self.tracking_path)
        tracking_book.close()

        success, _ = self.module["match_pm_tracking_data"](category="LCD")

        self.assertTrue(success)
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        worksheet = workbook["LCD透视表"]
        self.assertEqual(
            tuple(worksheet.cell(2, column).value for column in range(6, 10)),
            (500, 100, 40, 300),
        )
        workbook.close()

    def test_ic_card_pm_tracking_uses_exact_ic_card_content(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "IC卡透视表"
        worksheet.append([
            "采购凭证", "求和项:订单净值", "E2E项目名", "项目总金额(审批)",
            "支持文件名(FLD 审批)", "Price", "PlanCost", "总saving",
            "订单是否下完=price-订单净值",
        ])
        worksheet.append(["PO-IC-PM", 200, "IC卡项目测试", None, None, None, None, None, None])
        workbook.save(self.excel_path)
        workbook.close()

        tracking_book = openpyxl.Workbook()
        tracking_sheet = tracking_book.active
        tracking_sheet.title = "Base Data"
        tracking_sheet.append(["无关表头"])
        tracking_sheet.append(["Project Name", "Price", "PlanCost", "Content"])
        tracking_sheet.append(["IC卡项目测试-采购", 500, 100, "IC卡"])
        tracking_sheet.append(["IC卡项目测试-历史", 900, 200, "IC卡-历史"])
        tracking_book.save(self.tracking_path)
        tracking_book.close()

        success, _ = self.module["match_pm_tracking_data"](category="IC卡")

        self.assertTrue(success)
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        worksheet = workbook["IC卡透视表"]
        self.assertEqual(
            tuple(worksheet.cell(2, column).value for column in range(6, 10)),
            (500, 100, 40, 300),
        )
        workbook.close()


if __name__ == "__main__":
    unittest.main()
