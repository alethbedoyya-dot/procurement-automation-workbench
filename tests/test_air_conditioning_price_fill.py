import gc
import runpy
import tempfile
import unittest
from pathlib import Path

import openpyxl


PROJECT_DIR = Path(__file__).resolve().parents[1]
GUI_SCRIPT = next(
    path
    for path in PROJECT_DIR.glob("*.py")
    if path.name != "web_query.py" and path.stat().st_size > 50_000
)


class AirConditioningPriceFillTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.module = runpy.run_path(str(GUI_SCRIPT))
        self.excel_path = Path(self.temp_dir.name) / "采购记录.xlsx"
        self.price_path = Path(self.temp_dir.name) / "AC价格表.xlsx"
        self.download_dir = Path(self.temp_dir.name) / "downloads"
        self.module_globals = self.module["match_pm_tracking_data"].__globals__
        self.module_globals["EXCEL_FILE"] = str(self.excel_path)
        self.module_globals["DOWNLOAD_DIR"] = str(self.download_dir)

    def tearDown(self):
        gc.collect()
        self.temp_dir.cleanup()

    def _write_air_conditioning_data(self, rows):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "空调数据"
        worksheet.append([
            "采购凭证", "短文本", "订单净值", "老价格", "新价格", "Saving", "CHECK", "净价", "采购订单数量",
        ])
        for row in rows:
            worksheet.append(row)
            po_number = str(row[0] or "").strip()
            if po_number:
                # 查询完成但没有 FLD 时，实际程序也会保留一个无 FLD PO 目录。
                self._download_po_dir("无FLD文件", po_number).mkdir(
                    parents=True, exist_ok=True
                )
        workbook.save(self.excel_path)
        workbook.close()

    def _write_air_conditioning_pivot(self, rows, native_row_label_layout=False):
        workbook = openpyxl.load_workbook(self.excel_path)
        worksheet = workbook.create_sheet("空调透视表")
        if native_row_label_layout:
            worksheet.append([
                "行标签", "求和项:订单净值", "支持文件名(FLD 审批)", "总saving",
            ])
        else:
            worksheet.append([
                "采购凭证", "求和项:订单净值", "支持文件名(FLD 审批)", "总saving",
            ])
        for row in rows:
            worksheet.append(row)
            po_number = str(row[0] or "").strip()
            support_file_name = str(row[2] or "").strip() if len(row) > 2 else ""
            if po_number and support_file_name:
                fld_dir = self._download_po_dir("FLD文件", po_number)
                fld_dir.mkdir(parents=True, exist_ok=True)
                (fld_dir / support_file_name).write_bytes(b"test")
        workbook.save(self.excel_path)
        workbook.close()

    def _download_po_dir(self, bucket, po_number):
        return self.download_dir / "空调" / bucket / str(po_number)

    def _write_download_file(self, bucket, po_number, file_name):
        po_dir = self._download_po_dir(bucket, po_number)
        po_dir.mkdir(parents=True, exist_ok=True)
        (po_dir / file_name).write_bytes(b"test")

    def _write_price_list(self, rows):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "AC 25.10降价"
        header = [None] * 10
        header[4] = "SAP Discription-Ner"
        header[5] = "Discription"
        header[7] = "老价格来源"
        header[9] = "新价格来源"
        worksheet.append(header)
        for sap_description, description, old_price, new_price in rows:
            row = [None] * 10
            row[4] = sap_description
            row[5] = description
            row[7] = old_price
            row[9] = new_price
            worksheet.append(row)
        workbook.save(self.price_path)
        workbook.close()

    def _read_target_rows(self):
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        worksheet = workbook["空调数据"]
        rows = [tuple(row) for row in worksheet.iter_rows(min_row=2, values_only=True)]
        workbook.close()
        return rows

    def test_fills_prices_and_preserves_saving_or_check_when_net_price_is_missing(self):
        self._write_air_conditioning_data([
            ["PO-001", "HZHS AC 003", 100, None, None, "保持原值", "保持原值"],
            ["PO-002", "安装调试费（含辅料）", 100, None, None, "保持原值", "保持原值"],
            ["PO-003", "机房空调", 100, None, None, "保持原值", "保持原值"],
        ])
        self._write_air_conditioning_pivot([])
        self._write_price_list([
            ["HZHS AC 003", "普通物料", 2439.07, 2292.73],
            ["HZHS AC 022", "安装调试费（含辅料）", 400, 376],
            ["NJLD AC 013", "安装调试费（含辅料）", 400, 376],
            ["GZLD AC 012", "安装调试费（含辅料）", 400, 376],
        ])

        success, message = self.module["fill_air_conditioning_old_new_prices"](str(self.price_path))

        self.assertTrue(success)
        self.assertIn("机房空调", message)
        self.assertEqual(
            self._read_target_rows(),
            [
                ("PO-001", "HZHS AC 003", 100, 2439.07, 2292.73, "保持原值", "保持原值", None, None),
                ("PO-002", "安装调试费（含辅料）", 100, 400, 376, "保持原值", "保持原值", None, None),
                ("PO-003", "机房空调", 100, None, None, "保持原值", "保持原值", None, None),
            ],
        )

    def test_leaves_installation_row_empty_when_fallback_prices_conflict(self):
        self._write_air_conditioning_data([
            ["PO-001", "安装调试费（含辅料）", 100, None, None, None, None],
        ])
        self._write_air_conditioning_pivot([])
        self._write_price_list([
            ["HZHS AC 022", "安装调试费（含辅料）", 400, 376],
            ["NJLD AC 013", "安装调试费（含辅料）", 410, 376],
        ])

        success, message = self.module["fill_air_conditioning_old_new_prices"](str(self.price_path))

        self.assertTrue(success)
        self.assertIn("价格存在歧义", message)
        self.assertEqual(
            self._read_target_rows(),
            [("PO-001", "安装调试费（含辅料）", 100, None, None, None, None, None, None)],
        )

    def test_calculates_check_and_non_fld_saving_after_price_and_fld_po_overrides(self):
        self._write_air_conditioning_data([
            ["PO-BASE", "HZHS AC 003", 100, None, None, "Saving 保持", None, 2292.73, 2],
            ["PO-FLD", "HZHS AC 004", 350, None, None, "Saving 保持", None, 19695, 1],
            ["PO-FLD", "HZHS AC 005", 1900, None, None, "Saving 保持", None, 1900, 43],
        ])
        self._write_air_conditioning_pivot([
            ["PO-FLD", 19695, "FLD审批.msg", 24132],
        ])
        self._write_price_list([
            ["HZHS AC 003", "普通物料", 2439.07, 2292.73],
            ["HZHS AC 004", "普通物料", 400, 376],
            ["HZHS AC 005", "普通物料", 400, 376],
        ])

        success, message = self.module["fill_air_conditioning_old_new_prices"](str(self.price_path))

        self.assertTrue(success)
        self.assertIn("CHECK：已计算", message)
        self.assertEqual(
            self._read_target_rows(),
            [
                ("PO-BASE", "HZHS AC 003", 100, 2439.07, 2292.73, 292.68, 0, 2292.73, 2),
                ("PO-FLD", "HZHS AC 004", 350, 24132, 19695, 4437, 0, 19695, 1),
                ("PO-FLD", "HZHS AC 005", 1900, None, None, None, 17795, 1900, 43),
            ],
        )

    def test_second_run_preserves_completed_non_fld_price_saving_and_check(self):
        self._write_air_conditioning_data([
            ["PO-NO-FLD", "HZHS AC 003", 100, None, None, None, None, 350, 2],
        ])
        self._write_air_conditioning_pivot([])
        self._write_price_list([
            ["HZHS AC 003", "普通物料", 400, 376],
        ])

        success, _ = self.module["fill_air_conditioning_old_new_prices"](str(self.price_path))
        self.assertTrue(success)
        self.assertEqual(
            self._read_target_rows()[0][3:7],
            (400, 376, 100, 26),
        )

        workbook = openpyxl.load_workbook(self.excel_path)
        worksheet = workbook["空调数据"]
        worksheet["D2"] = 901
        worksheet["E2"] = 801
        worksheet["F2"] = 701
        worksheet["G2"] = 601
        worksheet["H2"] = 300
        worksheet["I2"] = 3
        workbook.save(self.excel_path)
        workbook.close()
        self._write_price_list([
            ["HZHS AC 003", "普通物料", 500, 450],
        ])

        success, _ = self.module["fill_air_conditioning_old_new_prices"](str(self.price_path))

        self.assertTrue(success)
        self.assertEqual(
            self._read_target_rows()[0][3:7],
            (901, 801, 701, 601),
        )

    def test_second_run_still_refreshes_fld_values_from_latest_pivot(self):
        self._write_air_conditioning_data([
            ["PO-FLD", "HZHS AC 003", 100, None, None, None, None, 100, 1],
        ])
        self._write_air_conditioning_pivot([
            ["PO-FLD", 200, "FLD审批.msg", 300],
        ])
        self._write_price_list([
            ["HZHS AC 003", "普通物料", 400, 376],
        ])

        success, _ = self.module["fill_air_conditioning_old_new_prices"](str(self.price_path))
        self.assertTrue(success)
        self.assertEqual(
            self._read_target_rows()[0][3:7],
            (300, 200, 100, 100),
        )

        workbook = openpyxl.load_workbook(self.excel_path)
        pivot = workbook["空调透视表"]
        pivot["B2"] = 250
        pivot["D2"] = 360
        workbook.save(self.excel_path)
        workbook.close()

        success, _ = self.module["fill_air_conditioning_old_new_prices"](str(self.price_path))

        self.assertTrue(success)
        self.assertEqual(
            self._read_target_rows()[0][3:7],
            (360, 250, 110, 150),
        )

    def test_support_filename_does_not_turn_ordinary_downloads_into_fld(self):
        self._write_air_conditioning_data([
            ["PO-ORDINARY", "HZHS AC 003", 100, 901, 801, 701, 601, 350, 2],
        ])
        self._write_air_conditioning_pivot([
            ["PO-ORDINARY", 999, "普通附件.msg", 1000],
        ])
        self._write_download_file("无FLD文件", "PO-ORDINARY", "普通附件.msg")
        # 模拟旧版本遗留：同一 PO 两边都有目录，但实际文件都不是 FLD。
        self._write_download_file("FLD文件", "PO-ORDINARY", "普通说明.pdf")
        self._write_price_list([
            ["HZHS AC 003", "普通物料", 500, 450],
        ])

        success, _ = self.module["fill_air_conditioning_old_new_prices"](str(self.price_path))

        self.assertTrue(success)
        self.assertEqual(
            self._read_target_rows()[0][3:7],
            (901, 801, 701, 601),
        )

    def test_actual_fld_file_refreshes_po_even_when_support_filename_is_ordinary(self):
        self._write_air_conditioning_data([
            ["PO-ACTUAL-FLD", "HZHS AC 003", 100, 901, 801, 701, 601, 100, 1],
        ])
        self._write_air_conditioning_pivot([
            ["PO-ACTUAL-FLD", 250, "普通附件.msg", 360],
        ])
        self._write_download_file(
            "无FLD文件", "PO-ACTUAL-FLD", "转发 FLD新物料价格审批.msg"
        )
        self._write_price_list([
            ["HZHS AC 003", "普通物料", 500, 450],
        ])

        success, _ = self.module["fill_air_conditioning_old_new_prices"](str(self.price_path))

        self.assertTrue(success)
        self.assertEqual(
            self._read_target_rows()[0][3:7],
            (360, 250, 110, 150),
        )

    def test_po_without_any_download_folder_is_reported_and_left_unchanged(self):
        self._write_air_conditioning_data([
            ["PO-MISSING", "HZHS AC 003", 100, 901, 801, 701, 601, 100, 1],
        ])
        self._write_air_conditioning_pivot([
            ["PO-MISSING", 250, "FLD审批.msg", 360],
        ])
        for bucket in ("FLD文件", "无FLD文件"):
            po_dir = self._download_po_dir(bucket, "PO-MISSING")
            if po_dir.exists():
                for child in po_dir.iterdir():
                    child.unlink()
                po_dir.rmdir()
        self._write_price_list([
            ["HZHS AC 003", "普通物料", 500, 450],
        ])

        success, message = self.module["fill_air_conditioning_old_new_prices"](str(self.price_path))

        self.assertTrue(success)
        self.assertIn("PO-MISSING", message)
        self.assertEqual(
            self._read_target_rows()[0][3:7],
            (901, 801, 701, 601),
        )

    def test_overwrites_every_detail_row_of_an_fld_po_from_pivot_totals(self):
        self._write_air_conditioning_data([
            ["PO-FLD", "HZHS AC 003", 350, None, None, "VLOOKUP 结果", "不修改"],
            ["PO-FLD", "HZHS AC 004", 1900, None, None, "VLOOKUP 结果", "不修改"],
            ["PO-NO-FLD", "HZHS AC 005", 500, None, None, "VLOOKUP 结果", "不修改"],
        ])
        self._write_air_conditioning_pivot([
            ["PO-FLD", 19695, "FLD审批.msg", 24132],
            ["PO-NO-FLD", 500, None, 100],
        ])
        self._write_price_list([
            ["HZHS AC 003", "普通物料", 400, 376],
            ["HZHS AC 004", "普通物料", 400, 376],
            ["HZHS AC 005", "普通物料", 400, 376],
        ])

        success, message = self.module["fill_air_conditioning_old_new_prices"](str(self.price_path))

        self.assertTrue(success)
        self.assertIn("FLD PO 覆盖：1 个 PO / 2 条明细", message)
        self.assertEqual(
            self._read_target_rows(),
            [
                ("PO-FLD", "HZHS AC 003", 350, 24132, 19695, 4437, "不修改", None, None),
                ("PO-FLD", "HZHS AC 004", 1900, None, None, None, "不修改", None, None),
                ("PO-NO-FLD", "HZHS AC 005", 500, 400, 376, "VLOOKUP 结果", "不修改", None, None),
            ],
        )

    def test_reads_fld_po_from_native_pivot_row_labels(self):
        self._write_air_conditioning_data([
            ["PO-FLD", "HZHS AC 003", 350, None, None, None, "不修改"],
        ])
        self._write_air_conditioning_pivot([
            ["Central", 19695, None, None],
            ["PO-FLD", 19695, "FLD审批.msg", 24132],
        ], native_row_label_layout=True)
        self._write_price_list([
            ["HZHS AC 003", "普通物料", 400, 376],
        ])

        success, _ = self.module["fill_air_conditioning_old_new_prices"](str(self.price_path))

        self.assertTrue(success)
        self.assertEqual(
            self._read_target_rows(),
            [("PO-FLD", "HZHS AC 003", 350, 24132, 19695, 4437, "不修改", None, None)],
        )

    def test_writes_fld_totals_to_the_anchor_of_existing_merged_detail_cells(self):
        self._write_air_conditioning_data([
            ["PO-FLD", "HZHS AC 003", 350, None, None, None, "不修改"],
            ["PO-FLD", "HZHS AC 004", 1900, None, None, None, "不修改"],
        ])
        workbook = openpyxl.load_workbook(self.excel_path)
        worksheet = workbook["空调数据"]
        for column in ("D", "E", "F"):
            worksheet.merge_cells(f"{column}2:{column}3")
        workbook.save(self.excel_path)
        workbook.close()
        self._write_air_conditioning_pivot([
            ["PO-FLD", 19695, "FLD审批.msg", 24132],
        ])
        self._write_price_list([
            ["HZHS AC 003", "普通物料", 400, 376],
            ["HZHS AC 004", "普通物料", 400, 376],
        ])

        success, _ = self.module["fill_air_conditioning_old_new_prices"](str(self.price_path))

        self.assertTrue(success)
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        worksheet = workbook["空调数据"]
        self.assertEqual((worksheet["D2"].value, worksheet["E2"].value, worksheet["F2"].value), (24132, 19695, 4437))
        self.assertEqual((worksheet["G2"].value, worksheet["G3"].value), ("不修改", "不修改"))
        workbook.close()

    def test_merges_the_three_fld_total_columns_for_each_contiguous_po_group(self):
        self._write_air_conditioning_data([
            ["PO-FLD", "HZHS AC 003", 350, None, None, None, None, 19695, 1],
            ["PO-FLD", "HZHS AC 004", 1900, None, None, None, None, 1900, 43],
            ["PO-NO-FLD", "HZHS AC 005", 500, None, None, None, None, 376, 2],
        ])
        self._write_air_conditioning_pivot([
            ["PO-FLD", 19695, "FLD审批.msg", 24132],
        ])
        self._write_price_list([
            ["HZHS AC 003", "普通物料", 400, 376],
            ["HZHS AC 004", "普通物料", 400, 376],
            ["HZHS AC 005", "普通物料", 400, 376],
        ])

        success, message = self.module["fill_air_conditioning_old_new_prices"](str(self.price_path))

        self.assertTrue(success)
        self.assertIn("FLD 分组合并", message)
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        worksheet = workbook["空调数据"]
        merged_ranges = {str(cell_range) for cell_range in worksheet.merged_cells.ranges}
        self.assertTrue({"D2:D3", "E2:E3", "F2:F3"}.issubset(merged_ranges))
        self.assertEqual((worksheet["D2"].value, worksheet["E2"].value, worksheet["F2"].value), (24132, 19695, 4437))
        self.assertEqual((worksheet["G2"].value, worksheet["G3"].value), (0, 17795))
        self.assertNotIn("D4:D4", merged_ranges)
        workbook.close()

        # 同一按钮重复执行时，既不叠加合并范围，也不能破坏 CHECK 的逐行结果。
        success, _ = self.module["fill_air_conditioning_old_new_prices"](str(self.price_path))
        self.assertTrue(success)
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        worksheet = workbook["空调数据"]
        self.assertEqual(
            {str(cell_range) for cell_range in worksheet.merged_cells.ranges},
            merged_ranges,
        )
        self.assertEqual((worksheet["G2"].value, worksheet["G3"].value), (0, 17795))
        workbook.close()

    def test_keeps_existing_fld_merge_unchanged_when_another_column_crosses_poes(self):
        self._write_air_conditioning_data([
            ["PO-FLD", "HZHS AC 003", 350, None, None, None, None, 19695, 1],
            ["PO-FLD", "HZHS AC 004", 1900, None, None, None, None, 1900, 43],
            ["PO-NO-FLD", "HZHS AC 005", 500, None, None, None, None, 376, 2],
        ])
        workbook = openpyxl.load_workbook(self.excel_path)
        worksheet = workbook["空调数据"]
        worksheet.merge_cells("D2:D3")
        worksheet.merge_cells("E2:E4")
        workbook.save(self.excel_path)
        workbook.close()
        self._write_air_conditioning_pivot([
            ["PO-FLD", 19695, "FLD审批.msg", 24132],
        ])
        self._write_price_list([
            ["HZHS AC 003", "普通物料", 400, 376],
            ["HZHS AC 004", "普通物料", 400, 376],
            ["HZHS AC 005", "普通物料", 400, 376],
        ])

        success, _ = self.module["fill_air_conditioning_old_new_prices"](str(self.price_path))

        self.assertTrue(success)
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        worksheet = workbook["空调数据"]
        merged_ranges = {str(cell_range) for cell_range in worksheet.merged_cells.ranges}
        self.assertIn("D2:D3", merged_ranges)
        self.assertIn("E2:E4", merged_ranges)
        workbook.close()

    def test_uses_a_valid_fld_summary_when_another_pivot_row_is_blank(self):
        self._write_air_conditioning_data([
            ["PO-FLD", "HZHS AC 003", 350, None, None, None, "不修改"],
        ])
        self._write_air_conditioning_pivot([
            ["PO-FLD", None, "FLD审批.msg", None],
            ["PO-FLD", 19695, "FLD审批.msg", 24132],
        ])
        self._write_price_list([
            ["HZHS AC 003", "普通物料", 400, 376],
        ])

        success, _ = self.module["fill_air_conditioning_old_new_prices"](str(self.price_path))

        self.assertTrue(success)
        self.assertEqual(
            self._read_target_rows(),
            [("PO-FLD", "HZHS AC 003", 350, 24132, 19695, 4437, "不修改", None, None)],
        )


if __name__ == "__main__":
    unittest.main()
