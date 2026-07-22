import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

import web_query


class PoExtractionTests(unittest.TestCase):
    def test_prefers_the_small_category_data_sheet_and_uses_read_only_mode(self):
        """查询前应直接读品类数据 Sheet，不能扫描整个透视表。"""
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "records.xlsx"
            workbook = openpyxl.Workbook()
            sheet1 = workbook.active
            sheet1.title = "Sheet1"
            sheet1.append(["采购凭证", "订单净值"])
            sheet1.append(["4000000001", 100])
            sheet1.append(["4000000002", 200])

            data_sheet = workbook.create_sheet("外包板数据")
            data_sheet.append(["采购凭证", "订单净值"])
            data_sheet.append(["4000000002", 200])
            data_sheet.append(["4000000001", 100])
            data_sheet.append(["4000000002", 200])

            pivot_sheet = workbook.create_sheet("外包板透视表")
            pivot_sheet.append(["采购凭证", "金额"])
            pivot_sheet.append(["4000000001", 100])
            pivot_sheet.append(["9999999999", 999])
            workbook.save(path)
            workbook.close()

            real_loader = openpyxl.load_workbook
            load_calls = []

            def tracked_loader(*args, **kwargs):
                load_calls.append(kwargs.copy())
                return real_loader(*args, **kwargs)

            with patch("openpyxl.load_workbook", side_effect=tracked_loader):
                result = web_query.extract_all_pos_from_pivot(
                    excel_path=str(path), sheet_name="外包板透视表"
                )

            self.assertEqual(result, ["4000000002", "4000000001"])
            self.assertTrue(load_calls[0]["read_only"])
            self.assertTrue(load_calls[0]["data_only"])


if __name__ == "__main__":
    unittest.main()
