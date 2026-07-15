import gc
import runpy
import tempfile
import unittest
from pathlib import Path

import openpyxl


PROJECT_DIR = Path(__file__).resolve().parents[1]
GUI_SCRIPT = next(
    path
    for path in PROJECT_DIR.glob("*.py")
    if path.name != "web_query.py" and path.stat().st_size > 50_000
)


class AirConditioningPmTrackingTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.module = runpy.run_path(str(GUI_SCRIPT))
        self.match_pm_tracking_data = self.module["match_pm_tracking_data"]
        self.module_globals = self.match_pm_tracking_data.__globals__
        self.excel_path = Path(self.temp_dir.name) / "采购记录.xlsx"
        self.tracking_path = Path(self.temp_dir.name) / "NI PM Saving Tracking.xlsx"
        self.module_globals["EXCEL_FILE"] = str(self.excel_path)
        self.module_globals["TRACKING_FILE"] = str(self.tracking_path)

    def tearDown(self):
        gc.collect()
        self.temp_dir.cleanup()

    def _write_pivot(self, category, rows):
        target_sheet = self.module["CATEGORIES"][category]["target_sheet"]
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = target_sheet
        worksheet.append([
            "采购凭证",
            "求和项:订单净值",
            "E2E项目名",
            "项目总金额(审批)",
            "支持文件名(FLD 审批)",
            "Price",
            "PlanCost",
            "总saving",
            "订单是否下完=price-订单净值",
        ])
        for row in rows:
            values = list(row)
            if len(values) == 8:
                # 旧测试数据未显式给出审批金额；插入空值以保持原字段顺序。
                values.insert(3, None)
            worksheet.append(values)
        workbook.save(self.excel_path)
        workbook.close()

    def _write_tracking(self, rows):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Base Data"
        worksheet.append(["无关表头"])
        worksheet.append(["Project Name", "Price", "PlanCost", "Content"])
        for row in rows:
            worksheet.append(row)
        workbook.save(self.tracking_path)
        workbook.close()

    def _run_air_conditioning_match(self):
        return self.match_pm_tracking_data(category="空调")

    def _read_values(self, row_number):
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        worksheet = workbook[self.module["CATEGORIES"]["空调"]["target_sheet"]]
        headers = {
            str(worksheet.cell(1, col).value): col
            for col in range(1, worksheet.max_column + 1)
        }
        values = tuple(
            worksheet.cell(row_number, headers[header]).value
            for header in ("Price", "PlanCost", "总saving", "订单是否下完=price-订单净值")
        )
        workbook.close()
        return values

    def test_air_conditioning_matches_manual_non_fld_project_when_e2e_name_exists(self):
        self._write_pivot("空调", [
            ["PO-FLD", 200, "空调项目A", "FLD审批.msg", None, None, None, None],
            ["PO-NO-FLD", 300, "空调项目B", None, None, None, None, None],
        ])
        self._write_tracking([
            ["空调项目A-采购", 500, 100, "AC空调"],
            ["空调项目B-采购", 900, 180, "AC空调"],
        ])

        success, _ = self._run_air_conditioning_match()

        self.assertTrue(success)
        self.assertEqual(self._read_values(2), (500, 100, 40, 300))
        self.assertEqual(self._read_values(3), (900, 180, 60, 600))

    def test_air_conditioning_accepts_exact_ac_content_when_project_has_multiple_rows(self):
        self._write_pivot("空调", [
            ["PO-AC", 200, "空调项目AC", "FLD审批.msg", None, None, None, None],
        ])
        self._write_tracking([
            ["空调项目AC-采购", 600, 120, "AC"],
            ["空调项目AC-历史", 700, 110, "AC空调-历史"],
        ])

        success, _ = self._run_air_conditioning_match()

        self.assertTrue(success)
        self.assertEqual(self._read_values(2), (600, 120, 40, 400))

    def test_air_conditioning_uses_exact_ac_content_when_project_has_multiple_rows(self):
        self._write_pivot("空调", [
            ["PO-AC", 200, "空调项目C", "FLD审批.msg", None, None, None, None],
        ])
        self._write_tracking([
            ["空调项目C-采购", 600, 120, "AC空调"],
            ["空调项目C-历史", 700, 110, "AC空调-历史"],
        ])

        success, _ = self._run_air_conditioning_match()

        self.assertTrue(success)
        self.assertEqual(self._read_values(2), (600, 120, 40, 400))

    def test_air_conditioning_skips_ambiguous_rows_without_an_exact_ac_content_match(self):
        self._write_pivot("空调", [
            ["PO-AMBIGUOUS", 200, "空调项目D", "FLD审批.msg", None, None, None, None],
        ])
        self._write_tracking([
            ["空调项目D-历史", 700, 110, "AC空调-历史"],
            ["空调项目D-采购", 600, 120, "装潢"],
        ])

        success, message = self._run_air_conditioning_match()

        self.assertTrue(success)
        self.assertEqual(self._read_values(2), (None, None, None, None))
        self.assertIn("PO-AMBIGUOUS", message)

    def test_air_conditioning_skips_ambiguous_rows_with_multiple_exact_ac_content_matches(self):
        self._write_pivot("空调", [
            ["PO-DUPLICATE-AC", 200, "空调项目F", "FLD审批.msg", None, None, None, None],
        ])
        self._write_tracking([
            ["空调项目F-采购", 600, 120, "AC空调"],
            ["空调项目F-追加", 700, 110, "AC空调"],
        ])

        success, message = self._run_air_conditioning_match()

        self.assertTrue(success)
        self.assertEqual(self._read_values(2), (None, None, None, None))
        self.assertIn("PO-DUPLICATE-AC", message)

    def test_air_conditioning_skips_project_when_ac_and_ac_air_conditioning_both_match(self):
        self._write_pivot("空调", [
            ["PO-AC-BOTH", 200, "空调项目双候选", "FLD审批.msg", None, None, None, None],
        ])
        self._write_tracking([
            ["空调项目双候选-AC", 600, 120, "AC"],
            ["空调项目双候选-AC空调", 700, 140, "AC空调"],
        ])

        success, message = self._run_air_conditioning_match()

        self.assertTrue(success)
        self.assertEqual(self._read_values(2), (None, None, None, None))
        self.assertIn("PO-AC-BOTH", message)

    def test_reports_price_vs_approval_amount_mismatches_and_missing_values_after_matching(self):
        self._write_pivot("空调", [
            ["PO-EQUAL", 200, "空调项目一致", 500.01, None, None, None, None, None],
            ["PO-DIFFERENT", 200, "空调项目不一致", 500, None, None, None, None, None],
            ["PO-MISSING", 200, "空调项目金额缺失", None, None, None, None, None, None],
        ])
        self._write_tracking([
            ["空调项目一致-采购", 500, 100, "AC"],
            ["空调项目不一致-采购", 500.02, 100, "AC"],
            ["空调项目金额缺失-采购", 300, 100, "AC"],
        ])

        success, message = self._run_air_conditioning_match()

        self.assertTrue(success)
        self.assertNotIn("PO-EQUAL", message)
        self.assertIn("PO-DIFFERENT", message)
        self.assertIn("PO-MISSING", message)
        self.assertIn("Price 与项目总金额(审批)不一致", message)
        self.assertIn("Price 与项目总金额(审批)无法比较", message)

    def test_air_conditioning_rejects_a_unique_project_match_when_content_differs(self):
        self._write_pivot("空调", [
            ["PO-UNIQUE", 200, "空调项目E", "FLD审批.msg", None, None, None, None],
        ])
        self._write_tracking([
            ["空调项目E-采购", 500, 100, "装潢"],
        ])

        success, message = self._run_air_conditioning_match()

        self.assertTrue(success)
        self.assertEqual(self._read_values(2), (None, None, None, None))
        self.assertIn("PO-UNIQUE", message)

    def test_air_conditioning_fuzzy_match_filters_content_before_scoring(self):
        project_name = (
            "示例地块一标段设备采购及相关服务"
            "示例城市项目一期设备更新项目"
        )
        self._write_pivot("空调", [
            ["PO-COMBINED", 200, project_name, "FLD审批.msg", None, None, None, None],
        ])
        self._write_tracking([
            ["示例城市项目一期设备更新项目", 139888, 152338, "装潢"],
            ["示例地块一标段设备采购及相关服务", 172900, 176930, "装潢"],
            [
                "示例地块一标段设备采购及相关服务&示例城市项目&示例城市项目一期设备更新项目",
                314000,
                319134,
                "AC空调",
            ],
        ])

        success, _ = self._run_air_conditioning_match()

        self.assertTrue(success)
        self.assertEqual(self._read_values(2)[:2], (314000, 319134))

    def test_decoration_matching_behavior_is_unchanged(self):
        self._write_pivot("装潢", [
            ["PO-DECORATION", 200, "装潢项目", None, None, None, None, None],
        ])
        self._write_tracking([
            ["装潢项目-采购", 500, 100, "装潢-华东"],
            ["装潢项目-其他", 700, 100, "空调"],
        ])

        success, _ = self.match_pm_tracking_data(category="装潢")

        self.assertTrue(success)
        workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        worksheet = workbook[self.module["CATEGORIES"]["装潢"]["target_sheet"]]
        headers = {
            str(worksheet.cell(1, col).value): col
            for col in range(1, worksheet.max_column + 1)
        }
        self.assertEqual(
            tuple(
                worksheet.cell(2, headers[header]).value
                for header in ("Price", "PlanCost", "总saving", "订单是否下完=price-订单净值")
            ),
            (500, 100, 40, 300),
        )
        workbook.close()


if __name__ == "__main__":
    unittest.main()
