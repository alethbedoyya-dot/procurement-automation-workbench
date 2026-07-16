# -*- coding: utf-8 -*-
"""
TKE VIEW 网站自动化查询模块
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
功能：
  1. 从「装潢透视表」Sheet 提取全部采购凭证号
  2. Playwright 启动 Edge → 登录 TKE VIEW（Microsoft SSO，会话持久化）
  3. 逐个 PO：跳转搜索页 → 填PO号 → 查找 → 点击请求ID → 详情页
  4. 详情页提取：项目名称 / 梯台明细合计数量 / 下载全部附件
  5. 附件分类：FLD 开头 → downloads/FLD文件/<PO号>/
               非 FLD   → downloads/无FLD文件/<PO号>/
  6. FLD .msg 解析审批价格 → 计算差异 → 写回 Excel

会话管理：
  → 首次运行需手动完成 Microsoft 账户登录
  → 登录后自动保存 cookies + localStorage 到 auth.json
  → 后续运行加载 auth.json，跳过登录
"""

import os
import re
import json
import sys
import time
import subprocess
import traceback
import threading
import weakref
from urllib.parse import parse_qs, urlparse

# ═══════════════════ 配置 ═══════════════════

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(SCRIPT_DIR, "采购记录0701.xlsx")
TARGET_SHEET = "装潢透视表"
HOME_URL = "https://view.tkelevator.com.cn/"
SEARCH_URL = "https://view.tkelevator.com.cn/vivid/niops/purchasing/materials/search"
AUTH_FILE = os.path.join(SCRIPT_DIR, "auth.json")
RUN_LOG_FILE = os.path.join(SCRIPT_DIR, "automation_run_log.txt")
BACKFILL_RUN_LOG_FILE = os.path.join(SCRIPT_DIR, "backfill_run_log.txt")
DOWNLOAD_DIR = os.path.join(SCRIPT_DIR, "downloads")
FLD_DIR = os.path.join(DOWNLOAD_DIR, "FLD文件")
NO_FLD_DIR = os.path.join(DOWNLOAD_DIR, "无FLD文件")
PAGE_SCAN_FILE = os.path.join(SCRIPT_DIR, "page_scan.txt")
PENDING_BACKFILL_FILENAME = "待回填任务.json"
COMPLETED_BACKFILL_PREFIX = "已完成回填任务_"
PDF_PARSE_TIMEOUT = 45
OCR_TIMEOUT = 60
TARGET_MSG_PREFIXES = ("fld",)
APPROVAL_PRICE_LABEL = "\u5ba1\u6279\u4ef7\u683c"

# 详情页会被复用为下一轮的搜索页。焦点模拟属于 CDP 会话级状态，
# 因此会话必须一直保留到该页面真正关闭时才释放。弱引用避免已关闭页面
# 因全局缓存而常驻内存；锁避免 GUI 工作线程出现并发访问时相互覆盖。
_PAGE_ACTIVITY_SESSIONS = weakref.WeakKeyDictionary()
_PAGE_ACTIVITY_SESSIONS_LOCK = threading.RLock()

# 仅传给本自动化启动的 Edge：关闭后台标签页的渲染/计时器节流，
# 防止标签页轮换时被 Windows 遮挡判定为后台而停止响应。
EDGE_BACKGROUND_PROTECTION_ARGS = (
    "--disable-background-timer-throttling",
    "--disable-renderer-backgrounding",
    "--disable-backgrounding-occluded-windows",
)

# 请求 ID 已点击后，部分页面会因站点响应或 Edge 标签页调度而晚于
# expect_page 的 12 秒窗口出现。公司实测最慢详情页在点击约 118 秒后才创建；
# 这里预留到 150 秒，仅观察既有页面，不发起第二次点击，因而不会重现一条请求
# 打开多个详情页的问题。正常 PO 在 expect_page 成功后立即继续，不受此值影响。
REQUEST_ID_LATE_PAGE_WAIT_SECONDS = 150

# ── 测试模式 ──
TEST_MODE = False       # True=仅处理前N个PO, False=全量
TEST_LIMIT = 4           # 测试模式下处理的PO数量


# ═══════════════════ 查询/回填任务清单 ═══════════════════

def _is_stop_requested(stop_requested):
    """兼容 threading.Event 或任意返回布尔值的停止检查函数。"""
    if stop_requested is None:
        return False
    try:
        return bool(stop_requested()) if callable(stop_requested) else bool(stop_requested.is_set())
    except Exception:
        return False


def _category_label(category_label=None):
    """返回下载目录和任务清单使用的品类名。"""
    return str(category_label or "装潢").strip() or "装潢"


def _category_download_dir(category_label=None):
    return os.path.join(DOWNLOAD_DIR, _category_label(category_label))


def _pending_backfill_path(category_label=None):
    return os.path.join(
        _category_download_dir(category_label), PENDING_BACKFILL_FILENAME
    )


def _configure_category_download_paths(category_label=None):
    """让网页下载和后续回填使用同一个品类目录。"""
    global FLD_DIR, NO_FLD_DIR
    category_dir = _category_download_dir(category_label)
    FLD_DIR = os.path.join(category_dir, "FLD文件")
    NO_FLD_DIR = os.path.join(category_dir, "无FLD文件")


def _normalise_po_list(values):
    """保留顺序地去重有效 PO，避免目录扫描受到无关文件夹干扰。"""
    pos = []
    seen = set()
    for value in values or []:
        po = str(value or "").strip()
        if re.fullmatch(r"\d{7,12}", po) and po not in seen:
            seen.add(po)
            pos.append(po)
    return pos


def audit_download_completeness(expected_pos, category_label=None, failed_pos=None):
    """按本次透视表 PO 与实际 PO 目录核对下载完整性。

    不能只比较文件夹数量：旧运行留下的目录、两个分类目录和“无附件但已
    查询”的 PO 都会让数量比较失真。目录存在表示该 PO 已进入下载阶段；
    本轮明确失败的 PO 即使恰好存在旧目录，仍会被列为待补查。
    """
    category_dir = _category_download_dir(category_label)
    expected = _normalise_po_list(expected_pos)
    expected_set = set(expected)
    folder_set = set()
    for bucket in ("FLD文件", "无FLD文件"):
        bucket_dir = os.path.join(category_dir, bucket)
        if not os.path.isdir(bucket_dir):
            continue
        try:
            for entry in os.scandir(bucket_dir):
                if entry.is_dir() and entry.name in expected_set:
                    folder_set.add(entry.name)
        except OSError:
            continue

    explicitly_failed = set(_normalise_po_list(failed_pos)) & expected_set
    folder_missing = expected_set - folder_set
    missing_set = folder_missing | explicitly_failed
    return {
        "expected_pos": expected,
        "folder_pos": [po for po in expected if po in folder_set],
        "folder_missing_pos": [po for po in expected if po in folder_missing],
        "failed_pos": [po for po in expected if po in explicitly_failed],
        "missing_pos": [po for po in expected if po in missing_set],
    }


def merge_query_results(existing_results, new_results, expected_pos, replaced_pos=None):
    """将补查结果按 PO 覆盖并回原任务，避免重复回填同一 PO。"""
    allowed_pos = set(_normalise_po_list(expected_pos))
    replaced = set(_normalise_po_list(replaced_pos))
    by_po = {}
    for result in existing_results or []:
        if not isinstance(result, dict):
            continue
        po = str(result.get("po") or "").strip()
        if po in allowed_pos and po not in replaced:
            by_po[po] = result
    for result in new_results or []:
        if not isinstance(result, dict):
            continue
        po = str(result.get("po") or "").strip()
        if po in allowed_pos:
            by_po[po] = result

    ordered_results = []
    for po in _normalise_po_list(expected_pos):
        if po in by_po:
            ordered_results.append(by_po[po])
    return ordered_results


def retry_missing_pos(
    target_sheet=None, category_label=None, status_callback=None, stop_requested=None,
):
    """只重查实际未完成的 PO；支持程序更新前留下的旧下载目录。"""
    ts = target_sheet or TARGET_SHEET
    category = _category_label(category_label)
    _configure_category_download_paths(category)
    expected_pos = _normalise_po_list(extract_all_pos_from_pivot(EXCEL_FILE, ts))
    if not expected_pos:
        return False, f"未从「{ts}」提取到可补查的 PO。"

    manifest = None
    manifest_path = _pending_backfill_path(category)
    if os.path.isfile(manifest_path):
        manifest, error = _load_pending_backfill_manifest(category, ts)
        if error:
            return False, error

    audit = audit_download_completeness(
        expected_pos,
        category,
        failed_pos=(manifest or {}).get("failed_pos", []),
    )
    missing_pos = audit["missing_pos"]
    if not manifest and not audit["folder_pos"]:
        return False, "未发现历史查询记录，请先点击「③ 打开网站查询」。"
    if not missing_pos:
        return True, f"完整性检查通过：{len(expected_pos)} 个 PO 均已有下载目录，无需补查。"

    if status_callback:
        status_callback("待补查 PO：" + "、".join(missing_pos))
    success, message = query_and_download_attachments(
        target_sheet=ts,
        category_label=category,
        status_callback=status_callback,
        stop_requested=stop_requested,
        po_numbers=missing_pos,
        resume_manifest=manifest,
        is_recovery_run=True,
    )
    return success, "补查 PO：" + "、".join(missing_pos) + "\n\n" + message


def _save_pending_backfill_manifest(manifest):
    """原子写入待回填任务，避免中断时留下半截 JSON。"""
    category = _category_label(manifest.get("category_label"))
    path = _pending_backfill_path(category)
    os.makedirs(os.path.dirname(path), exist_ok=True)

    payload = dict(manifest)
    payload["category_label"] = category
    payload["saved_at"] = time.strftime("%Y-%m-%d %H:%M:%S")
    temp_path = f"{path}.tmp"
    try:
        with open(temp_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        os.replace(temp_path, path)
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass
    return path


def _load_pending_backfill_manifest(category_label, target_sheet):
    """读取并校验当前品类的待回填任务。"""
    category = _category_label(category_label)
    path = _pending_backfill_path(category)
    if not os.path.exists(path):
        other_categories = []
        if os.path.isdir(DOWNLOAD_DIR):
            for entry in os.scandir(DOWNLOAD_DIR):
                if entry.is_dir() and entry.name != category:
                    other_path = os.path.join(entry.path, PENDING_BACKFILL_FILENAME)
                    if os.path.isfile(other_path):
                        other_categories.append(entry.name)
        if other_categories:
            return None, (
                f"未找到「{category}」的待回填任务；发现「{', '.join(other_categories)}」"
                "的任务，请切换到对应品类后回填。"
            )
        return None, f"未找到「{category}」的待回填任务，请先执行“查询并下载附件”。"

    try:
        with open(path, "r", encoding="utf-8") as f:
            manifest = json.load(f)
    except (OSError, json.JSONDecodeError) as e:
        return None, f"待回填任务文件无法读取：{e}"

    if not isinstance(manifest, dict):
        return None, "待回填任务文件格式错误，请重新执行“查询并下载附件”。"
    if manifest.get("category_label") != category:
        return None, "待回填任务的品类与当前选择不一致，请重新执行查询。"
    if manifest.get("target_sheet") != target_sheet:
        return None, (
            f"待回填任务目标为「{manifest.get('target_sheet', '未知')}」，"
            f"当前目标为「{target_sheet}」，请重新执行查询。"
        )
    if not isinstance(manifest.get("results"), list):
        return None, "待回填任务缺少结果列表，请重新执行“查询并下载附件”。"
    return manifest, None


def _archive_pending_backfill_manifest(category_label):
    """将已完整回填的任务归档，避免下次误重复执行。"""
    path = _pending_backfill_path(category_label)
    category_dir = os.path.dirname(path)
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    archived_path = os.path.join(
        category_dir, f"{COMPLETED_BACKFILL_PREFIX}{timestamp}.json"
    )
    suffix = 1
    while os.path.exists(archived_path):
        archived_path = os.path.join(
            category_dir, f"{COMPLETED_BACKFILL_PREFIX}{timestamp}_{suffix}.json"
        )
        suffix += 1
    os.replace(path, archived_path)
    return archived_path


# ═══════════════════ PO 号提取 ═══════════════════

def extract_all_pos_from_pivot(excel_path=None, sheet_name=None):
    """
    从透视表 Sheet 中提取所有有效的采购凭证号（去重）。
    用 Sheet1「采购凭证」列做白名单过滤，排除透视表中的干扰数字（金额/数量等）。
    """
    import openpyxl

    path = excel_path or EXCEL_FILE
    name = sheet_name or TARGET_SHEET

    if not os.path.exists(path):
        return []

    wb = openpyxl.load_workbook(path, data_only=True)

    # ── 步骤1: 从 Sheet1 读取合法 PO 白名单 ──
    valid_pos = set()
    if "Sheet1" in wb.sheetnames:
        ws1 = wb["Sheet1"]
        po_col = None
        for col_idx in range(1, ws1.max_column + 1):
            header = str(ws1.cell(1, col_idx).value or "").strip()
            if "采购凭证" in header:
                po_col = col_idx
                break
        if po_col:
            for row_idx in range(2, ws1.max_row + 1):
                val = str(ws1.cell(row_idx, po_col).value or "").strip()
                if val and re.match(r'^\d{7,12}$', val):
                    valid_pos.add(val)

    # ── 步骤2: 从透视表提取数字，只保留白名单中的 ──
    if name not in wb.sheetnames:
        wb.close()
        return []

    ws = wb[name]
    seen = set()
    pos = []
    skipped = 0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            val = str(cell.value or "").strip()
            if val and re.match(r'^\d{7,12}$', val) and val not in seen:
                seen.add(val)
                if valid_pos and val not in valid_pos:
                    skipped += 1
                    continue
                pos.append(val)

    wb.close()
    if skipped:
        print(f"[PO提取] 已过滤 {skipped} 个非 PO 数字（如金额/数量等）")
    return pos


# ═══════════════════ 辅助函数 ═══════════════════

def _col_letter(idx):
    """1→A, 26→Z, 27→AA, ..."""
    s = ""
    while idx > 0:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


def _is_on_login_page(page):
    """判断当前页面是否为登录页面。"""
    url = page.url.lower()
    return (
        "login.php" in url
        or "login.microsoftonline" in url
        or "microsoft" in url
    )


# ═══════════════════ 元素定位 ═══════════════════

def _locate_po_input(page, log):
    """多策略定位「PO号」输入框，返回第一个可见 input 或 None。"""
    selectors = [
        "//label[contains(text(),'PO号')]/following-sibling::*/descendant::input[1]",
        "//label[contains(text(),'PO号')]/following::input[1]",
        "//td[contains(text(),'PO')]/following-sibling::td//input[1]",
        "//th[contains(text(),'PO')]/following-sibling::td//input[1]",
        "//span[contains(text(),'PO号')]/ancestor::td/following-sibling::td//input[1]",
        "input[placeholder*='PO']",
        "input[placeholder*='po']",
        "input[name*='po']",
        "input[name*='PO']",
        "input[name*='purchase']",
        "input[id*='po']",
        "input[id*='PO']",
        "//label[contains(text(),'PO')]/following::input[1]",
        "//*[contains(text(),'PO号')]/following::input[1]",
    ]
    for sel in selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=800):
                log(f"  ✓ PO输入框定位成功（{sel[:50]}...）")
                return el
        except Exception:
            continue
    return None


def _locate_search_button(page, log):
    """多策略定位「查找」按钮。找不到时诊断页面所有可见按钮。"""
    selectors = [
        "button:has-text('查找')",
        "//button[contains(text(),'查找')]",
        "//button[normalize-space(text())='查找']",
        "//a[contains(text(),'查找')]",
        "//input[@type='submit'][@value='查找']",
        "//input[@type='button'][@value='查找']",
        "//*[@role='button'][contains(text(),'查找')]",
        "button:has-text('查')",
        "//button[contains(@class,'search')]",
        "//button[contains(@class,'btn')][contains(text(),'查')]",
        "//*[contains(@class,'el-button')][contains(text(),'查找')]",
    ]
    for sel in selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=800):
                log(f"  ✓ 「查找」按钮定位成功（{sel[:50]}）")
                return el
        except Exception:
            continue

    # ── 诊断：列出所有可见按钮 ──
    try:
        all_btns = page.locator("button:visible").all()
        log(f"[诊断] 共 {len(all_btns)} 个可见按钮：")
        for i, btn in enumerate(all_btns[:15]):
            try:
                txt = (btn.inner_text() or "").strip()[:40]
                cls = (btn.get_attribute("class") or "")[:40]
                log(f"  [{i}] text='{txt}' | class='{cls}'")
            except Exception:
                pass
        input_btns = page.locator("input[type='submit']:visible, input[type='button']:visible").all()
        if input_btns:
            log(f"[诊断] 共 {len(input_btns)} 个可见 input 按钮：")
            for i, ib in enumerate(input_btns[:10]):
                try:
                    v = (ib.get_attribute("value") or "")[:40]
                    log(f"  [{i}] value='{v}'")
                except Exception:
                    pass
    except Exception:
        pass

    return None


def _wait_for_search_results(page, log, max_retries=2):
    """等待搜索结果出现，支持重试。"""
    wait_selectors = [
        "//th[contains(text(),'请求ID')]",
        "//a[contains(text(),'LM')]",
        "table:visible",
    ]
    for attempt in range(max_retries):
        if attempt > 0:
            log(f"  搜索结果未出现，第 {attempt + 1} 次重试...")
            try:
                search_btn = _locate_search_button(page, log)
                if search_btn:
                    search_btn.click()
            except Exception:
                pass
            time.sleep(3)
        time.sleep(2)
        for sel in wait_selectors:
            try:
                page.wait_for_selector(sel, state="visible", timeout=8000)
                log("  ✓ 检测到搜索结果")
                return True
            except Exception:
                continue
        time.sleep(3)
    return False


# ═══════════════════ 请求ID点击 & 导航 ═══════════════════

def _request_id_digits(request_id):
    return "".join(re.findall(r"\d+", str(request_id or "")))


def _is_detail_url_for_request(url, expected_request_id):
    """判断 URL 是否是指定 LM 请求对应的物料审批详情页。"""
    expected_digits = _request_id_digits(expected_request_id)
    if not expected_digits:
        return False
    try:
        parsed = urlparse(str(url or ""))
        if not parsed.path.rstrip("/").endswith("/materials/requests/approval"):
            return False
        actual_request_id = parse_qs(parsed.query).get("requestId", [""])[0]
        return _request_id_digits(actual_request_id) == expected_digits
    except Exception:
        return False


def _find_browser_target_for_request(context, request_id):
    """从 Chromium 的浏览器级目标列表定位尚未进入 ``context.pages`` 的详情页。

    少数 TKE 请求会先在 Edge 窗口中展示详情页，再晚很久才注册给 Playwright
    的浏览器上下文。浏览器级 CDP 目标列表不依赖该注册过程，能够在页面已肉眼
    可见时立即取得其真实 URL。
    """
    browser = getattr(context, "browser", None)
    if browser is None:
        return None
    session = None
    try:
        session = browser.new_browser_cdp_session()
        targets = session.send("Target.getTargets").get("targetInfos", [])
        for target in targets:
            if target.get("type") != "page":
                continue
            if _is_detail_url_for_request(target.get("url"), request_id):
                return target
    except Exception:
        return None
    finally:
        if session is not None:
            try:
                session.detach()
            except Exception:
                pass
    return None


def _close_browser_target(context, target_id):
    """关闭已被受控页面接管的外部 Edge 标签，避免留下重复详情页。"""
    if not target_id:
        return
    browser = getattr(context, "browser", None)
    if browser is None:
        return
    session = None
    try:
        session = browser.new_browser_cdp_session()
        session.send("Target.closeTarget", {"targetId": target_id})
    except Exception:
        pass
    finally:
        if session is not None:
            try:
                session.detach()
            except Exception:
                pass


def _find_detail_page_for_request(context, request_id):
    """在所有标签页中定位当前请求 ID 的详情页。

    TKE 有时不会创建新标签，而是复用一张已经存在的空白/旧详情标签。
    这时 Playwright 不会触发 ``expect_page``，仅比较“新页面”也会漏掉它。
    通过详情 URL 的 requestId 与当前 LM 请求 ID 精确比对，可避免误把其他
    PO 的残留详情页作为本轮结果。
    """
    try:
        pages = list(context.pages)
    except Exception:
        return None
    for candidate in pages:
        try:
            if candidate.is_closed():
                continue
            if _is_detail_url_for_request(candidate.url, request_id):
                return candidate
        except Exception:
            continue
    return None

def _click_request_id(page, context, log):
    """
    在搜索结果中找到请求ID并点击，等待新标签页打开。
    返回 (请求ID文本, 详情页page对象 或 None)。
    """
    request_el = None
    selectors = [
        "//a[contains(text(),'LM')]",
        "//td//a",
        "//a[contains(@href,'request')]",
        "//td[contains(text(),'LM')]",
        "//span[contains(text(),'LM')]",
        "//*[contains(text(),'LM80')]",
    ]
    for sel in selectors:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=1000):
                text = el.inner_text().strip()
                if len(text) >= 6:
                    request_el = el
                    log(f"  找到请求ID: {text} (selector: {sel[:40]})")
                    break
        except Exception:
            continue

    # ── 诊断：列出搜索页所有链接 ──
    if request_el is None:
        try:
            all_links = page.locator("a:visible").all()
            log(f"[诊断] 搜索结果页共 {len(all_links)} 个可见链接：")
            for i, link in enumerate(all_links[:15]):
                try:
                    txt = (link.inner_text() or "").strip()[:50]
                    href = (link.get_attribute("href") or "")[:60]
                    log(f"  [{i}] text='{txt}' | href='{href}'")
                except Exception:
                    pass
            tables = page.locator("table:visible").all()
            log(f"[诊断] 共 {len(tables)} 个可见表格")
        except Exception:
            pass

    if request_el is None:
        log("  ⚠ 未找到请求ID链接")
        return ("", None)

    rid_text = request_el.inner_text().strip()
    current_pages = list(context.pages)
    current_url = page.url

    # 少数请求 ID 只响应浏览器认可的真实点击；JS evaluate click 可能完全不触发
    # 导航。这里仍只发送一次 Playwright click，并在点击前订阅新页面事件，避免
    # 旧版“失败后再次点击”导致同一请求打开多个详情页的问题。
    try:
        log("  正在点击请求ID（单次真实点击，原子等待新标签页）...")
        with context.expect_page(timeout=12000) as page_info:
            request_el.click()
        detail_page = page_info.value
        detail_page.bring_to_front()
        try:
            detail_page.wait_for_load_state("domcontentloaded", timeout=15000)
        except Exception:
            pass
        log("  ✓ 请求ID已打开详情页")
        return (rid_text, detail_page)
    except Exception as e:
        # 不再对同一链接发出第二次点击。若页面刚好在超时边界才创建，下面的
        # 短暂检查仍可接住它；否则本 PO 安全失败，不留下重复详情页。
        log(f"  ⚠ 请求ID未在规定时间打开新标签页: {str(e)[:100]}")

    log(
        f"  正在等待详情页登记（最多 {REQUEST_ID_LATE_PAGE_WAIT_SECONDS} 秒，不重复点击请求ID）..."
    )
    for elapsed_seconds in range(REQUEST_ID_LATE_PAGE_WAIT_SECONDS):
        # 不只看点击后新建的标签页：站点可能把详情加载进点击前已存在的标签。
        detail_page = _find_detail_page_for_request(context, rid_text)
        if detail_page is not None:
            detail_page.bring_to_front()
            try:
                detail_page.wait_for_load_state("domcontentloaded", timeout=15000)
            except Exception:
                pass
            source = "已有标签页" if detail_page in current_pages else "新标签页"
            log(f"  ✓ 请求ID在 {elapsed_seconds} 秒后通过{source}识别到详情页")
            return (rid_text, detail_page)

        # 实测中详情页已经在 Edge 窗口内可见，但迟迟没有出现在 context.pages。
        # 此时通过浏览器级 CDP 目标列表拿到精确 URL，再在当前自动化上下文中打开
        # 同一页，后续下载仍由 Playwright 正常接管。每两秒检查一次即可，无需盲等。
        if elapsed_seconds % 2 == 0:
            browser_target = _find_browser_target_for_request(context, rid_text)
            if browser_target is not None:
                target_url = browser_target.get("url")
                controlled_page = None
                try:
                    log("  ⓘ 详情页已在 Edge 可见但未登记，正在接管该请求页面...")
                    controlled_page = context.new_page()
                    controlled_page.goto(
                        target_url, wait_until="domcontentloaded", timeout=30000
                    )
                    if not _is_detail_url_for_request(controlled_page.url, rid_text):
                        raise RuntimeError("受控页面未进入当前请求的详情地址")
                    controlled_page.bring_to_front()
                    _close_browser_target(context, browser_target.get("targetId"))
                    log(f"  ✓ 请求ID在 {elapsed_seconds} 秒后已由受控页面接管")
                    return (rid_text, controlled_page)
                except Exception as e:
                    if controlled_page is not None:
                        try:
                            controlled_page.close()
                        except Exception:
                            pass
                    log(f"  ⚠ 详情页接管失败，继续等待原页面登记: {str(e)[:100]}")

        new_pages = [p for p in context.pages if p not in current_pages]
        if new_pages:
            detail_page = new_pages[0]
            detail_page.bring_to_front()
            try:
                detail_page.wait_for_load_state("domcontentloaded", timeout=15000)
            except Exception:
                pass
            log(f"  ✓ 请求ID延迟 {elapsed_seconds} 秒打开详情页")
            return (rid_text, detail_page)
        if page.url != current_url and "search" not in page.url.lower():
            try:
                page.wait_for_load_state("domcontentloaded", timeout=15000)
            except Exception:
                pass
            log(f"  ✓ 请求ID在 {elapsed_seconds} 秒后已在当前标签页打开详情")
            return (rid_text, page)
        if elapsed_seconds and elapsed_seconds % 10 == 0:
            log(f"  仍在等待迟到详情页（已等待 {elapsed_seconds} 秒）...")
        time.sleep(1)

    log("  ✗ 请求ID未打开详情页：本 PO 记为失败，不写入待回填任务")
    try:
        for index, candidate in enumerate(context.pages):
            if not candidate.is_closed():
                log(f"  [诊断] 仍打开标签页[{index}]: {str(candidate.url)[:180]}")
    except Exception:
        pass
    return (rid_text, None)


# ═══════════════════ 详情页数据提取 ═══════════════════

def _extract_project_name(page, log):
    """从详情页文本中提取「项目名称」字段值。"""
    try:
        body_text = page.inner_text("body")
    except Exception:
        return None
    if not body_text:
        return None

    lines = [l.strip() for l in body_text.split("\n") if l.strip()]

    # 搜索 "项目名称" 标签，取紧随其后的中文值
    for i, line in enumerate(lines):
        if "项目名称" not in line:
            continue
        # 尝试在当前行取「项目名称」之后的文本
        m = re.search(r'项目名称\s*[：:]*\s*([\u4e00-\u9fff][\u4e00-\u9fff\w（）()\-·、，\s]+)', line)
        if m:
            val = m.group(1).strip()
            if len(val) > 2:
                log(f"  项目名称（同行）: {val[:60]}")
                return val
        # 尝试下一行
        if i + 1 < len(lines):
            next_line = lines[i + 1].strip()
            if next_line and not next_line.startswith("项目") and any('\u4e00' <= c <= '\u9fff' for c in next_line):
                log(f"  项目名称（下行）: {next_line[:60]}")
                return next_line

    # 兜底：正则搜索
    m = re.search(r'项目名称\s*[：:]*\s*([\u4e00-\u9fff][\u4e00-\u9fff\w（）()\-·、，]+)', body_text)
    if m:
        val = m.group(1).strip()
        if len(val) > 2:
            log(f"  项目名称（正则）: {val[:60]}")
            return val

    log("  ⚠ 未提取到项目名称")
    return None


def _extract_wbs(page, log):
    """从详情页头部提取「项目WBS No」字段值。"""
    try:
        body_text = page.inner_text("body")
    except Exception:
        return None
    if not body_text:
        return None

    pattern = re.compile(
        r"项目\s*WBS\s*(?:No\.?)?\s*[：:]?\s*([A-Za-z0-9][A-Za-z0-9._/-]*)",
        re.IGNORECASE,
    )
    for line in (line.strip() for line in body_text.split("\n")):
        if "WBS" not in line.upper():
            continue
        match = pattern.search(line)
        if match:
            wbs = match.group(1).strip()
            log(f"  项目WBS No: {wbs}")
            return wbs

    # 少数页面会把标签和值渲染为相邻两行。
    lines = [line.strip() for line in body_text.split("\n") if line.strip()]
    for index, line in enumerate(lines[:-1]):
        if "WBS" not in line.upper() or "项目" not in line:
            continue
        candidate = lines[index + 1]
        if re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._/-]*", candidate):
            log(f"  项目WBS No（下行）: {candidate}")
            return candidate

    log("  ⚠ 未提取到项目WBS No")
    return None


def _extract_order_qty(page, log):
    """
    从详情页「梯台明细」表中提取合计行的数量值。
    策略1：找表格→定位「数量」列→定位「合计」行→取交叉单元格。
    策略2（兜底）：找不到「合计」行时，取表格末行的数量列值。
    """
    try:
        # 滚动到梯台明细区域
        try:
            target = page.locator("//*[contains(text(),'梯台明细')]").first
            target.scroll_into_view_if_needed()
            time.sleep(0.5)
        except Exception:
            pass

        for table in page.locator("table").all():
            try:
                if not table.is_visible():
                    continue
                tbl_text = table.inner_text()
                if "数量" not in tbl_text:
                    continue

                # 找「数量」列索引
                qty_col = None
                ths = table.locator("th").all()
                for i, th in enumerate(ths):
                    if (th.inner_text() or "").strip() == "数量":
                        qty_col = i
                        break
                if qty_col is None:
                    first_row = table.locator("tr").first
                    tds = first_row.locator("td").all()
                    for i, td in enumerate(tds):
                        if (td.inner_text() or "").strip() == "数量":
                            qty_col = i
                            break
                if qty_col is None:
                    continue

                # 策略1：找「合计」行 → 取数量列单元格
                for row in table.locator("tr").all():
                    row_text = row.inner_text() or ""
                    if "合计" not in row_text:
                        continue
                    cells = row.locator("td").all()
                    if len(cells) > qty_col:
                        val = cells[qty_col].inner_text().strip()
                        if val:
                            log(f"  梯台明细合计数量: {val}")
                            return val

                # 策略2：兜底 — 末行取值（合计行可能不以"合计"文字出现）
                all_rows = table.locator("tr").all()
                for row in reversed(all_rows):
                    row_text = (row.inner_text() or "").strip()
                    if not row_text:
                        continue
                    cells = row.locator("td, th").all()
                    if len(cells) > qty_col:
                        val = cells[qty_col].inner_text().strip()
                        if val and re.match(r'^[\d,.]+$', val):
                            log(f"  梯台明细合计数量（末行兜底）: {val}")
                            return val
                    break  # 只尝试最后一个非空行
            except Exception:
                continue
    except Exception:
        pass

    log("  ⚠ 未提取到梯台明细合计数量")
    return None


def _extract_target_msg_name(text):
    """Return a FLD .msg filename from link text or URL-like text."""
    if not text:
        return ""
    value = str(text).strip()
    try:
        from urllib.parse import unquote
        value = unquote(value)
    except Exception:
        pass

    # Only filenames starting with FLD are valid approval mail files.
    pattern = r"(?i)(?:^|[\\/\s=&?])(fld[^\\/\r\n<>]*?\.msg)"
    match = re.search(pattern, value)
    if not match:
        return ""
    name = match.group(1).strip().strip("'\"")
    return os.path.basename(name)


def _is_target_msg_filename(name):
    """True only for .msg files whose basename starts with FLD."""
    extracted = _extract_target_msg_name(name)
    if not extracted:
        return False
    lower = extracted.lower()
    return lower.endswith(".msg") and lower.startswith(TARGET_MSG_PREFIXES)


def _extract_any_file_name(text):
    """Extract a filename (with extension) from link text or URL-like text (any file type)."""
    if not text:
        return ""
    value = str(text).strip()
    try:
        from urllib.parse import unquote
        value = unquote(value)
    except Exception:
        pass
    # Match filenames with common extensions
    pattern = r"(?:^|[\\/\s=&?])([^\\/\r\n<>]*?\.[a-zA-Z0-9]{2,5})(?:\s|$|[\\/?&])"
    match = re.search(pattern, value)
    if not match:
        return ""
    name = match.group(1).strip().strip("'\"")
    return os.path.basename(name)


def _link_file_name(link):
    """Extract filename from a link element (any file type, not just FLD .msg)."""
    candidates = []
    try:
        candidates.append(link.inner_text().strip())
    except Exception:
        pass
    for attr in ("download", "title", "aria-label", "href"):
        try:
            value = link.get_attribute(attr)
            if value:
                candidates.append(value)
        except Exception:
            pass
    for candidate in candidates:
        name = _extract_any_file_name(candidate)
        if name:
            return name
    return ""


def _classify_file(file_name):
    """Classify a file as 'fld' or 'non_fld'.

    FLD 判定规则（满足任一即可）：
      1. 文件名以 fld 开头（如 FLD新物料价格审批--xxx.msg）
      2. 文件名包含 fld 且是 .msg/.eml 邮件附件（如 转发 FLD新物料价格审批--xxx.msg）
    仅规则2不适用于非邮件文件，避免误判（如 xxx_fld_report.pdf）。
    """
    lower = (file_name or "").lower()
    if lower.startswith("fld"):
        return "fld"
    if "fld" in lower and (lower.endswith(".msg") or lower.endswith(".eml")):
        return "fld"
    return "non_fld"


def classify_downloaded_po_attachments(
    po_number, category_label=None, download_dir=None
):
    """根据已下载的实际附件判定一个 PO 是否有 FLD。

    同时扫描 ``FLD文件/<PO>`` 与 ``无FLD文件/<PO>``，兼容旧版本把同一
    PO 留在两个目录中的情况。目录名不参与判定；任意实际文件符合
    ``_classify_file`` 的 FLD 规则即为 ``fld``。目录存在但没有 FLD
    （包括空目录）为 ``non_fld``；两个目录均不存在为 ``missing``；扫描
    过程中发生错误且没有发现 FLD 时为 ``error``。
    """
    po = str(po_number or "").strip()
    root = str(download_dir or DOWNLOAD_DIR)
    category_dir = os.path.join(root, _category_label(category_label))
    po_dirs = [
        os.path.join(category_dir, bucket, po)
        for bucket in ("FLD文件", "无FLD文件")
    ]
    existing_dirs = [path for path in po_dirs if os.path.isdir(path)]
    if not existing_dirs:
        return {
            "status": "missing",
            "po": po,
            "files": [],
            "fld_files": [],
            "folders": [],
            "errors": [],
        }

    files = []
    fld_files = []
    errors = []

    def _record_walk_error(exc):
        errors.append(str(exc))

    for po_dir in existing_dirs:
        try:
            for current_dir, _, file_names in os.walk(
                po_dir, onerror=_record_walk_error
            ):
                for file_name in file_names:
                    full_path = os.path.join(current_dir, file_name)
                    files.append(full_path)
                    if _classify_file(file_name) == "fld":
                        fld_files.append(full_path)
        except OSError as exc:
            errors.append(str(exc))

    if fld_files:
        status = "fld"
    elif errors:
        status = "error"
    else:
        status = "non_fld"

    return {
        "status": status,
        "po": po,
        "files": files,
        "fld_files": fld_files,
        "folders": existing_dirs,
        "errors": errors,
    }


def _extract_project_from_fld_name(filename, log):
    """
    兜底：从 FLD 文件名中提取项目名称。
    文件名格式: FLD新物料价格审批--<项目名>--<分类>--价格审批回复.msg
    策略：-- 分割后，取中间片段中「包含中文且最长」的作为项目名。
    """
    name = os.path.basename(str(filename or ""))
    parts = name.split("--")
    if len(parts) < 3:
        return None
    best = None
    for part in parts[1:-1]:  # 跳过首段(FLD前缀)和末段(价格审批回复.msg)
        part = part.strip()
        if any('\u4e00' <= c <= '\u9fff' for c in part):
            if best is None or len(part) > len(best):
                best = part
    if best:
        log(f"  项目名称（FLD文件名兜底）: {best[:60]}")
    return best


def _target_msg_name_from_link(link):
    candidates = []
    try:
        candidates.append(link.inner_text().strip())
    except Exception:
        pass
    for attr in ("download", "title", "aria-label", "href"):
        try:
            value = link.get_attribute(attr)
            if value:
                candidates.append(value)
        except Exception:
            pass

    for candidate in candidates:
        name = _extract_target_msg_name(candidate)
        if name:
            return name
    return ""


# 附件文件扩展名白名单（用于过滤非附件链接如 view.tkelevator.com.cn）
_ATTACHMENT_EXTENSIONS = {
    '.msg', '.eml', '.pdf', '.xlsx', '.xls', '.xlsm',
    '.doc', '.docx', '.zip', '.rar', '.7z',
    '.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff',
}

def _looks_like_attachment(href, name):
    """判断链接是否像附件（非页面导航链接）。"""
    # 下载链接模式
    if href and ('/download' in href.lower() or 'download' in href.lower()):
        return True
    # 文件名包含附件扩展名
    name_lower = name.lower()
    for ext in _ATTACHMENT_EXTENSIONS:
        if name_lower.endswith(ext):
            return True
    # href 包含附件扩展名
    if href:
        href_lower = href.lower()
        for ext in _ATTACHMENT_EXTENSIONS:
            if ext in href_lower:
                return True
    return False


def _collect_all_attachment_links(page):
    """Collect ALL attachment links from the detail page (not just FLD .msg).
    同时扫描「支持文件」区域和「项目协议」区域的附件链接。"""
    selectors = [
        "//th[contains(text(),'\u652f\u6301\u6587\u4ef6')]/ancestor::table//a",
        "//td[contains(text(),'\u652f\u6301\u6587\u4ef6')]/ancestor::table//a",
        "//th[contains(text(),'\u9879\u76ee\u534f\u8bae')]/ancestor::table//a",
        "//td[contains(text(),'\u9879\u76ee\u534f\u8bae')]/ancestor::table//a",
        "a",
    ]
    found = []
    seen = set()
    for selector in selectors:
        try:
            links = page.locator(selector).all()
        except Exception:
            continue
        for link in links:
            try:
                if not link.is_visible():
                    continue
            except Exception:
                pass
            name = _link_file_name(link)
            if not name:
                continue
            try:
                href = link.get_attribute("href") or ""
            except Exception:
                href = ""
            # 过滤非附件链接（如 view.tkelevator.com.cn）
            if not _looks_like_attachment(href, name):
                continue
            key = (name.lower(), href)
            if key in seen:
                continue
            seen.add(key)
            found.append((link, name))
    return found


def _download_support_files(page, log, po):
    """
    下载详情页中的全部附件，存入同一目录。
    有 FLD 附件 → downloads/FLD文件/<PO号>/（全部文件，含非FLD）
    无 FLD 附件 → downloads/无FLD文件/<PO号>/
    返回: {fld_files: [...], non_fld_files: [...]}
    """
    fld_target = os.path.join(FLD_DIR, str(po))
    non_fld_target = os.path.join(NO_FLD_DIR, str(po))

    # 等待「支持文件」或「项目协议」区域出现后再滚动
    try:
        page.wait_for_selector(
            "//th[contains(text(),'支持文件')] | //td[contains(text(),'支持文件')] | //*[contains(text(),'项目协议')]",
            timeout=8000
        )
    except Exception:
        pass
    try:
        target = page.locator("//*[contains(text(),'审批记录')]").first
        target.scroll_into_view_if_needed()
    except Exception:
        pass
    # 同时滚动到「项目协议」区域（附件可能在两个区域中任一出现）
    try:
        target = page.locator("//*[contains(text(),'项目协议')]").first
        target.scroll_into_view_if_needed()
    except Exception:
        pass
    # 滚动到页面底部再回到顶部，触发所有懒加载区域渲染
    try:
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(1)
        page.evaluate("window.scrollTo(0, 0)")
        time.sleep(0.5)
    except Exception:
        pass

    fld_files = []
    non_fld_files = []
    all_links = _collect_all_attachment_links(page)
    # 首扫为空 → 主动等待文件链接出现（有则立刻返回，无则12s超时）
    if not all_links:
        log("  附件链接尚未出现，等待中...")
        try:
            page.wait_for_selector(
                "//th[contains(text(),'支持文件')]/ancestor::table//a[contains(text(),'.')]",
                timeout=12000
            )
            time.sleep(0.5)
            all_links = _collect_all_attachment_links(page)
        except Exception:
            pass
    if not all_links:
        log(f"  未发现任何附件链接（当前URL: {page.url}）。")
        # ── 诊断：保存页面文本 + 列出所有可见链接 ──
        try:
            body_text = page.inner_text("body")
            if body_text:
                with open(PAGE_SCAN_FILE, "w", encoding="utf-8") as f:
                    f.write(body_text)
                log(f"  [诊断] 页面文本已保存到 {PAGE_SCAN_FILE}")
        except Exception:
            pass
        try:
            all_visible_links = page.locator("a:visible").all()
            log(f"  [诊断] 页面共 {len(all_visible_links)} 个可见链接：")
            for idx, alink in enumerate(all_visible_links[:30]):
                try:
                    txt = (alink.inner_text() or "").strip()[:80]
                    href = (alink.get_attribute("href") or "")[:100]
                    log(f"    [{idx}] '{txt}' -> {href}")
                except Exception:
                    pass
            # 检查关键标记是否存在
            for marker in ["支持文件", "审批记录", "项目协议", "项目经理"]:
                try:
                    found = page.locator(f"//*[contains(text(),'{marker}')]").count()
                    log(f"  [诊断] '{marker}' 出现次数: {found}")
                except Exception:
                    pass
        except Exception:
            pass
        # 没有附件也归入「无FLD文件」：保留 PO 目录作为“已查询但无附件”的记录。
        os.makedirs(non_fld_target, exist_ok=True)
        return {"fld_files": fld_files, "non_fld_files": non_fld_files}

    # 有任一 FLD 文件 → 全部存入 FLD文件夹；否则存入 无FLD文件夹
    has_any_fld = any(_classify_file(name) == "fld" for _, name in all_links)
    if has_any_fld:
        target_dir = fld_target
        dir_label = "FLD文件"
    else:
        target_dir = non_fld_target
        dir_label = "无FLD文件"
    os.makedirs(target_dir, exist_ok=True)

    for link, file_name in all_links:
        try:
            category = _classify_file(file_name)
            log(f"  正在下载: {file_name} [{category}]")
            with page.expect_download(timeout=30000) as download_info:
                link.click()
            download = download_info.value

            safe_name = file_name.replace("/", "_").replace("\\", "_").replace(":", "_")
            suggested = download.suggested_filename or ""
            if suggested:
                suggested = os.path.basename(suggested)
                safe_name = suggested.replace("/", "_").replace("\\", "_").replace(":", "_")

            save_path = os.path.join(target_dir, safe_name)
            download.save_as(save_path)
            if category == "fld":
                fld_files.append(safe_name)
            else:
                non_fld_files.append(safe_name)
            log(f"    ✓ 已保存: {dir_label}/{po}/{safe_name}")
            time.sleep(0.5)
        except Exception as e:
            log(f"    ✗ 下载失败: {str(e)[:100]}")

    log(f"  附件下载完成: FLD {len(fld_files)} 个, 非FLD {len(non_fld_files)} 个 → {dir_label}")
    return {"fld_files": fld_files, "non_fld_files": non_fld_files}


def _has_fld_msg_attachment(page, log):
    """
    检查详情页附件中是否存在文件名以 FLD 开头的 .msg 文件。
    多策略扫描，返回 True/False。
    """
    # 先滚动到审批记录区域（懒加载）
    try:
        target = page.locator("//*[contains(text(),'审批记录')]").first
        target.scroll_into_view_if_needed()
        time.sleep(1)
    except Exception:
        pass
    # 同时滚动到「项目协议」区域（附件可能在此区域）
    try:
        target = page.locator("//*[contains(text(),'项目协议')]").first
        target.scroll_into_view_if_needed()
        time.sleep(0.5)
    except Exception:
        pass

    all_links = _collect_all_attachment_links(page)
    for _, name in all_links:
        if _classify_file(name) == "fld":
            log(f"  ✓ 发现 FLD 附件: {name[:80]}")
            return True

    log("  未发现 FLD 开头的附件")
    return False


def _has_any_msg_attachment(page):
    """快速扫描是否存在任何 .msg 附件（不管是不是 FLD）"""
    try:
        links = page.locator("a").all()
        for link in links[:50]:
            try:
                if ".msg" in link.inner_text().strip().lower():
                    return True
            except Exception:
                continue
    except Exception:
        pass
    return False


# ═══════════════════ FLD .msg 审批价格提取 ═══════════════════

def _compact_text(value):
    return re.sub(r"\s+", "", str(value or ""))


def _extract_amount_from_text(value):
    text = str(value or "").replace("\xa0", " ")
    amount_pattern = r"(?<![\d])(?:\d{1,3}(?:,\d{3})+|\d{4,}|\d{1,3})(?:\.\d+)?(?![\d])"
    for match in re.finditer(amount_pattern, text):
        raw = match.group(0)
        cleaned = raw.replace(",", "")
        try:
            if float(cleaned) > 100:
                return cleaned
        except ValueError:
            continue
    return None


def _extract_approval_price_from_tables(soup, log):
    for table_no, table in enumerate(soup.find_all("table"), 1):
        rows = []
        for tr in table.find_all("tr"):
            cells = [
                cell.get_text(" ", strip=True)
                for cell in tr.find_all(["td", "th"])
            ]
            if cells:
                rows.append(cells)

        for row_idx, row in enumerate(rows):
            price_cols = [
                col_idx
                for col_idx, cell_text in enumerate(row)
                if APPROVAL_PRICE_LABEL in _compact_text(cell_text)
            ]
            if not price_cols:
                continue

            for col_idx in price_cols:
                # Vertical/key-value table: label and value can be in the same row.
                for candidate in row[col_idx + 1: col_idx + 3]:
                    amount = _extract_amount_from_text(candidate)
                    if amount:
                        log(f"  [MSG] 表格{table_no} 同行提取审批价格: {amount}")
                        return amount

                # Horizontal table: label is a header, value is in the next data row.
                for next_row in rows[row_idx + 1:]:
                    if col_idx >= len(next_row):
                        continue
                    amount = _extract_amount_from_text(next_row[col_idx])
                    if amount:
                        log(f"  [MSG] 表格{table_no} 第{col_idx + 1}列审批价格: {amount}")
                        return amount
    return None


def _extract_approval_price_from_plain_text(text, log):
    lines = [_compact_text(line) for line in str(text or "").splitlines()]
    lines = [line for line in lines if line]
    for idx, line in enumerate(lines):
        if APPROVAL_PRICE_LABEL not in line:
            continue

        # Only use this fallback when the label is isolated. If a whole header row
        # is flattened into one line, the next line may start with audit IDs.
        if len(line) > len(APPROVAL_PRICE_LABEL) + 4:
            continue

        amount = _extract_amount_from_text(line.replace(APPROVAL_PRICE_LABEL, "", 1))
        if amount:
            log(f"  [MSG] 纯文本同行提取审批价格: {amount}")
            return amount

        for next_line in lines[idx + 1: idx + 5]:
            amount = _extract_amount_from_text(next_line)
            if amount:
                log(f"  [MSG] 纯文本下一行提取审批价格: {amount}")
                return amount
    return None


def _extract_approval_price_from_msg(msg_path, log):
    """
    解析 FLD .msg 邮件，从正文表格中提取「审批价格」。
    返回金额字符串（如 "246900"），失败返回 None。
    """
    import extract_msg
    from bs4 import BeautifulSoup

    log(f"  [MSG] 解析: {os.path.basename(msg_path)}")
    try:
        msg = extract_msg.Message(msg_path)
        html = msg.htmlBody or ""
        body_text = msg.body or ""
        msg.close()

        if not html.strip() and not body_text.strip():
            log("  [MSG] 邮件正文为空（HTML和纯文本均无内容）")
            return None

        # 优先用 HTML 解析，回退到纯文本
        if html.strip():
            source = html
            log(f"  [MSG] 正文来源: HTML, 长度: {len(source)}")
        else:
            source = body_text
            log(f"  [MSG] 正文来源: 纯文本(无HTML), 长度: {len(source)}")

        soup = BeautifulSoup(source, "html.parser")
        
        # 诊断：输出解析后的纯文本前 200 字符
        plain_preview = soup.get_text()[:200]
        log(f"  [MSG] 文本预览: {plain_preview}")
        
        amount = _extract_approval_price_from_tables(soup, log)
        if amount:
            return amount

        amount = _extract_approval_price_from_plain_text(soup.get_text("\n"), log)
        if amount:
            return amount

        log("  [MSG] 未找到「审批价格」")
        return None
    except ImportError:
        log("  [MSG] extract-msg 模块未安装，请运行: pip install extract-msg beautifulsoup4")
        return None
    except Exception as e:
        log(f"  [MSG] 解析失败: {e}")
        return None


def _extract_approval_price_from_eml(eml_path, log):
    """
    解析 FLD .eml 邮件，从正文表格中提取「审批价格」。
    返回金额字符串（如 "366200"），失败返回 None。
    """
    import email
    from email import policy
    from bs4 import BeautifulSoup

    log(f"  [EML] 解析: {os.path.basename(eml_path)}")
    try:
        with open(eml_path, 'r', encoding='utf-8', errors='replace') as f:
            msg = email.message_from_file(f, policy=policy.default)

        html_body = ""
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == 'text/html':
                    payload = part.get_payload(decode=True)
                    if payload:
                        html_body = payload.decode('utf-8', errors='replace')
                    break
        else:
            payload = msg.get_payload(decode=True)
            if payload:
                html_body = payload.decode('utf-8', errors='replace')

        if not html_body.strip():
            log("  [EML] 邮件正文为空（无HTML内容）")
            return None

        log(f"  [EML] 正文来源: HTML, 长度: {len(html_body)}")
        soup = BeautifulSoup(html_body, "html.parser")

        plain_preview = soup.get_text()[:200]
        log(f"  [EML] 文本预览: {plain_preview}")

        amount = _extract_approval_price_from_tables(soup, log)
        if amount:
            return amount

        amount = _extract_approval_price_from_plain_text(soup.get_text("\n"), log)
        if amount:
            return amount

        log("  [EML] 未解析到审批价格")
        return None
    except Exception as e:
        log(f"  [EML] 解析异常: {e}")
        return None


# ═══════════════════ PDF 金额提取 ═══════════════════

def _extract_total_from_pdf(pdf_path, log):
    """
    读取 PDF 文件，提取「总计（不含税）」后的金额。
    使用子进程执行 pdfplumber 解析，防止卡死主流程。
    """
    log(f"  [PDF] 解析: {os.path.basename(pdf_path)}")
    worker_code = r'''
import json, sys, pdfplumber, re

pdf_path = sys.argv[1]
logs = []

def emit(status, total=None, error=""):
    print(json.dumps({
        "status": status, "total": total, "error": error, "logs": logs,
    }, ensure_ascii=False))

try:
    with pdfplumber.open(pdf_path) as pdf:
        for page_no, page in enumerate(pdf.pages, 1):
            logs.append(f"  [PDF] 第 {page_no}/{len(pdf.pages)} 页")
            # 归一化旋转（摆正颠倒/横排页面）
            rot = getattr(page, 'rotation', 0) or 0
            if rot != 0:
                logs.append(f"  [PDF] 检测到旋转 {rot}°，正在摆正...")
                try:
                    page = page.rotate(360 - rot)
                except Exception:
                    pass
            text = page.extract_text() or ""
            # 补充表格文字（总计可能在表格中）
            try:
                tables = page.extract_tables()
                for tbl in tables:
                    for row in tbl:
                        for cell in row:
                            if cell and ("总计" in str(cell) or "小计" in str(cell) or "合计" in str(cell)):
                                text += "\n" + str(cell)
            except Exception:
                pass
            if not text.strip():
                continue
            lines = text.split("\n")
            for i, line in enumerate(lines):
                # 匹配关键词：精确匹配优先，"总计"兜底
                if "总计（不含税）" not in line and "总计(不含税)" not in line and "总计" not in line:
                    continue
                logs.append(f"  [PDF] 匹配行: {line[:150]}")
                # ── 正则提取数字（解决千分位空格/逗号切分BUG）──
                # 匹配数字模式：支持 "9 ,594.00"、"9,594.00"、"6520.00" 等
                candidates = []
                for raw in re.findall(r'[\d][\d, ]*\.?\d*', line):
                    cleaned = raw.replace(",", "").replace(" ", "").replace("\uffe5", "").replace("\xa5", "")
                    try:
                        v = float(cleaned)
                        if v > 100:
                            candidates.append((v, cleaned))
                    except ValueError:
                        continue
                # 也检查下一行
                if i + 1 < len(lines):
                    for raw in re.findall(r'[\d][\d, ]*\.?\d*', lines[i + 1]):
                        cleaned = raw.replace(",", "").replace(" ", "").replace("\uffe5", "").replace("\xa5", "")
                        try:
                            v = float(cleaned)
                            if v > 100:
                                candidates.append((v, cleaned))
                        except ValueError:
                            continue
                if candidates:
                    # 取最大值（排除可能的小额杂数）
                    best_val, best_str = max(candidates, key=lambda x: x[0])
                    logs.append(f"  [PDF] 提取总计: {best_str}")
                    emit("ok", best_str)
                    sys.exit(0)
    emit("ok")
except Exception as e:
    emit("error", error=str(e))
'''

    try:
        completed = subprocess.run(
            [sys.executable, "-c", worker_code, pdf_path],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
            timeout=PDF_PARSE_TIMEOUT,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
    except subprocess.TimeoutExpired:
        log(f"  [PDF] 解析超时（{PDF_PARSE_TIMEOUT}秒），跳过")
        return None
    except Exception as e:
        log(f"  [PDF] 子进程启动失败: {e}")
        return None

    output = (completed.stdout or "").strip()
    if not output:
        stderr = (completed.stderr or "").strip()
        if stderr:
            log(f"  [PDF] 无结果: {stderr[:200]}")
        return None

    try:
        data = json.loads(output.splitlines()[-1])
    except Exception:
        log(f"  [PDF] 解析JSON失败: {output[:200]}")
        return None

    for line in data.get("logs", []):
        log(line)

    if data.get("status") == "error":
        log(f"  [PDF] 读取失败: {data.get('error')}")
        return None

    return data.get("total")


# ═══════════════════ OCR 回退 ═══════════════════

def _extract_total_from_pdf_ocr(pdf_path, log):
    """
    OCR 回退：当 pdfplumber 无法提取文字时，将 PDF 首页转为图片，
    用 Tesseract 中文识别，搜索「总计（不含税）」后的金额。
    返回金额字符串，失败返回 None。
    """
    log(f"  [OCR] 尝试 OCR: {os.path.basename(pdf_path)}")
    worker_code = r'''
import json, sys, re, os

try:
    import pytesseract
    from pdf2image import convert_from_path
except ImportError as e:
    print(json.dumps({"status": "error", "error": f"OCR模块未安装: {e}", "logs": []}, ensure_ascii=False))
    sys.exit(0)

pdf_path = sys.argv[1]
logs = []

def emit(status, total=None, error=""):
    print(json.dumps({
        "status": status, "total": total, "error": error, "logs": logs,
    }, ensure_ascii=False))

try:
    # 检测 Tesseract 路径
    tesseract_paths = [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        os.path.join(os.environ.get("LOCALAPPDATA", ""), "Programs", "Tesseract-OCR", "tesseract.exe"),
    ]
    tesseract_found = None
    for tp in tesseract_paths:
        if os.path.exists(tp):
            tesseract_found = tp
            pytesseract.pytesseract.tesseract_cmd = tp
            break
    if not tesseract_found:
        # 尝试 PATH 中的 tesseract
        import shutil
        found = shutil.which("tesseract")
        if found:
            pytesseract.pytesseract.tesseract_cmd = found
            tesseract_found = found
    if not tesseract_found:
        emit("error", error="Tesseract OCR 未安装。请从 https://github.com/UB-Mannheim/tesseract/wiki 下载安装，并勾选中文语言包。")
        sys.exit(0)

    logs.append(f"  [OCR] Tesseract: {tesseract_found}")

    # 只转换首页（供应商报价单通常是单页）
    images = convert_from_path(pdf_path, first_page=1, last_page=1, dpi=300)
    if not images:
        emit("error", error="PDF转图片失败")
        sys.exit(0)

    img = images[0]
    logs.append(f"  [OCR] 图片尺寸: {img.size}")

    # 中文 OCR
    text = pytesseract.image_to_string(img, lang="chi_sim+eng", config="--psm 6")
    if not text or not text.strip():
        emit("ok")
        sys.exit(0)

    logs.append(f"  [OCR] 识别文字 {len(text)} 字符")

    # 搜索「总计（不含税）」及其变体
    lines = text.split("\n")
    candidates = []
    for i, line in enumerate(lines):
        # 多种关键词匹配
        if not ("总计" in line or "合计" in line or "总价" in line or "总金额" in line):
            continue
        logs.append(f"  [OCR] 候选行: {line[:150]}")
        # 正则提取本行及下一行的数字
        for check_line in (line, lines[i + 1] if i + 1 < len(lines) else ""):
            for raw in re.findall(r'[\d][\d, ]*\.?\d*', check_line):
                cleaned = raw.replace(",", "").replace(" ", "").replace("\uffe5", "").replace("\xa5", "")
                try:
                    v = float(cleaned)
                    if v > 100:
                        candidates.append((v, cleaned))
                except ValueError:
                    continue

    if candidates:
        best_val, best_str = max(candidates, key=lambda x: x[0])
        logs.append(f"  [OCR] 提取总计: {best_str}")
        emit("ok", best_str)
    else:
        logs.append("  [OCR] 未找到有效金额")
        emit("ok")
except Exception as e:
    emit("error", error=str(e))
'''

    try:
        completed = subprocess.run(
            [sys.executable, "-c", worker_code, pdf_path],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
            timeout=OCR_TIMEOUT,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
    except subprocess.TimeoutExpired:
        log(f"  [OCR] 识别超时（{OCR_TIMEOUT}秒），跳过")
        return None
    except Exception as e:
        log(f"  [OCR] 子进程启动失败: {e}")
        return None

    output = (completed.stdout or "").strip()
    if not output:
        stderr = (completed.stderr or "").strip()
        if stderr:
            log(f"  [OCR] 无结果: {stderr[:200]}")
        return None

    try:
        data = json.loads(output.splitlines()[-1])
    except Exception:
        log(f"  [OCR] 解析JSON失败: {output[:200]}")
        return None

    for line in data.get("logs", []):
        log(line)

    if data.get("status") == "error":
        log(f"  [OCR] 识别失败: {data.get('error')}")
        return None

    return data.get("total")


# ═══════════════════ Excel 读写 ═══════════════════

def _read_order_value_from_pivot(po, target_sheet=None):
    """从透视表读取指定 PO 的订单净值（列B「求和项:订单净值」）。"""
    import openpyxl
    sheet = target_sheet or TARGET_SHEET
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return None
        ws = wb[sheet]
        po_str = str(po).strip()
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row, 1).value or "").strip() == po_str:
                val = ws.cell(row, 2).value
                wb.close()
                return str(val) if val is not None else None
        wb.close()
    except Exception:
        pass
    return None


def _write_to_pivot_excel(po, project_name=None, order_qty=None,
                          support_files=None, total_amount=None, order_diff=None,
                          wbs=None, target_sheet=None):
    """
    将提取数据写回透视表 Sheet，按采购凭证行对齐。
    """
    import openpyxl

    sheet = target_sheet or TARGET_SHEET

    if not os.path.exists(EXCEL_FILE):
        return f"Excel 文件不存在: {EXCEL_FILE}"

    wb = openpyxl.load_workbook(EXCEL_FILE)
    if sheet not in wb.sheetnames:
        wb.close()
        return f"Sheet「{sheet}」不存在"

    ws = wb[sheet]

    # 1. 扫描第 1 行定位扩展列
    ext_cols = {}
    for col in range(1, ws.max_column + 1):
        val = str(ws.cell(1, col).value or "").strip()
        if val == "E2E项目名":
            ext_cols["E2E项目名"] = col
        elif val == "WBS":
            ext_cols["WBS"] = col
        elif val == "E2E订单数量":
            ext_cols["E2E订单数量"] = col
        elif val == "项目总金额(审批)":
            ext_cols["项目总金额(审批)"] = col
        elif val == "订单差异(单独计算)":
            ext_cols["订单差异(单独计算)"] = col
        elif val == "支持文件名(FLD 审批)":
            ext_cols["支持文件名(FLD 审批)"] = col

    if not ext_cols:
        wb.close()
        return "未找到扩展列，请先点击「添加扩展列」按钮"

    # 2. 扫描列A 找匹配 PO 的行
    po_str = str(po).strip()
    target_rows = []
    for row in range(2, ws.max_row + 1):
        val = str(ws.cell(row, 1).value or "").strip()
        if val == po_str:
            target_rows.append(row)

    if not target_rows:
        wb.close()
        return f"在列A中未找到采购凭证 {po_str}"

    # 3. 写入数据
    filled = []
    for tr in target_rows:
        if wbs and "WBS" in ext_cols:
            ws.cell(tr, ext_cols["WBS"], wbs)
            filled.append("WBS")
        if project_name and "E2E项目名" in ext_cols:
            ws.cell(tr, ext_cols["E2E项目名"], project_name)
            filled.append("E2E项目名")
        if order_qty and "E2E订单数量" in ext_cols:
            try:
                ws.cell(tr, ext_cols["E2E订单数量"], float(str(order_qty).replace(",", "")))
            except ValueError:
                ws.cell(tr, ext_cols["E2E订单数量"], order_qty)
            filled.append("E2E订单数量")
        if total_amount is not None and "项目总金额(审批)" in ext_cols:
            try:
                ws.cell(tr, ext_cols["项目总金额(审批)"], float(str(total_amount).replace(",", "")))
            except ValueError:
                ws.cell(tr, ext_cols["项目总金额(审批)"], total_amount)
            filled.append("项目总金额(审批)")
        if order_diff is not None and "订单差异(单独计算)" in ext_cols:
            try:
                ws.cell(tr, ext_cols["订单差异(单独计算)"], float(str(order_diff).replace(",", "")))
            except ValueError:
                ws.cell(tr, ext_cols["订单差异(单独计算)"], order_diff)
            filled.append("订单差异(单独计算)")
        if support_files and "支持文件名(FLD 审批)" in ext_cols:
            ws.cell(tr, ext_cols["支持文件名(FLD 审批)"], support_files)
            filled.append("支持文件名(FLD 审批)")

    wb.save(EXCEL_FILE)
    wb.close()
    return f"已写入 {len(target_rows)} 行（PO {po_str}）：{', '.join(filled)}"


# ═══════════════════ 单 PO 处理 ═══════════════════

def _drop_page_activity_session(page, expected_session=None):
    """从缓存移除页面 CDP 会话；用于页面/会话已经自行关闭的情形。"""
    with _PAGE_ACTIVITY_SESSIONS_LOCK:
        session = _PAGE_ACTIVITY_SESSIONS.get(page)
        if session is None or (
            expected_session is not None and session is not expected_session
        ):
            return None
        _PAGE_ACTIVITY_SESSIONS.pop(page, None)
        return session


def _release_page_activity_session(page):
    """释放页面关闭后不再需要的 CDP 会话。"""
    session = _drop_page_activity_session(page)
    if session is None:
        return
    try:
        session.detach()
    except Exception:
        pass


def _remember_page_activity_session(page, session):
    """缓存会话，并在页面或会话自行关闭时立即清理缓存。"""
    with _PAGE_ACTIVITY_SESSIONS_LOCK:
        _PAGE_ACTIVITY_SESSIONS[page] = session

    try:
        page_ref = weakref.ref(page)

        def clear_when_page_closed(_closed_page):
            closed_page = page_ref()
            if closed_page is not None:
                _drop_page_activity_session(closed_page, session)

        page.on("close", clear_when_page_closed)
    except Exception:
        # 某些测试替身或旧版对象没有事件接口；弱引用缓存仍会在对象释放时清理。
        pass

    try:
        page_ref = weakref.ref(page)

        def clear_when_session_closed(_closed_session):
            session_page = page_ref()
            if session_page is not None:
                _drop_page_activity_session(session_page, session)

        session.on("close", clear_when_session_closed)
    except Exception:
        pass


def _send_page_activity_commands(session, log):
    """向一个已建立的 CDP 会话发送焦点与生命周期恢复命令。"""
    focus_restored = False
    lifecycle_restored = False
    try:
        # Chromium DevTools Protocol: 模拟已聚焦、活动的页面。
        session.send("Emulation.setFocusEmulationEnabled", {"enabled": True})
        focus_restored = True
    except Exception as e:
        log(f"  ⚠ 页面焦点模拟未生效: {str(e)[:100]}")

    try:
        # 若标签被浏览器冻结，显式切回 active 后再发起导航。
        session.send("Page.setWebLifecycleState", {"state": "active"})
        lifecycle_restored = True
    except Exception as e:
        log(f"  ⚠ 页面 active 生命周期恢复未生效: {str(e)[:100]}")
    return focus_restored, lifecycle_restored


def _restore_page_activity(context, page, log):
    """恢复页面前台焦点与 active 生命周期，避免后台页被 Edge 冻结。

    用户手动点击后流程会立即恢复，说明页面可能在标签切换/关闭后进入了
    后台或 frozen 状态。此处用 Chromium CDP 直接模拟“已聚焦且活动”的页面，
    不依赖鼠标点击，也不会改变页面的业务数据。
    """
    if page is None:
        return

    try:
        if page.is_closed():
            return
    except Exception:
        return

    try:
        page.bring_to_front()
    except Exception as e:
        log(f"  ⚠ 无法切换页面到前台: {str(e)[:100]}")

    with _PAGE_ACTIVITY_SESSIONS_LOCK:
        session = _PAGE_ACTIVITY_SESSIONS.get(page)
    reused_session = session is not None
    if session is None:
        try:
            session = context.new_cdp_session(page)
            _remember_page_activity_session(page, session)
        except Exception as e:
            log(f"  ⚠ 无法创建页面活动状态会话: {str(e)[:100]}")
            return

    focus_restored, lifecycle_restored = _send_page_activity_commands(session, log)
    if not (focus_restored or lifecycle_restored) and reused_session:
        # 会话可能已被目标页或浏览器断开；只在复用的旧会话完全失效时重建一次。
        log("  ⚠ 页面活动状态会话已失效，正在重建...")
        _release_page_activity_session(page)
        try:
            session = context.new_cdp_session(page)
            _remember_page_activity_session(page, session)
            focus_restored, lifecycle_restored = _send_page_activity_commands(session, log)
        except Exception as e:
            log(f"  ⚠ 重建页面活动状态会话失败: {str(e)[:100]}")

    if focus_restored or lifecycle_restored:
        restored_parts = []
        if focus_restored:
            restored_parts.append("焦点")
        if lifecycle_restored:
            restored_parts.append("active 生命周期")
        log(f"  页面活动状态已恢复（{' + '.join(restored_parts)}）")


def _close_other_pages(context, keep_page, log):
    """关闭上下文中除 keep_page 外的页面，防止详情标签页持续堆积。"""
    close_requested_count = 0
    remaining_pages = []
    # 关闭一个页面时，网站脚本可能又打开一个临时页；最多再扫描两轮，
    # 让本轮 PO 的收尾稳定回到唯一查询页，避免无界循环。
    attempted_pages = set()
    for _ in range(3):
        candidates = []
        for candidate in list(context.pages):
            if candidate == keep_page:
                continue
            try:
                if candidate.is_closed():
                    _release_page_activity_session(candidate)
                elif id(candidate) not in attempted_pages:
                    candidates.append(candidate)
            except Exception as e:
                log(f"  ⚠ 无法检查多余页面状态: {str(e)[:100]}")
                if id(candidate) not in attempted_pages:
                    candidates.append(candidate)

        if not candidates:
            remaining_pages = []
            break

        for candidate in candidates:
            attempted_pages.add(id(candidate))
            try:
                # Page.close() 默认会等待标签页彻底关闭；公司电脑中隐藏详情页
                # 会在这一步冻结，直到人工点击当前页面才恢复。run_before_unload=True
                # 会触发关闭请求但不等待，从而不阻塞下一 PO；对 beforeunload 弹窗
                # 显式接受，确保旧页面仍会关闭。
                try:
                    candidate.on("dialog", lambda dialog: dialog.accept())
                except Exception:
                    pass
                _release_page_activity_session(candidate)
                candidate.close(run_before_unload=True)
                close_requested_count += 1
            except Exception as e:
                log(f"  ⚠ 无法关闭多余页面: {str(e)[:100]}")

    for candidate in list(context.pages):
        if candidate == keep_page:
            continue
        try:
            if not candidate.is_closed():
                remaining_pages.append(candidate)
            else:
                _release_page_activity_session(candidate)
        except Exception:
            remaining_pages.append(candidate)
    if close_requested_count:
        log(f"  已发起关闭 {close_requested_count} 个多余详情页（不等待，可继续下一 PO）")
    if remaining_pages:
        log(f"  ⚠ 仍有 {len(remaining_pages)} 个多余页面正在关闭，不阻塞下一 PO")
    return not remaining_pages


def _is_search_page_ready(page, response=None):
    """确认页面是可用的查询页，而不是同路径的错误页或登录中转页。"""
    try:
        if page.is_closed():
            return False
        actual = urlparse(page.url)
        expected = urlparse(SEARCH_URL)
        same_url = (
            actual.scheme == expected.scheme
            and actual.netloc == expected.netloc
            and actual.path.rstrip("/") == expected.path.rstrip("/")
            and actual.query == expected.query
            and actual.fragment == expected.fragment
        )
        if not same_url:
            return False
        # Playwright 的 goto 在 404/500 时不会抛异常，必须检查主资源响应。
        if response is not None and not response.ok:
            return False
        return True
    except Exception:
        return False


def _find_open_search_page(context, preferred_page=None):
    """在上下文中寻找仍可用的标准查询页，优先使用当前主流程页面。"""
    candidates = []
    if preferred_page is not None:
        candidates.append(preferred_page)
    try:
        candidates.extend(page for page in context.pages if page not in candidates)
    except Exception:
        pass
    for candidate in candidates:
        if _is_search_page_ready(candidate):
            return candidate
    return None


def _prepare_next_search_page(context, search_page, detail_page, log):
    """将当前活跃的详情页复用为下一个 PO 的搜索页。

    请求 ID 会在新标签页打开详情。下载完成后若关闭该页并新建搜索页，
    Edge 可能把新页按后台页处理，导致下一次输入必须依赖人工点击唤醒。
    因此关闭已经处于后台的旧搜索页，而让仍活跃的详情页直接导航回搜索页。
    """
    if detail_page is None:
        # 未打开详情页时只清理已验证的查询页之外的标签。若没有有效查询页，
        # 宁可保留现状也不能误关掉可能仍可恢复的最后一个页面。
        verified_search_page = _find_open_search_page(context, search_page)
        if verified_search_page is None:
            log("  ⚠ 未找到可用查询页，保留当前标签页，不执行清理")
            return search_page
        _close_other_pages(context, verified_search_page, log)
        return verified_search_page

    # 详情页刚刚用于下载，仍是浏览器中的活跃标签；保留它可以避免
    # 新建标签后的后台节流/焦点丢失问题。先恢复页面活动状态，再关闭旧页。
    log("  正在恢复详情页活动状态...")
    _restore_page_activity(context, detail_page, log)

    try:
        log("  正在返回查询物料请求页面...")
        response = detail_page.goto(
            SEARCH_URL, wait_until="domcontentloaded", timeout=20000
        )
        if not _is_search_page_ready(detail_page, response):
            raise RuntimeError(f"详情页返回后不在查询页: {getattr(detail_page, 'url', '')}")
        log("  查询物料请求页面已就绪")
        # 导航完成会让 Chromium 重新计算标签页可见性；再次恢复活动状态，确保
        # 随后的非阻塞清理不会让新查询页退回后台，也不再依赖人工点击唤醒。
        log("  正在确认查询页活动状态...")
        _restore_page_activity(context, detail_page, log)
        # 替代页已经可用后，才关闭旧搜索页和点击请求 ID 时遗留的详情页。
        # 这样切换过程中始终至少保留一个可用页面，避免 Edge 进入后台状态。
        all_pages_closed = _close_other_pages(context, detail_page, log)
        if all_pages_closed:
            log("  已复用活跃详情页作为下一 PO 的搜索页")
        else:
            log("  ⚠ 当前查询页可继续使用，但仍有页面等待浏览器自行关闭")
        return detail_page
    except Exception as e:
        # 导航失败时才新建页面。旧搜索页在替代页真正可用前始终保留，
        # 避免二次失败时把最后一个可用页面也关闭。
        log(f"  ⚠ 详情页无法返回搜索页，改用新页: {str(e)[:100]}")
        new_search_page = None
        try:
            new_search_page = context.new_page()
            _restore_page_activity(context, new_search_page, log)
            log("  正在打开备用查询页面...")
            response = new_search_page.goto(
                SEARCH_URL, wait_until="domcontentloaded", timeout=20000
            )
            if not _is_search_page_ready(new_search_page, response):
                raise RuntimeError(f"备用页未进入查询页: {getattr(new_search_page, 'url', '')}")
            log("  备用查询页面已就绪")
        except Exception as nav_error:
            log(f"  ⚠ 新搜索页导航失败: {str(nav_error)[:100]}")
            if new_search_page is not None:
                try:
                    _release_page_activity_session(new_search_page)
                    if not new_search_page.is_closed():
                        new_search_page.close()
                except Exception:
                    pass
            retained_search_page = _find_open_search_page(context, search_page)
            if retained_search_page is not None:
                _restore_page_activity(context, retained_search_page, log)
                log("  ⚠ 已保留原查询页，等待下一 PO 继续")
                return retained_search_page
            log("  ⚠ 没有可验证的查询页，保留当前页面引用以便后续恢复")
            return search_page

        all_pages_closed = _close_other_pages(context, new_search_page, log)
        if all_pages_closed:
            log("  已启用备用查询页面作为下一 PO 的搜索页")
        else:
            log("  ⚠ 备用查询页可继续使用，但仍有页面等待浏览器自行关闭")
        return new_search_page


def _process_one_po(po, context, search_page, log):
    """
    处理单个 PO 的完整浏览器流程，并返回下一个 PO 使用的搜索页。
    搜索 → 点击请求ID → 提取项目名称/数量 → 下载附件。
    返回 dict 或 None。
    """
    result = {"po": po}

    try:
        # ── 0. 确保搜索页获得焦点与活动状态 ──
        _restore_page_activity(context, search_page, log)
        time.sleep(0.5)

        # ── 1. 获取搜索表单（优先复用当前页，只有首次才导航）──
        po_input = _locate_po_input(search_page, log)
        if po_input is None:
            search_page.goto(SEARCH_URL, wait_until="domcontentloaded", timeout=20000)
            time.sleep(3)
            po_input = _locate_po_input(search_page, log)

        if po_input is None:
            log(f"  ⚠ 无法定位 PO 输入框，跳过")
            return (None, search_page)

        # ── 2. 激活页面 + 填入 PO 号 ──
        # 点击 PO 标签（激活页面 + 聚焦输入框，比 body.click 精准）
        try:
            search_page.locator("//label[contains(text(),'PO号')]").first.click(timeout=2000)
        except Exception:
            po_input.click(force=True)
        time.sleep(0.3)

        # 三击全选 + 删除（替代 fill("")，兼容 Vue v-model）
        po_input.click(click_count=3, force=True)
        time.sleep(0.1)
        search_page.keyboard.press("Delete")
        time.sleep(0.1)
        search_page.keyboard.type(po, delay=50)
        time.sleep(0.3)

        # ── 3. 触发搜索 ──
        search_btn = _locate_search_button(search_page, log)
        if search_btn is None:
            log("  未找到「查找」按钮，按 Enter 提交...")
            search_page.keyboard.press("Enter")
        else:
            search_btn.click(force=True)
        log("  已触发搜索...")

        if not _wait_for_search_results(search_page, log):
            log(f"  ⚠ 搜索结果未出现，跳过")
            return (None, search_page)

        # ── 4. 点击请求ID → 详情页 ──
        rid_text, detail_page = _click_request_id(search_page, context, log)
        result["request_id"] = rid_text

        new_search_page = search_page  # 默认兜底：失败时保留原搜索页引用

        if detail_page is None:
            log(f"  ✗ 未能进入详情页：未执行附件下载，本 PO 不计为完成")
            new_search_page = _prepare_next_search_page(
                context, search_page, None, log
            )
            return (None, new_search_page)
        else:
            log(f"  详情页已打开，等待加载...")
            try:
                detail_page.wait_for_selector(
                    "//th[contains(text(),'支持文件')] | //td[contains(text(),'支持文件')] | //*[contains(text(),'项目经理')]",
                    timeout=15000
                )
            except Exception:
                pass
            try:
                detail_page.wait_for_load_state("domcontentloaded", timeout=5000)
            except Exception:
                pass
            try:
                detail_page.wait_for_load_state("networkidle", timeout=30000)
            except Exception:
                pass

            # WBS 是所有 PO 都需要回填的独立字段，不受 FLD 附件条件限制。
            result["wbs"] = _extract_wbs(detail_page, log)

            # ── 5. 下载全部附件（按 PO 号分文件夹，FLD/非FLD 分类）──
            try:
                dl_result = _download_support_files(detail_page, log, po)
                result["fld_files"] = dl_result.get("fld_files", [])
                result["non_fld_files"] = dl_result.get("non_fld_files", [])
                log(f"  附件下载: FLD {len(result['fld_files'])} 个, 非FLD {len(result['non_fld_files'])} 个")

                has_fld = len(result["fld_files"]) > 0

                # ── 6. 有 FLD 附件时：提取项目名称和梯台明细数量（用于 Excel 写回）──
                if has_fld:
                    result["project_name"] = _extract_project_name(detail_page, log)
                    # 兜底：网页提取失败时从 FLD 文件名提取项目名称
                    if not result["project_name"] and result["fld_files"]:
                        result["project_name"] = _extract_project_from_fld_name(
                            result["fld_files"][0], log
                        )
                    result["order_qty"] = _extract_order_qty(detail_page, log)
                else:
                    log(f"  ⓘ PO {po} 无 FLD 附件，跳过项目名称/数量提取")
                    result["project_name"] = None
                    result["order_qty"] = None
            finally:
                # ── 7. 将活跃详情页轮换为下个 PO 的搜索页 ──
                new_search_page = _prepare_next_search_page(
                    context, search_page, detail_page, log
                )

        log(f"  ✓ PO {po} 网页阶段完成")
        return (result, new_search_page)

    except Exception as e:
        log(f"  ✗ PO {po} 异常: {e}")
        return (None, search_page)


def _finalize_po_result(result, log, target_sheet=None):
    """
    本地处理：PDF解析 + 差异计算 + Excel写回。
    不依赖浏览器，避免 PDF 解析阻塞后续 PO 查询。
    """
    po = result.get("po")

    # ── FLD .msg 解析 → 提取审批价格 ──
    total_amount = None
    fld_files = result.get("fld_files", [])
    if fld_files:
        log("  解析 FLD 附件...")
        fld_target = os.path.join(FLD_DIR, str(po))

        # 找第一个 FLD 邮件附件（.msg 优先），根据扩展名选择解析器
        fld_mail = None
        is_eml = False
        for fname in fld_files:
            if fname.lower().endswith(".msg"):
                fld_mail = fname
                break
            if fname.lower().endswith(".eml"):
                fld_mail = fname
                is_eml = True
                break

        if fld_mail:
            fpath = os.path.join(fld_target, fld_mail)
            if os.path.exists(fpath):
                if is_eml:
                    total_amount = _extract_approval_price_from_eml(fpath, log)
                else:
                    total_amount = _extract_approval_price_from_msg(fpath, log)
            else:
                log(f"  FLD 邮件文件不存在: {fld_mail}")
        else:
            log("  FLD 附件中未找到 .msg 或 .eml 文件")

        if total_amount:
            log(f"  审批价格: {total_amount}")
        else:
            log("  未从 FLD 邮件中解析到审批价格")
    else:
        log("  无 FLD 附件，跳过审批价格解析")
    result["total_amount"] = total_amount

    # ── 读取订单净值 → 计算差异 ──
    order_value = _read_order_value_from_pivot(po, target_sheet=target_sheet)
    result["order_value"] = order_value
    order_diff = None

    if total_amount is not None and order_value is not None:
        try:
            ta = float(str(total_amount).replace(",", ""))
            ov = float(str(order_value).replace(",", ""))
            order_diff = round(ta - ov, 2)
            log(f"  订单差异: {order_diff} (={ta} - {ov})")
        except (ValueError, TypeError):
            log("  ⚠ 差异计算失败")
    result["order_diff"] = order_diff

    # ── 统计非 PDF 附件 ──
    non_pdf = [f for f in fld_files if not f.lower().endswith(".pdf")]
    result["non_pdf_count"] = len(non_pdf)

    # ── 写回 Excel ──
    support_files_str = "; ".join(fld_files) if fld_files else ""
    diff_str = str(order_diff) if order_diff is not None else None
    total_str = str(total_amount) if total_amount is not None else None

    excel_msg = _write_to_pivot_excel(
        po,
        project_name=result.get("project_name"),
        order_qty=result.get("order_qty"),
        support_files=support_files_str,
        total_amount=total_str,
        order_diff=diff_str,
        wbs=result.get("wbs"),
        target_sheet=target_sheet,
    )
    log(f"  {excel_msg}")
    log(f"  ✓ PO {po} 本地处理完成")
    return result


# ═══════════════════ 主入口 ═══════════════════

def query_and_download_attachments(
    target_sheet=None, category_label=None, status_callback=None, stop_requested=None,
    po_numbers=None, resume_manifest=None, is_recovery_run=False,
):
    """
    网页查询与下载阶段：
      1. 从透视表提取全部 PO 号
      2. 启动 Edge → 登录 → 循环处理每个 PO（搜索/提取/下载）
      3. 保存待回填任务，供独立回填阶段使用

    target_sheet: 目标透视表 Sheet 名称（默认 "装潢透视表"）
    category_label: 品类标签（"装潢"/"空调"），用于分文件夹存储下载附件
    stop_requested: 可选 threading.Event 或返回 bool 的函数。停止会在当前 PO
        完成后的安全检查点生效，并保留已下载文件和待回填任务。
    po_numbers: 仅处理指定 PO；供“补查缺失 PO”使用，永不扩大为全量。
    resume_manifest: 补查前的待回填任务。补查成功结果会按 PO 覆盖写回该任务。
    """
    ts = target_sheet or TARGET_SHEET
    category = _category_label(category_label)
    _configure_category_download_paths(category)

    if _is_stop_requested(stop_requested):
        return True, "⏹ 查询尚未开始，已按用户请求停止。"

    with open(RUN_LOG_FILE, "w", encoding="utf-8") as f:
        f.write(f"自动化运行日志 {time.strftime('%Y-%m-%d %H:%M:%S')}\n")

    def log(msg):
        if status_callback:
            status_callback(msg)
        try:
            with open(RUN_LOG_FILE, "a", encoding="utf-8") as f:
                f.write(f"{time.strftime('%H:%M:%S')} {msg}\n")
        except Exception:
            pass

    # ── 1. 提取全部 PO ──
    log("正在从透视表提取全部采购凭证号...")
    all_pos = _normalise_po_list(extract_all_pos_from_pivot(sheet_name=ts))
    if not all_pos:
        return False, "透视表中未找到任何采购凭证号。\n请确认透视表已生成。"
    log(f"共提取到 {len(all_pos)} 个采购凭证号")

    if po_numbers is not None:
        requested_pos = set(_normalise_po_list(po_numbers))
        test_pos = [po for po in all_pos if po in requested_pos]
        if not test_pos:
            return False, "待补查 PO 已不在当前透视表中，未启动浏览器。"
        audit_expected_pos = all_pos if is_recovery_run else test_pos
        log(f"[补查模式] 仅处理 {len(test_pos)} 个缺失 PO：{'、'.join(test_pos)}")
    elif TEST_MODE:
        test_pos = all_pos[:TEST_LIMIT]
        audit_expected_pos = test_pos
        log(f"[测试模式] 仅处理前 {len(test_pos)} 个 PO")
    else:
        test_pos = all_pos
        audit_expected_pos = all_pos
        log(f"[正式模式] 处理全部 {len(test_pos)} 个 PO")

    # ── 2. 启动浏览器 ──
    log("正在启动 Edge 浏览器...")
    from playwright.sync_api import sync_playwright
    playwright = sync_playwright().start()

    try:
        browser = playwright.chromium.launch(
            channel="msedge",
            headless=False,
            args=[
                "--start-maximized",
                "--disable-features=TranslateUI",           # 关闭翻译弹窗
                *EDGE_BACKGROUND_PROTECTION_ARGS,
            ],
        )

        context_kwargs = {"no_viewport": True, "accept_downloads": True}
        if os.path.exists(AUTH_FILE):
            log("检测到已保存的登录状态，尝试自动登录...")
            try:
                context = browser.new_context(**{**context_kwargs, "storage_state": AUTH_FILE})
                log("已加载登录状态。")
            except Exception:
                os.remove(AUTH_FILE)
                context = browser.new_context(**context_kwargs)
                log("登录状态文件已过期，需要重新登录。")
        else:
            context = browser.new_context(**context_kwargs)

        page = context.new_page()

        # ── 3. 登录检测 ──
        log("正在打开 TKE VIEW...")
        page.goto(HOME_URL, wait_until="domcontentloaded", timeout=30000)
        time.sleep(2)

        if _is_on_login_page(page):
            log("════════════════════════════════")
            log("⚠ 需要登录 — 请在浏览器中完成 Microsoft 账户验证")
            log("  登录完成后程序将自动继续...")
            log("════════════════════════════════")
            try:
                page.wait_for_url(
                    lambda url: "login.php" not in url.lower()
                    and "login.microsoftonline" not in url.lower(),
                    timeout=300000
                )
                log("✅ 登录成功！正在保存登录状态...")
                context.storage_state(path=AUTH_FILE)
                log("登录状态已保存。")
                time.sleep(2)
            except Exception:
                return False, "登录超时（超过 5 分钟），请重试。"
        else:
            log("✅ 已登录，会话有效。")

        try:
            page.close()
        except Exception:
            pass

        # 创建首个搜索页；每个详情页完成后会轮换为下一 PO 的搜索页。
        search_page = context.new_page()
        search_page.set_default_timeout(30000)
        log("搜索页已就绪")

        # ── 4. 循环处理本轮 PO（全量或补查清单）──
        total = len(test_pos)

        success_count = 0
        fail_count = 0
        failed_pos = []
        web_results = []
        crashed = False
        stopped = False

        for i, po in enumerate(test_pos):
            if _is_stop_requested(stop_requested):
                stopped = True
                log("⏹ 已收到停止请求：当前已完成的文件将保留，后续 PO 不再处理。")
                break
            log("")
            log(f"━━━ [{i+1}/{total}] PO: {po} ━━━")
            try:
                result, search_page = _process_one_po(po, context, search_page, log)
                if result is not None:
                    success_count += 1
                    web_results.append(result)
                    log(f"✓ [{i+1}/{total}] PO {po} 网页阶段完成")
                else:
                    fail_count += 1
                    failed_pos.append(po)
                    log(f"⚠ [{i+1}/{total}] PO {po} 未完成")
            except Exception as e:
                log(f"  ✗ 浏览器异常: {e}")
                fail_count += 1
                failed_pos.append(po)
                error_str = str(e).lower()
                if any(kw in error_str for kw in ("epipe", "closed", "target closed", "connection")):
                    log("⚠ 浏览器连接断开，停止后续处理。")
                    crashed = True
                    break

            time.sleep(0.5)

        # 正常完成时关闭搜索页；用户停止时保留浏览器，便于查看当前页面。
        if not stopped:
            try:
                _release_page_activity_session(search_page)
                search_page.close()
            except Exception:
                pass

        # 保留浏览器供查看
        if stopped:
            log("⏹ 网页查询已按用户请求停止，浏览器窗口保留供查看。")
        elif not crashed:
            log("网页阶段完成，浏览器窗口保留供查看。")

        # ── 5. 完整性审计 + 保存待回填任务 ──
        combined_results = merge_query_results(
            (resume_manifest or {}).get("results", []),
            web_results,
            audit_expected_pos,
            replaced_pos=test_pos if is_recovery_run else None,
        )
        audit = audit_download_completeness(
            audit_expected_pos, category, failed_pos=failed_pos
        )
        missing_pos = audit["missing_pos"]
        if missing_pos:
            log(
                f"⚠ 完整性检查：应有 {len(audit['expected_pos'])} 个 PO，"
                f"已有目录 {len(audit['folder_pos'])} 个，缺失 {len(missing_pos)} 个"
            )
            log("⚠ 待补查 PO：" + "、".join(missing_pos))
        else:
            log(f"✅ 完整性检查通过：{len(audit['expected_pos'])} 个 PO 均已有目录")

        manifest = {
            "version": 2,
            "created_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "category_label": category,
            "target_sheet": ts,
            "total_pos": len(audit_expected_pos),
            "attempted_pos": test_pos,
            "success_count": len(audit_expected_pos) - len(missing_pos),
            "failed_pos": missing_pos,
            "results": combined_results,
        }
        try:
            manifest_path = _save_pending_backfill_manifest(manifest)
            log(f"待回填任务已保存: {manifest_path}")
        except Exception as e:
            return False, f"网页查询完成，但保存待回填任务失败：\n{e}"

        # ── 6. 汇总 ──
        log("")
        if crashed:
            log(f"⚠ 浏览器异常中断！本轮完成: {success_count} / 失败: {fail_count}")
        elif stopped:
            log(f"⏹ 用户停止：本轮已完成 {success_count} / 失败 {fail_count} / 总数 {total}")
        else:
            prefix = "补查" if is_recovery_run else "网页阶段"
            log(f"✅ {prefix}完成：本轮成功 {success_count} / 失败 {fail_count} / 总数 {total}")

        auth_info = "登录状态已保存，下次自动登录。" if os.path.exists(AUTH_FILE) else ""
        if stopped:
            summary_prefix = "⏹ 查询已按请求停止"
        elif is_recovery_run:
            summary_prefix = "↻ 缺失 PO 补查完成！"
        else:
            summary_prefix = "✅ 查询与附件下载阶段完成！"
        summary = (
            f"{summary_prefix}\n\n"
            f"网页查询成功: {success_count} / 失败: {fail_count} / 总数: {total}\n"
            f"待回填 PO: {len(combined_results)}\n\n"
            "请点击「④ 回填已下载文件」执行附件解析、差异计算与 Excel 写回。"
        )
        no_fld_download_pos = [r.get("po") for r in combined_results if not r.get("fld_files")]
        if no_fld_download_pos:
            summary += f"\n\nℹ {len(no_fld_download_pos)} 个PO无FLD附件（仅下载非FLD文件至「无FLD文件」目录）："
            for p in no_fld_download_pos:
                summary += f"\n  - {p}"
        if missing_pos:
            summary += f"\n\n⚠ 完整性检查发现缺失 PO ({len(missing_pos)} 个):\n"
            for fp in missing_pos[:20]:
                summary += f"  - {fp}\n"
            if len(missing_pos) > 20:
                summary += f"  ... 等共 {len(missing_pos)} 个\n"
            summary += "请点击「↻ 补查缺失 PO」仅重新查询这些 PO。"
        summary += f"\n\n{auth_info}"
        return not crashed, summary

    except Exception as e:
        return False, f"网页操作异常：\n{e}\n\n{traceback.format_exc()}"


def backfill_downloaded_results(target_sheet=None, category_label=None, status_callback=None):
    """本地阶段：解析已下载附件、计算差异并写回当前品类的 Excel 透视表。"""
    ts = target_sheet or TARGET_SHEET
    category = _category_label(category_label)
    _configure_category_download_paths(category)
    manifest, error = _load_pending_backfill_manifest(category, ts)
    if error:
        return False, error

    with open(BACKFILL_RUN_LOG_FILE, "w", encoding="utf-8") as f:
        f.write(f"回填运行日志 {time.strftime('%Y-%m-%d %H:%M:%S')}\n")

    def log(msg):
        if status_callback:
            status_callback(msg)
        try:
            with open(BACKFILL_RUN_LOG_FILE, "a", encoding="utf-8") as f:
                f.write(f"{time.strftime('%H:%M:%S')} {msg}\n")
        except Exception:
            pass

    results = manifest["results"]
    finalize_success = 0
    skipped_without_fld = 0
    wbs_backfilled = 0
    failures = []

    log(f"开始回填「{category}」：共 {len(results)} 个网页结果")
    for index, result in enumerate(results, 1):
        po = result.get("po", "?")
        log("")
        log(f"━━━ [回填 {index}/{len(results)}] PO: {po} ━━━")
        if not result.get("fld_files"):
            skipped_without_fld += 1
            wbs = result.get("wbs")
            if wbs:
                excel_msg = _write_to_pivot_excel(
                    po, wbs=wbs, target_sheet=ts
                )
                wbs_backfilled += 1
                log(f"  无 FLD 附件，跳过附件解析；{excel_msg}")
            else:
                log("  ⏭ 无 FLD 附件，跳过附件解析；未取得 WBS，不写入空值")
            continue
        try:
            _finalize_po_result(result, log, target_sheet=ts)
            finalize_success += 1
            if result.get("wbs"):
                wbs_backfilled += 1
        except Exception as e:
            failures.append((po, str(e)))
            log(f"  ✗ 本地处理异常: {e}")

    if failures:
        log(f"⚠ 回填未完全成功，保留待回填任务以便重试：失败 {len(failures)} 个")
        failed_pos = "、".join(str(po) for po, _ in failures[:20])
        return False, (
            f"⚠ 回填未完成：成功 {finalize_success}，失败 {len(failures)}，"
            f"无 FLD 跳过 {skipped_without_fld}。\n"
            f"失败 PO：{failed_pos}\n"
            "待回填任务已保留，修复附件或 Excel 后可再次点击回填。"
        )

    try:
        archived_path = _archive_pending_backfill_manifest(category)
    except Exception as e:
        return False, f"Excel 回填完成，但归档待回填任务失败：\n{e}"

    log(
        f"✅ 回填完成：FLD 完整回填 {finalize_success}，"
        f"WBS 回填 {wbs_backfilled}，无 FLD 跳过附件解析 {skipped_without_fld}"
    )
    log(f"任务已归档: {archived_path}")
    return True, (
        f"✅ 回填完成！\n\n"
        f"回填成功: {finalize_success}\n"
        f"WBS 回填: {wbs_backfilled}\n"
        f"无 FLD 跳过: {skipped_without_fld}\n"
        f"任务归档: {archived_path}"
    )


def open_website_and_search(target_sheet=None, category_label=None, status_callback=None):
    """兼容旧调用：依次执行查询下载与 Excel 回填。"""
    query_success, query_message = query_and_download_attachments(
        target_sheet=target_sheet,
        category_label=category_label,
        status_callback=status_callback,
    )
    if not query_success:
        return False, query_message
    backfill_success, backfill_message = backfill_downloaded_results(
        target_sheet=target_sheet,
        category_label=category_label,
        status_callback=status_callback,
    )
    return backfill_success, f"{query_message}\n\n{backfill_message}"
